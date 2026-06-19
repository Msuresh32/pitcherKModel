"""
Daily picks Excel — styled to match the picks dashboard.
Columns: #, Pitcher, Bet, Book, Odds, Projection, Gap, Hit%, Edge, Kelly, Units,
         Exchange Order, Fair Odds, Make Order, Worst Play, Expected, Confidence, Game, Start

Make Order logic: minimum odds at which EV >= 12%.
Since user averages 10-15 cents CLV, the greedy initial ask is current_odds + 12 cents.

Sort: composite score = edge_pct + 2 * abs(gap), descending (strongest plays first).
Units: 2.5U = edge >= 35% | 2U = edge >= 18% AND |gap| >= 0.5 | 1U = everything else.
"""
import sys, io, argparse
from datetime import datetime, timezone, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, ".")

MIN_EDGE_PCT  = 15.0
MIN_EV_FLOOR  = 0.12       # Make Order floor: accept if EV >= 12%
CLV_CENTS     = 12         # greedy ask: request this many cents better than current
ET_OFFSET     = timedelta(hours=-4)   # EDT (UTC-4)

# ── Colours ───────────────────────────────────────────────────────────────────
C_HEADER   = "0D0D0D"
C_OVER     = "D6F5DD"
C_UNDER    = "FFE8E6"
C_GREEN_TXT= "1A6B2E"
C_RED_TXT  = "8B1A1A"
C_NAVY     = "1A2744"
C_FAIR_BG  = "FDECEA"
C_MO_POS   = "E8F5E9"
C_MO_NEG   = "FDECEA"
C_GREY_BG  = "F7F7F7"
C_ALT      = "FCFCFC"
C_UNIT_25_BG  = "FFF3CD"   # gold background for 2.5U
C_UNIT_25_TXT = "7D5200"   # dark amber text
C_UNIT_2_BG   = "D4EDDA"   # light green for 2U
C_UNIT_2_TXT  = "1A6B2E"   # dark green text
C_UNIT_1_TXT  = "666666"   # gray text for 1U

def tb(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def cell_style(cell, value=None, fmt=None, bold=False, size=9,
               color="000000", bg=None, align="center", wrap=False, border=True):
    if value is not None:
        cell.value = value
    if fmt:
        cell.number_format = fmt
    cell.font = Font(name="Inter", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = fill(bg)
    if border:
        cell.border = tb()


# ── Odds helpers ──────────────────────────────────────────────────────────────
def dec_to_american(decimal):
    if pd.isna(decimal) or decimal <= 1:
        return np.nan
    if decimal >= 2.0:
        return round((decimal - 1) * 100)
    else:
        return round(-100 / (decimal - 1))

def american_to_dec(odds):
    o = float(odds)
    return 1 + o/100 if o > 0 else 1 + 100/abs(o)

def fmt_american(odds):
    if pd.isna(odds): return ""
    o = int(round(float(odds)))
    return f"+{o}" if o > 0 else str(o)

def prob_to_american(p):
    if pd.isna(p) or p <= 0 or p >= 1: return np.nan
    if p >= 0.5:
        return round(-100 * p / (1 - p))
    else:
        return round(100 * (1 - p) / p)

def make_order_american(hit_prob, target_ev=MIN_EV_FLOOR):
    """Minimum odds at which EV >= target_ev."""
    if pd.isna(hit_prob) or hit_prob <= 0 or hit_prob >= 1:
        return np.nan
    decimal_floor = (1 + target_ev) / hit_prob
    return dec_to_american(decimal_floor)

def greedy_ask(current_odds, cents=CLV_CENTS):
    """Current odds + CLV_CENTS (the initial limit to post)."""
    if pd.isna(current_odds): return np.nan
    o = float(current_odds)
    asked = o + cents   # positive odds → +cents; negative → less negative (better for bettor)
    return int(round(asked))

def compute_units(edge_pct: float, gap_abs: float) -> str:
    """edge_pct: 0-100 scale. gap_abs: absolute value of (projection - line)."""
    if edge_pct >= 35:
        return "2.5U"
    if edge_pct >= 18 and gap_abs >= 0.5:
        return "2U"
    return "1U"

def composite_score(edge_pct: float, gap_abs: float) -> float:
    """Primary sort key — edge dominates, gap breaks ties."""
    return edge_pct + 2.0 * gap_abs

def exchange_order(pitcher, best_side, line):
    threshold = int(line) + 1   # e.g. 4.5 → 5, 5.5 → 6
    yn = "YES" if best_side == "over" else "NO"
    return f"{pitcher} {threshold}+ Ks: {yn}"

def parse_game_time(utc_str):
    try:
        dt = datetime.fromisoformat(utc_str.replace("Z", "+00:00"))
        et  = dt + ET_OFFSET
        return et.strftime("%-I:%M %p").replace("AM","AM").replace("PM","PM")
    except Exception:
        try:
            # Windows doesn't support %-I
            dt = datetime.fromisoformat(utc_str.replace("Z", "+00:00"))
            et  = dt + ET_OFFSET
            return et.strftime("%I:%M %p").lstrip("0")
        except Exception:
            return ""

def game_label(row):
    away = row.get("away_team", "")
    home = row.get("home_team", "")
    if away and home:
        return f"{away} @ {home}"
    return row.get("opponent", "")


# ── Load data ─────────────────────────────────────────────────────────────────
def load_picks(date_str):
    path = Path(f"data/exports/daily_pitcher_props_{date_str}.csv")
    if not path.exists():
        raise FileNotFoundError(f"No picks file for {date_str}: {path}")
    df = pd.read_csv(path)
    df = df[df["market"] == "strikeouts"].copy()
    df = df[df["edge_pct"] >= MIN_EDGE_PCT].copy()
    df["_gap_abs"] = (df["projection"] - df["line"]).abs()
    df["_score"]   = df.apply(lambda r: composite_score(r["edge_pct"], r["_gap_abs"]), axis=1)
    df = df.sort_values("_score", ascending=False).reset_index(drop=True)
    return df

def build_picks(df):
    rows = []
    for _, r in df.iterrows():
        side        = r["best_side"]
        line        = float(r["line"])
        proj        = float(r.get("projection", np.nan))
        over_prob   = float(r.get("over_probability", np.nan))
        hit_prob    = over_prob if side == "over" else (1 - over_prob)
        entry_odds  = r["over_odds"] if side == "over" else r["under_odds"]
        edge        = float(r.get("edge_pct", np.nan)) / 100
        kelly       = float(r.get("kelly_fraction", np.nan))
        fair_odds   = float(r.get("fair_odds", prob_to_american(1 - hit_prob)))
        hit_pct_str = r.get("hit_probability_pct", "")
        try:
            hit_pct = float(str(hit_pct_str).replace("%","")) / 100
        except Exception:
            hit_pct = hit_prob

        mo    = make_order_american(hit_prob)
        greed = greedy_ask(entry_odds)
        fair  = prob_to_american(1 - hit_prob) if side == "over" else prob_to_american(hit_prob)

        worst = (f"{fmt_american(mo)} max for {int(MIN_EV_FLOOR*100)}% edge; "
                 f"{fmt_american(fair)} breakeven")

        gap_val = round(proj - line, 2)
        units   = compute_units(float(r.get("edge_pct", 0)), abs(gap_val))

        rows.append({
            "Pitcher":        r["pitcher_name"],
            "Bet":            f"{'Over' if side=='over' else 'Under'} {line:.1f} Ks",
            "Book":           r.get("bookmaker_title", r.get("bookmaker", "")),
            "side":           side,
            "Odds":           entry_odds,
            "Projection":     round(proj, 2),
            "Gap":            gap_val,
            "Hit%":           hit_pct,
            "Edge":           edge,
            "Kelly":          kelly,
            "Units":          units,
            "Exchange Order": exchange_order(r["pitcher_name"], side, line),
            "Fair Odds":      fair,
            "Make Order":     mo,
            "Greedy Ask":     greed,
            "Worst Play":     worst,
            "Expected":       round(proj, 2),
            "Confidence":     r.get("confidence_tier", "Medium"),
            "Game":           game_label(r),
            "Start":          parse_game_time(str(r.get("commence_time",""))),
        })
    return pd.DataFrame(rows)


# ── Build Excel ───────────────────────────────────────────────────────────────
def build_excel(picks: pd.DataFrame, date_str: str, out_path: Path):
    wb = Workbook()
    del wb["Sheet"]
    ws = wb.create_sheet("Picks")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    COL_W = [
        ("",         3.5),
        ("Pitcher",  18),
        ("Bet",      12),
        ("Book",     12),
        ("Odds",      7),
        ("Proj",      7),
        ("Gap",       6),
        ("Hit %",     7),
        ("Edge",      7),
        ("Kelly",     7),
        ("Units",     6),
        ("Exchange Order", 24),
        ("Fair Odds",      9),
        ("Make Order",    10),
        ("Worst Play",    36),
        ("Expected",       8),
        ("Confidence",    11),
        ("Game",          30),
        ("Start",          9),
    ]
    N = len(COL_W)

    # ── Title row ──────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(N)}1")
    t = ws["A1"]
    t.value = f"Pitcher Strikeout Picks — {date_str}   |   Edge ≥ {int(MIN_EDGE_PCT)}%   |   {len(picks)} Plays"
    t.font      = Font(name="Inter", bold=True, size=12, color="FFFFFF")
    t.fill      = fill(C_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Header row ─────────────────────────────────────────────────────────────
    for c, (label, width) in enumerate(COL_W, 1):
        cell = ws.cell(row=2, column=c, value=label)
        cell.font      = Font(name="Inter", bold=True, size=9, color="FFFFFF")
        cell.fill      = fill(C_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = tb("444466")
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[2].height = 18

    # ── Data rows ──────────────────────────────────────────────────────────────
    for i, (_, row) in enumerate(picks.iterrows(), 1):
        r       = i + 2
        side    = row["side"]
        row_bg  = C_OVER if side == "over" else C_UNDER

        def c(col, val, fmt=None, bold=False, color="111111", bg=None, align="center"):
            cell = ws.cell(row=r, column=col, value=val)
            cell_style(cell, fmt=fmt, bold=bold, size=9,
                       color=color, bg=bg or row_bg, align=align)

        mo      = row["Make Order"]
        mo_bg   = C_MO_POS if (not pd.isna(mo) and mo >= 0) else C_MO_NEG
        mo_clr  = C_GREEN_TXT if (not pd.isna(mo) and mo >= 0) else C_RED_TXT
        fair    = row["Fair Odds"]
        entry   = row["Odds"]
        gap     = row["Gap"]

        units     = row["Units"]
        unit_bg   = C_UNIT_25_BG if units == "2.5U" else (C_UNIT_2_BG if units == "2U" else row_bg)
        unit_clr  = C_UNIT_25_TXT if units == "2.5U" else (C_UNIT_2_TXT if units == "2U" else C_UNIT_1_TXT)

        c(1,  i,                         fmt="#,##0",      color="999999")
        c(2,  row["Pitcher"],                               bold=True, align="left",  color=C_NAVY)
        c(3,  row["Bet"],                                   bold=True,
              color=C_GREEN_TXT if side=="over" else C_RED_TXT)
        c(4,  row["Book"],                                  color="333333")
        c(5,  fmt_american(entry),                          bold=True,
              color=C_GREEN_TXT if (not pd.isna(entry) and entry>=0) else C_RED_TXT)
        c(6,  row["Projection"],         fmt="0.00")
        c(7,  row["Gap"],                fmt="+0.00;-0.00",
              bold=True, color=C_GREEN_TXT if gap>=0 else C_RED_TXT)
        c(8,  row["Hit%"],               fmt="0.0%")
        c(9,  row["Edge"],               fmt="0.0%",       bold=True,
              color=C_GREEN_TXT if row["Edge"]>=0.20 else "333333")
        c(10, row["Kelly"],              fmt="0.0%")

        # Units cell — gold=2.5U, green=2U, gray=1U
        unit_cell = ws.cell(row=r, column=11, value=units)
        unit_cell.font      = Font(name="Inter", size=9, bold=True, color=unit_clr)
        unit_cell.fill      = fill(unit_bg)
        unit_cell.alignment = Alignment(horizontal="center", vertical="center")
        unit_cell.border    = tb()

        c(12, row["Exchange Order"],                        align="left", color="444444")
        c(13, fmt_american(fair),                           bg=C_FAIR_BG, bold=True, color=C_RED_TXT)

        # Make Order cell — show greedy ask / floor
        mo_cell = ws.cell(row=r, column=14)
        mo_text = fmt_american(mo)
        mo_cell.value     = f"{mo_text} or better"
        mo_cell.font      = Font(name="Inter", size=9, bold=True, color=mo_clr)
        mo_cell.fill      = fill(mo_bg)
        mo_cell.alignment = Alignment(horizontal="center", vertical="center")
        mo_cell.border    = tb()

        c(15, row["Worst Play"],                            align="left", color="555555")
        c(16, row["Expected"],           fmt="0.00")

        conf    = row["Confidence"]
        conf_clr = {"High": C_GREEN_TXT, "Medium": "8B6914", "Low": C_RED_TXT}.get(conf, "333333")
        c(17, conf,                                         bold=True, color=conf_clr)
        c(18, row["Game"],                                  align="left", color="333333")
        c(19, row["Start"],                                 color="333333")

        ws.row_dimensions[r].height = 16

    # ── Greedy ask note ────────────────────────────────────────────────────────
    note_row = len(picks) + 3
    ws.merge_cells(f"A{note_row}:{get_column_letter(N)}{note_row}")
    note = ws.cell(row=note_row, column=1,
                   value=f"Make Order = minimum odds for {int(MIN_EV_FLOOR*100)}% EV floor.  "
                         f"Initial ask = current odds +{CLV_CENTS} cents (greedy, based on avg CLV).  "
                         f"Accept anything better than Make Order.")
    note.font      = Font(name="Inter", size=8, italic=True, color="888888")
    note.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[note_row].height = 14

    # Auto-filter
    ws.auto_filter.ref = f"A2:{get_column_letter(N)}{len(picks)+2}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


# ── Main ──────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"))
args = parser.parse_args()

df   = load_picks(args.date)
picks = build_picks(df)
out  = Path(f"site/picks/todays_picks_{args.date}.xlsx")
build_excel(picks, args.date, out)

print(f"\n{len(picks)} picks for {args.date} (sorted by strength):")
for i, (_, r) in enumerate(picks.iterrows(), 1):
    print(f"  {i:>2}. [{r['Units']:<4}] {r['Pitcher']:<22} {r['Bet']:<14} "
          f"{fmt_american(r['Odds']):<7} edge={r['Edge']*100:.1f}%  gap={r['Gap']:+.2f}  "
          f"Make Order: {fmt_american(r['Make Order'])} or better")

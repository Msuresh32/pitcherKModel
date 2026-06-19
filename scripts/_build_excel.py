"""
Build full Excel report: all picks, running bankroll, P&L charts, summary.

Bet sizing: half-Kelly capped at 3% of current bankroll (realistic).
Starting bankroll: $10,000.
"""
import sys, io
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, ".")

# ── Config ────────────────────────────────────────────────────────────────────
STARTING_BANKROLL = 10_000
UNIT_SIZE         = 100        # 1U = $100 (1% of starting bankroll), fixed — does not compound
MAX_UNITS         = 3          # hard cap: 3U = $300 per bet
HALF_KELLY_FACTOR = 0.5
MIN_EDGE_PCT      = 15.0
OUTPUT_PATH       = "data/exports/pitcher_model_report.xlsx"

# ── Colours ───────────────────────────────────────────────────────────────────
CLR_HEADER_DARK  = "1F3864"   # dark navy
CLR_HEADER_MID   = "2F5496"   # mid blue
CLR_HEADER_LIGHT = "D6E4F7"   # pale blue
CLR_WIN          = "D9EAD3"   # light green
CLR_LOSS         = "FCE5CD"   # light orange/red
CLR_ACCENT       = "C9DAF8"   # accent blue
CLR_SUMMARY_BG   = "F3F3F3"
CLR_POSITIVE     = "274E13"   # dark green text
CLR_NEGATIVE     = "990000"   # dark red text
CLR_NEUTRAL      = "1F3864"   # navy text

# ── Helpers ───────────────────────────────────────────────────────────────────
def american_to_decimal(odds):
    if pd.isna(odds): return np.nan
    o = float(odds)
    return 1 + o / 100 if o > 0 else 1 + 100 / abs(o)

def implied_prob(odds):
    o = float(odds)
    return 100 / (100 + o) if o > 0 else abs(o) / (abs(o) + 100)

def devig_pair(over_o, under_o):
    if pd.isna(over_o) or pd.isna(under_o): return np.nan, np.nan
    ip_o, ip_u = implied_prob(over_o), implied_prob(under_o)
    d = ip_o + ip_u
    return (ip_o / d, ip_u / d) if d > 0 else (np.nan, np.nan)

def fmt_odds(o):
    if pd.isna(o): return ""
    return f"+{int(o)}" if o > 0 else str(int(o))

def mean_decimal_odds(american_odds):
    vals = [american_to_decimal(o) for o in american_odds if pd.notna(o)]
    vals = [v for v in vals if pd.notna(v)]
    return float(np.mean(vals)) if vals else np.nan

def clv_pct_from_prices(entry_decimal, close_american):
    if pd.isna(entry_decimal) or pd.isna(close_american):
        return np.nan
    close_decimal = american_to_decimal(close_american)
    if pd.isna(close_decimal) or close_decimal <= 1:
        return np.nan
    return (entry_decimal / close_decimal - 1) * 100

def decimal_to_american(dec):
    if pd.isna(dec) or dec <= 1: return np.nan
    return round((dec - 1) * 100) if dec >= 2.0 else round(-100 / (dec - 1))

def clv_pct_decimal(entry_dec, close_dec):
    """CLV in percentage points using decimal prices (positive = beat the close)."""
    if pd.isna(entry_dec) or pd.isna(close_dec) or close_dec <= 1: return np.nan
    return (entry_dec / close_dec - 1) * 100

def compute_units_bk(edge_pct: float, gap_abs: float) -> float:
    """Edge + gap based fixed unit sizing. Returns float (1.0, 2.0, 2.5)."""
    if edge_pct >= 35: return 2.5
    if edge_pct >= 18 and gap_abs >= 0.5: return 2.0
    return 1.0

def side_str(best_side, line):
    return f"O {line}" if best_side == "over" else f"U {line}"

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ── Load & prepare bets ───────────────────────────────────────────────────────
def load_bets():
    sources = [
        ("thresh_sel_2025_dk_edges.csv", "data/processed_2024"),
        ("wf2026_p1_mar_apr_edges.csv",  "data/processed"),
        ("wf2026_p2_may_edges.csv",       "data/processed_apr2026"),
        ("wf2026_p3_jun_edges.csv",       "data/processed"),
    ]
    dfs = []
    for fname, d in sources:
        p = Path(d) / fname
        if p.exists():
            df = pd.read_csv(p)
            df["_source"] = fname
            dfs.append(df[df["market"] == "strikeouts"].copy())
    if not dfs:
        raise FileNotFoundError("No bet files found.")
    bets = pd.concat(dfs, ignore_index=True)
    bets = (bets.sort_values("edge_pct", ascending=False)
                .drop_duplicates(subset=["game_date","pitcher_name","line","best_side"])
                .reset_index(drop=True))
    bets = bets[bets["edge_pct"] >= MIN_EDGE_PCT].copy()
    bets["game_date"] = pd.to_datetime(bets["game_date"])
    bets = bets.sort_values("game_date").reset_index(drop=True)

    bets["bet_odds"] = np.where(
        bets["best_side"] == "over", bets["over_odds"], bets["under_odds"])
    bets["decimal_odds"] = bets["bet_odds"].apply(american_to_decimal)

    nv = bets.apply(lambda r: pd.Series(devig_pair(r["over_odds"], r["under_odds"])), axis=1)
    bets["nv_over"]  = nv[0]
    bets["nv_under"] = nv[1]
    bets["nv_entry"] = np.where(bets["best_side"] == "over", bets["nv_over"], bets["nv_under"])

    bets["won"] = np.where(
        bets["best_side"] == "over",
        bets["strikeouts"] > bets["line"],
        bets["strikeouts"] < bets["line"],
    )
    return bets

def compute_bankroll(bets, starting=STARTING_BANKROLL):
    """
    Fixed-unit sizing: 1U = $100.
    2.5U = edge>=35% | 2U = edge>=18% + |gap|>=0.5 | 1U = everything else.
    Unit size is FIXED — does not compound with bankroll. Max bet = $300.
    """
    bets = bets.copy()
    bankroll = starting
    bet_sizes, units_list, profits, running_bks = [], [], [], []

    for _, row in bets.iterrows():
        edge_pct = float(row.get("edge_pct", 0) or 0)
        gap_abs  = abs(float(row.get("gap", 0) or 0))
        units    = compute_units_bk(edge_pct, gap_abs)
        bet_size = round(units * UNIT_SIZE, 2)

        dec = row["decimal_odds"]
        won = row["won"]
        profit = round(bet_size * (dec - 1), 2) if (won and not pd.isna(dec)) else -bet_size
        if pd.isna(won):
            profit = 0.0

        bankroll = round(bankroll + profit, 2)
        bet_sizes.append(bet_size)
        units_list.append(units)
        profits.append(profit)
        running_bks.append(bankroll)

    bets["units"]           = units_list
    bets["bet_size"]        = bet_sizes
    bets["pnl"]             = profits
    bets["running_bankroll"] = running_bks
    return bets

# ── Style helpers ─────────────────────────────────────────────────────────────
def write_header_row(ws, row, cols, bg=CLR_HEADER_DARK, fg="FFFFFF", bold=True, fontsize=10):
    for c, (text, width) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.fill = header_fill(bg)
        cell.font = Font(name="Calibri", bold=bold, color=fg, size=fontsize)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        if width:
            ws.column_dimensions[get_column_letter(c)].width = width

def style_data_cell(cell, fmt=None, bold=False, color=None, align="center", fill_color=None):
    cell.font = Font(name="Calibri", size=9, bold=bold,
                     color=color if color else "000000")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = thin_border()
    if fmt:
        cell.number_format = fmt
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)

# ── Sheet 1: All Picks ────────────────────────────────────────────────────────
BOOKS = [
    ("draftkings",  "DK"),
    ("fanduel",     "FD"),
    ("betonlineag", "BOL"),
    ("betrivers",   "BR"),
    ("pinnacle",    "PIN"),
]

def load_all_odds():
    """Load open + close odds for all books, keyed by (game_date, player_name, line, bookmaker)."""
    dfs = []
    for path in ["data/odds/historical_pitcher_props_2025.csv",
                 "data/odds/full_2026_odds.csv"]:
        if Path(path).exists():
            df = pd.read_csv(path, usecols=lambda c: c in [
                "game_date","player_name","bookmaker","line",
                "over_odds","under_odds","snapshot_type","fetched_at"])
            dfs.append(df)

    # Pinnacle close cache
    pin_path = Path("data/odds/pinnacle_close_cache.csv")
    if pin_path.exists():
        pin = pd.read_csv(pin_path)
        pin["bookmaker"] = "pinnacle"
        pin["snapshot_type"] = "close"
        dfs.append(pin[["game_date","player_name","bookmaker","line",
                         "over_odds","under_odds","snapshot_type"]])

    if not dfs:
        return pd.DataFrame()

    odds = pd.concat(dfs, ignore_index=True)
    odds["game_date"] = pd.to_datetime(odds["game_date"])
    odds["line"] = pd.to_numeric(odds["line"], errors="coerce")
    odds["over_odds"] = pd.to_numeric(odds["over_odds"], errors="coerce")
    odds["under_odds"] = pd.to_numeric(odds["under_odds"], errors="coerce")

    # Keep only rows with at least one side
    odds = odds[odds["over_odds"].notna() | odds["under_odds"].notna()].copy()

    # For open: take first valid row per (date, book, player, line)
    opens = (odds[odds["snapshot_type"] == "open"]
             .sort_values("fetched_at", na_position="last")
             .groupby(["game_date","bookmaker","player_name","line"],
                      as_index=False).first())

    # For close: take last row with both sides where possible
    close_both = odds[(odds["snapshot_type"] == "close") &
                      odds["over_odds"].notna() & odds["under_odds"].notna()]
    closes = (close_both
              .sort_values("fetched_at", na_position="last")
              .groupby(["game_date","bookmaker","player_name","line"],
                       as_index=False).last())

    # Fallback: any close row
    close_any = (odds[odds["snapshot_type"] == "close"]
                 .sort_values("fetched_at", na_position="last")
                 .groupby(["game_date","bookmaker","player_name","line"],
                           as_index=False).last())
    closes = pd.concat([closes, close_any]).drop_duplicates(
        subset=["game_date","bookmaker","player_name","line"], keep="first")

    return opens, closes


def get_odds_for_bet(row, opens, closes, book_key):
    mask_o = (
        (opens["game_date"] == row["game_date"]) &
        (opens["bookmaker"] == book_key) &
        (opens["player_name"] == row["pitcher_name"]) &
        (opens["line"] == row["line"])
    )
    mask_c = (
        (closes["game_date"] == row["game_date"]) &
        (closes["bookmaker"] == book_key) &
        (closes["player_name"] == row["pitcher_name"]) &
        (closes["line"] == row["line"])
    )
    side = row["best_side"]
    def pick(df, mask):
        sub = df[mask]
        if sub.empty: return np.nan
        o = sub.iloc[0]["over_odds"] if side == "over" else sub.iloc[0]["under_odds"]
        return float(o) if pd.notna(o) else np.nan

    return pick(opens, mask_o), pick(closes, mask_c)


def build_picks_sheet(wb, bets, opens=None, closes=None):
    ws = wb.create_sheet("All Picks")
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30

    # Column layout:
    # Core cols (1-17) + per-book open/close pairs (5 books × 2 = 10 cols)
    n_book_cols = len(BOOKS) * 2
    total_cols  = 18 + n_book_cols

    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    title = ws["A1"]
    title.value = "MLB Pitcher Strikeout Props Model — All Picks (Edge ≥ 15%)"
    title.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    title.fill = header_fill(CLR_HEADER_DARK)
    title.alignment = Alignment(horizontal="center", vertical="center")

    core_headers = [
        ("#",            4),
        ("Date",         11),
        ("Pitcher",      18),
        ("Team",         6),
        ("Opp",          6),
        ("Bet",          9),
        ("Entry Odds",   9),
        ("Projection",   10),
        ("Actual Ks",    9),
        ("Edge %",       8),
        ("Over Prob",    9),
        ("Units",        6),
        ("Bet ($)",      9),
        ("Won",          5),
        ("P&L ($)",      10),
        ("Running BK",   12),
        ("Year",         6),
        ("CLV %",        8),
    ]
    book_headers = []
    for _, short in BOOKS:
        book_headers.append((f"{short} Open", 9))
        book_headers.append((f"{short} Close", 9))

    all_headers = core_headers + book_headers

    # Two-row header: first row group labels, second row column names
    # Group label row
    ws.merge_cells(f"A2:R2")  # core
    grp = ws["A2"]
    grp.value = "Model Output"
    grp.font  = Font(bold=True, color="FFFFFF", size=9)
    grp.fill  = header_fill(CLR_HEADER_MID)
    grp.alignment = Alignment(horizontal="center", vertical="center")

    col_offset = 19
    for b_idx, (_, short) in enumerate(BOOKS):
        c1 = get_column_letter(col_offset + b_idx * 2)
        c2 = get_column_letter(col_offset + b_idx * 2 + 1)
        ws.merge_cells(f"{c1}2:{c2}2")
        cell = ws[f"{c1}2"]
        cell.value = short
        cell.font  = Font(bold=True, color="FFFFFF", size=9)
        cell.fill  = header_fill("3D6B9E" if b_idx % 2 == 0 else CLR_HEADER_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 16

    write_header_row(ws, 3, all_headers, bg=CLR_HEADER_MID, fontsize=9)
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"

    prev_bk = STARTING_BANKROLL
    for i, (_, row) in enumerate(bets.iterrows(), 1):
        r = i + 3
        won = row["won"]
        fill = CLR_WIN if won else CLR_LOSS
        bk_change = row["running_bankroll"] - prev_bk
        prev_bk = row["running_bankroll"]

        proj_val = row.get("strikeouts_projection", row.get("projection", np.nan))
        try:
            proj_val = round(float(proj_val), 2)
        except Exception:
            proj_val = np.nan

        book_prices = {}
        if opens is not None and closes is not None:
            for book_key, _ in BOOKS:
                book_prices[book_key] = get_odds_for_bet(row, opens, closes, book_key)

        avg_open_dec = mean_decimal_odds(
            open_odds for book_key, (open_odds, _) in book_prices.items()
            if book_key != "pinnacle"
        )
        pin_close_am = book_prices.get("pinnacle", (np.nan, np.nan))[1]
        if pd.notna(pin_close_am):
            close_dec = american_to_decimal(pin_close_am)
        else:
            # Fallback: average closing price across all API books
            close_dec = mean_decimal_odds(
                c for _, (_, c) in book_prices.items() if pd.notna(c)
            )
        clv_pct = clv_pct_decimal(avg_open_dec, close_dec)

        core_vals = [
            (i,                                          "#,##0",                False, None),
            (row["game_date"].strftime("%Y-%m-%d"),      "@",                    False, None),
            (row["pitcher_name"],                        "@",                    False, None),
            (row.get("team",""),                         "@",                    False, None),
            (row.get("opponent",""),                     "@",                    False, None),
            (side_str(row["best_side"], row["line"]),    "@",                    True,  CLR_NEUTRAL),
            (fmt_odds(row["bet_odds"]),                  "@",                    False, None),
            (proj_val,                                   "0.00",                 False, None),
            (row.get("strikeouts", np.nan),              "0.0",                  False, None),
            (row["edge_pct"] / 100,                      "0.0%",                 True,
             CLR_POSITIVE if row["edge_pct"] >= 20 else CLR_NEUTRAL),
            (row.get("over_probability", np.nan),        "0.0%",                 False, None),
            (row["units"],                               "0.0",                  True,  CLR_NEUTRAL),
            (row["bet_size"],                            "$#,##0",               False, None),
            ("W" if won else "L",                        "@",                    True,
             CLR_POSITIVE if won else CLR_NEGATIVE),
            (row["pnl"],                                 "+$#,##0.00;-$#,##0.00", True,
             CLR_POSITIVE if row["pnl"] >= 0 else CLR_NEGATIVE),
            (row["running_bankroll"],                    "$#,##0",               False, None),
            (row["game_date"].year,                      "0",                    False, None),
            (clv_pct / 100 if pd.notna(clv_pct) else "-",    "+0.0%;-0.0%",       True,
             CLR_POSITIVE if pd.notna(clv_pct) and clv_pct >= 0 else CLR_NEGATIVE),
        ]

        for c, (val, fmt, bold, color) in enumerate(core_vals, 1):
            v = "" if (isinstance(val, float) and np.isnan(val)) else val
            style_data_cell(ws.cell(row=r, column=c, value=v),
                            fmt=fmt, bold=bold, color=color, fill_color=fill)

        # Book odds columns
        for b_idx, (book_key, _) in enumerate(BOOKS):
            c_open  = 19 + b_idx * 2
            c_close = 20 + b_idx * 2
            o_odds, c_odds = book_prices.get(book_key, (np.nan, np.nan))

            for c_num, val in [(c_open, o_odds), (c_close, c_odds)]:
                safe = float(val) if val is not None else np.nan
                disp = fmt_odds(safe) if pd.notna(safe) else "-"
                cell = ws.cell(row=r, column=c_num, value=disp)
                style_data_cell(cell, fmt="@", fill_color=fill)

        ws.row_dimensions[r].height = 14

    # Auto-filter on header row 3
    ws.auto_filter.ref = f"A3:{get_column_letter(total_cols)}{len(bets)+3}"

    # Totals row
    tr = len(bets) + 4
    totals = {1: "TOTAL", 13: (bets["bet_size"].sum(), "$#,##0"),
              15: (bets["pnl"].sum(), "+$#,##0;-$#,##0"),
              16: (bets["running_bankroll"].iloc[-1], "$#,##0")}
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=tr, column=c)
        cell.fill   = header_fill(CLR_HEADER_LIGHT)
        cell.border = thin_border()
        if c in totals:
            v = totals[c]
            if isinstance(v, str):
                cell.value = v
                cell.font  = Font(bold=True, size=10)
            else:
                cell.value = v[0]
                cell.number_format = v[1]
                cell.font  = Font(bold=True, size=10,
                    color=CLR_POSITIVE if v[0] >= 0 else CLR_NEGATIVE)

    return ws


# ── Sheet 2: Bankroll Chart ───────────────────────────────────────────────────
def build_bankroll_sheet(wb, bets):
    ws = wb.create_sheet("Bankroll Chart")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Running Bankroll ($10,000 start)"
    ws["A1"].font = Font(bold=True, size=12, color=CLR_NEUTRAL)
    ws["A2"] = "Bet #"
    ws["B2"] = "Bankroll"
    ws["C2"] = "Date"

    for i, (_, row) in enumerate(bets.iterrows(), 1):
        ws.cell(row=i+2, column=1, value=i)
        ws.cell(row=i+2, column=2, value=row["running_bankroll"])
        ws.cell(row=i+2, column=3, value=row["game_date"].strftime("%Y-%m-%d"))

    # Starting point
    ws.cell(row=3, column=2).value  # already set

    n = len(bets)
    chart = LineChart()
    chart.title = "Running Bankroll"
    chart.style = 10
    chart.y_axis.title = "Bankroll ($)"
    chart.x_axis.title = "Bet Number"
    chart.y_axis.numFmt = '$#,##0'
    chart.height = 14
    chart.width  = 26

    from openpyxl.chart import Reference as ChartReference
    data = ChartReference(ws, min_col=2, min_row=2, max_row=n+2)
    chart.add_data(data, titles_from_data=True)
    chart.series[0].graphicalProperties.line.solidFill = "2F5496"
    chart.series[0].graphicalProperties.line.width = 20000

    # Starting $10k reference line
    ws["E2"] = "Starting BK"
    for i in range(1, n+2):
        ws.cell(row=i+2, column=5, value=STARTING_BANKROLL)
    ref_data = ChartReference(ws, min_col=5, min_row=2, max_row=n+2)
    chart.add_data(ref_data, titles_from_data=True)
    chart.series[1].graphicalProperties.line.solidFill = "CCCCCC"
    chart.series[1].graphicalProperties.line.dashDot = "dash"
    chart.series[1].graphicalProperties.line.width = 12000

    ws.add_chart(chart, "G2")
    return ws


# ── Sheet 3: Monthly Summary ──────────────────────────────────────────────────
def build_monthly_sheet(wb, bets):
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_view.showGridLines = False

    bets = bets.copy()
    bets["month"] = bets["game_date"].dt.to_period("M").astype(str)

    monthly = (bets.groupby("month")
               .agg(
                   bets_count  =("pnl",    "count"),
                   wins        =("won",    "sum"),
                   total_staked=("bet_size","sum"),
                   total_pnl   =("pnl",    "sum"),
                   avg_edge    =("edge_pct","mean"),
               )
               .reset_index())
    monthly["win_rate"]  = monthly["wins"] / monthly["bets_count"]
    monthly["roi"]       = monthly["total_pnl"] / monthly["total_staked"]
    monthly["cum_pnl"]   = monthly["total_pnl"].cumsum()

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"].value = "Monthly Performance Summary"
    ws["A1"].font  = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill  = header_fill(CLR_HEADER_DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    hdrs = [("Month",10),("Bets",7),("Wins",7),("Win Rate",10),
            ("Staked ($)",13),("P&L ($)",11),("ROI",8),("Cum P&L ($)",13)]
    write_header_row(ws, 2, hdrs, bg=CLR_HEADER_MID)

    for i, mrow in enumerate(monthly.itertuples(), 1):
        r = i + 2
        fill = CLR_WIN if mrow.total_pnl >= 0 else CLR_LOSS
        data = [
            (mrow.month,         "@",                    False, None),
            (mrow.bets_count,    "#,##0",                False, None),
            (int(mrow.wins),     "#,##0",                False, None),
            (mrow.win_rate,      "0.0%",                 False, None),
            (mrow.total_staked,  "$#,##0",               False, None),
            (mrow.total_pnl,     "+$#,##0;-$#,##0",      True,
             CLR_POSITIVE if mrow.total_pnl >= 0 else CLR_NEGATIVE),
            (mrow.roi,           "+0.0%;-0.0%",           True,
             CLR_POSITIVE if mrow.roi >= 0 else CLR_NEGATIVE),
            (mrow.cum_pnl,       "+$#,##0;-$#,##0",      False, None),
        ]
        for c, (val, fmt, bold, color) in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_data_cell(cell, fmt=fmt, bold=bold, color=color, fill_color=fill)
        ws.row_dimensions[r].height = 16

    # Monthly P&L bar chart
    n_months = len(monthly)
    data_col = 6   # P&L column
    chart = BarChart()
    chart.type   = "col"
    chart.title  = "Monthly P&L ($)"
    chart.y_axis.title = "P&L ($)"
    chart.style  = 10
    chart.height = 12
    chart.width  = 22
    chart.y_axis.numFmt = '$#,##0'

    data_ref = Reference(ws, min_col=data_col, min_row=2, max_row=n_months+2)
    cats_ref = Reference(ws, min_col=1,        min_row=3, max_row=n_months+2)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "2F5496"

    ws.add_chart(chart, "J2")
    return ws


# ── Sheet 4: Edge Band Analysis ───────────────────────────────────────────────
def build_edge_sheet(wb, bets):
    ws = wb.create_sheet("Edge Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Performance by Edge Band"
    ws["A1"].font  = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill  = header_fill(CLR_HEADER_DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    bands = [(15,18,"15-18%"),(18,21,"18-21%"),(21,25,"21-25%"),(25,30,"25-30%"),(30,999,"30%+")]
    hdrs = [("Edge Band",11),("Bets",7),("Win Rate",10),("Avg Odds",9),("Staked ($)",12),("P&L ($)",11),("ROI",8)]
    write_header_row(ws, 2, hdrs, bg=CLR_HEADER_MID)

    for i, (lo, hi, label) in enumerate(bands, 1):
        sub = bets[(bets["edge_pct"] >= lo) & (bets["edge_pct"] < hi)]
        if sub.empty:
            continue
        r = i + 2
        roi = sub["pnl"].sum() / sub["bet_size"].sum() if sub["bet_size"].sum() > 0 else 0
        data = [
            (label,                    "@",              False, None),
            (len(sub),                 "#,##0",          False, None),
            (sub["won"].mean(),        "0.0%",           False, None),
            (sub["bet_odds"].mean(),   "+0.0;-0.0",      False, None),
            (sub["bet_size"].sum(),    "$#,##0",         False, None),
            (sub["pnl"].sum(),         "+$#,##0;-$#,##0", True,
             CLR_POSITIVE if sub["pnl"].sum() >= 0 else CLR_NEGATIVE),
            (roi,                      "+0.0%;-0.0%",    True,
             CLR_POSITIVE if roi >= 0 else CLR_NEGATIVE),
        ]
        fill = CLR_WIN if sub["pnl"].sum() >= 0 else CLR_LOSS
        for c, (val, fmt, bold, color) in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_data_cell(cell, fmt=fmt, bold=bold, color=color, fill_color=fill)
        ws.row_dimensions[r].height = 16

    # Total row
    tr = len(bands) + 3
    sub = bets
    roi_all = sub["pnl"].sum() / sub["bet_size"].sum()
    totals = [
        ("TOTAL (≥15%)",     "@",              True, None),
        (len(sub),           "#,##0",          True, None),
        (sub["won"].mean(),  "0.0%",           True, None),
        (sub["bet_odds"].mean(), "+0.0;-0.0",  True, None),
        (sub["bet_size"].sum(), "$#,##0",      True, None),
        (sub["pnl"].sum(),   "+$#,##0;-$#,##0", True,
         CLR_POSITIVE if sub["pnl"].sum() >= 0 else CLR_NEGATIVE),
        (roi_all,            "+0.0%;-0.0%",    True,
         CLR_POSITIVE if roi_all >= 0 else CLR_NEGATIVE),
    ]
    for c, (val, fmt, bold, color) in enumerate(totals, 1):
        cell = ws.cell(row=tr, column=c, value=val)
        style_data_cell(cell, fmt=fmt, bold=bold, color=color, fill_color=CLR_HEADER_LIGHT)
    return ws


# ── Sheet 5: Summary Dashboard ────────────────────────────────────────────────
def build_summary_sheet(wb, bets):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Move to front after creation (done outside)
    ws.merge_cells("A1:E1")
    ws["A1"].value = "Pitcher Strikeout Props — Model Performance Summary"
    ws["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill  = header_fill(CLR_HEADER_DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14

    def stat_block(title, stats, start_row, bg=CLR_HEADER_LIGHT):
        ws.merge_cells(f"A{start_row}:E{start_row}")
        ws[f"A{start_row}"].value = title
        ws[f"A{start_row}"].font  = Font(bold=True, size=11, color="FFFFFF")
        ws[f"A{start_row}"].fill  = header_fill(CLR_HEADER_MID)
        ws[f"A{start_row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[start_row].height = 22
        r = start_row + 1
        for label, val, fmt in stats:
            cell_l = ws.cell(row=r, column=1, value=label)
            cell_v = ws.cell(row=r, column=2, value=val)
            cell_l.font = Font(name="Calibri", size=10, bold=True, color=CLR_NEUTRAL)
            cell_l.fill = PatternFill("solid", fgColor=CLR_SUMMARY_BG)
            cell_l.border = thin_border()
            cell_v.number_format = fmt
            cell_v.font = Font(name="Calibri", size=10, bold=False)
            cell_v.fill = PatternFill("solid", fgColor="FFFFFF")
            cell_v.border = thin_border()
            if isinstance(val, (int, float)) and val >= 0 and "P" in fmt.upper():
                cell_v.font = Font(name="Calibri", size=10, color=CLR_POSITIVE, bold=True)
            elif isinstance(val, (int, float)) and val < 0 and ("$" in fmt or "%" in fmt):
                cell_v.font = Font(name="Calibri", size=10, color=CLR_NEGATIVE, bold=True)
            ws.row_dimensions[r].height = 18
            r += 1
        return r + 1

    # ── Overall stats ──────────────────────────────────────────────────────
    n = len(bets)
    wins = int(bets["won"].sum())
    total_staked = bets["bet_size"].sum()
    total_pnl = bets["pnl"].sum()
    roi = total_pnl / total_staked
    win_rate = wins / n
    final_bk = bets["running_bankroll"].iloc[-1]
    bk_gain = final_bk - STARTING_BANKROLL
    max_dd = (bets["running_bankroll"] - bets["running_bankroll"].cummax()).min()
    avg_edge = bets["edge_pct"].mean()
    avg_odds = bets["bet_odds"].mean()
    avg_bet = bets["bet_size"].mean()

    from scipy import stats as scipy_stats
    t_stat = bets["pnl"].mean() / (bets["pnl"].std(ddof=1) / np.sqrt(n))
    p_val = scipy_stats.t.sf(t_stat, df=n-1)

    nv = bets.apply(lambda r: pd.Series(devig_pair(r["over_odds"], r["under_odds"])), axis=1)
    bets["nv_over"]  = nv[0]; bets["nv_under"] = nv[1]
    bets["nv_entry"] = np.where(bets["best_side"]=="over", bets["nv_over"], bets["nv_under"])
    breakeven_exp_wins = bets["nv_entry"].sum()
    excess_wins = wins - breakeven_exp_wins
    var_wins = (bets["nv_entry"] * (1 - bets["nv_entry"])).sum()
    z_wins = excess_wins / np.sqrt(var_wins)
    p_binom = scipy_stats.norm.sf(z_wins)

    next_r = stat_block("OVERALL PERFORMANCE", [
        ("Starting Bankroll",     STARTING_BANKROLL,    "$#,##0"),
        ("Final Bankroll",        final_bk,             "$#,##0"),
        ("Total Profit",          bk_gain,              "+$#,##0;-$#,##0"),
        ("Bankroll Growth",       bk_gain/STARTING_BANKROLL, "+0.0%;-0.0%"),
        ("Total Staked",          total_staked,         "$#,##0"),
        ("P&L on Staked",         total_pnl,            "+$#,##0;-$#,##0"),
        ("ROI (on staked)",       roi,                  "+0.0%;-0.0%"),
        ("Max Drawdown",          max_dd,               "+$#,##0;-$#,##0"),
        ("Max Drawdown %",        max_dd/STARTING_BANKROLL, "+0.0%;-0.0%"),
    ], 3)

    next_r = stat_block("BET STATISTICS", [
        ("Total Bets",            n,                    "#,##0"),
        ("Wins",                  wins,                 "#,##0"),
        ("Losses",                n - wins,             "#,##0"),
        ("Win Rate",              win_rate,             "0.00%"),
        ("Avg Break-Even Rate",   bets["nv_entry"].mean(), "0.00%"),
        ("Excess Win Rate",       win_rate - bets["nv_entry"].mean(), "+0.00%;-0.00%"),
        ("Avg Edge %",            avg_edge/100,         "0.0%"),
        ("Avg Bet Odds",          avg_odds,             "+0.0;-0.0"),
        ("Avg Bet Size",          avg_bet,              "$#,##0"),
    ], next_r)

    next_r = stat_block("STATISTICAL SIGNIFICANCE", [
        ("Profit t-stat",         t_stat,               "0.000"),
        ("Profit p-value",        p_val,                "0.00000"),
        ("Gen. Binomial z-stat",  z_wins,               "0.000"),
        ("Binomial p-value",      p_binom,              "0.0000%"),
        ("Excess wins vs fair",   excess_wins,          "+0.0;-0.0"),
    ], next_r)

    next_r = stat_block("PERIOD BREAKDOWN", [
        ("2025 bets (in-sample OOS)", int((bets["game_date"].dt.year==2025).sum()), "#,##0"),
        ("2026 bets (true OOS)",      int((bets["game_date"].dt.year==2026).sum()), "#,##0"),
        ("Date range",               f"{bets['game_date'].min().strftime('%Y-%m-%d')} → {bets['game_date'].max().strftime('%Y-%m-%d')}", "@"),
        ("Bet sizing method",
         f"Half-Kelly, capped {MAX_UNITS}U (1U=${UNIT_SIZE}, fixed)", "@"),
    ], next_r)

    return ws


# ── Main ──────────────────────────────────────────────────────────────────────
print("Loading bets...")
bets = load_bets()
print(f"  {len(bets)} bets at edge >= {MIN_EDGE_PCT}%")

print("Computing bankroll...")
bets = compute_bankroll(bets)

print("Loading opening/closing odds for all books...")
try:
    opens, closes = load_all_odds()
    print(f"  Opens: {len(opens):,} rows, Closes: {len(closes):,} rows")
except Exception as e:
    print(f"  Warning: could not load odds ({e}), book columns will be empty")
    opens, closes = None, None

print("Building Excel workbook...")
wb = Workbook()
del wb["Sheet"]

build_summary_sheet(wb, bets)
build_picks_sheet(wb, bets, opens, closes)
build_bankroll_sheet(wb, bets)
build_monthly_sheet(wb, bets)
build_edge_sheet(wb, bets)

Path("data/exports").mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT_PATH)
print(f"\nSaved: {OUTPUT_PATH}")

final_bk = bets["running_bankroll"].iloc[-1]
total_pnl = bets["pnl"].sum()
roi = total_pnl / bets["bet_size"].sum()
print(f"  Final bankroll: ${final_bk:,.0f}  (started ${STARTING_BANKROLL:,})")
print(f"  Net profit: ${total_pnl:+,.0f}  |  ROI: {roi:+.1%}")
print(f"  Win rate: {bets['won'].mean():.1%} on {len(bets)} bets")

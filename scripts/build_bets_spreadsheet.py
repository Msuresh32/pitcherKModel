"""
Build a clean Excel spreadsheet focused on projection, gap, and direction.
Odds are excluded — user line-shops independently.
"""
import argparse
import pandas as pd
import numpy as np
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

ODDS_FILE   = "data/odds/june_2026_odds.csv"
EDGES_FILE  = "data/processed/clv_e10_june_edges.csv"
OUTPUT_FILE = "data/exports/june_2026_bets.xlsx"

# colours
C_HEADER_BG = "1F3864"
C_WIN       = "C6EFCE"
C_WIN_FONT  = "276221"
C_LOSS      = "FFCCCC"
C_LOSS_FONT = "9C0006"
C_ALT_ROW   = "EEF4FB"
C_CLV_POS   = "E2EFDA"
C_CLV_NEG   = "FCE4D6"
C_PUSH      = "FFF2CC"


def american_to_decimal(odds):
    try:
        o = float(odds)
    except Exception:
        return np.nan
    if pd.isna(o):
        return np.nan
    return (100 / abs(o) + 1) if o < 0 else (o / 100 + 1)


def fmt_american(odds):
    try:
        o = int(round(float(odds)))
        return f"+{o}" if o >= 0 else str(o)
    except Exception:
        return "—"


def build_clv_lookup(odds_path):
    """Best open & close price per (date, pitcher_lower, line, market)."""
    df = pd.read_csv(odds_path)
    df["game_date"] = pd.to_datetime(df["game_date"]).dt.strftime("%Y-%m-%d")
    df["_key"] = df["pitcher_name"].str.strip().str.lower()
    df["line"] = df["line"].astype(float)

    result = {}
    for snap_type in ("open", "close"):
        sub = df[df["snapshot_type"] == snap_type]
        for _, row in sub.iterrows():
            key = (row["game_date"], row["_key"], row["line"],
                   str(row.get("market", "strikeouts")))
            prev = result.get(key, {})
            o = row["over_odds"];  prev_o = prev.get(f"{snap_type}_over",  np.nan)
            u = row["under_odds"]; prev_u = prev.get(f"{snap_type}_under", np.nan)
            prev[f"{snap_type}_over"]  = o  if pd.isna(prev_o) else max(prev_o, o)
            prev[f"{snap_type}_under"] = u  if pd.isna(prev_u) else max(prev_u, u)
            result[key] = prev
    return result


def load_bets(edges_path, min_edge):
    df = pd.read_csv(edges_path)
    df = df[df["market"] == "strikeouts"].copy()
    df = df[df["edge_pct"] >= min_edge].copy()
    df["game_date"] = pd.to_datetime(df["game_date"]).dt.strftime("%Y-%m-%d")
    df["gap"] = df["strikeouts_projection"] - df["line"]
    df["won"] = df.apply(
        lambda r: (r["strikeouts"] > r["line"]) if r["best_side"] == "over"
                  else (r["strikeouts"] < r["line"]), axis=1)
    # deduplicate: same pitcher/date/line/side from multiple bookmakers — keep best edge
    df = (df.sort_values("edge_pct", ascending=False)
            .drop_duplicates(subset=["game_date", "pitcher_name", "line", "best_side"])
            .sort_values(["game_date", "pitcher_name"])
            .reset_index(drop=True))
    return df


def confidence_tier(gap, edge):
    if abs(gap) >= 1.5 or edge >= 25:
        return "High"
    elif abs(gap) >= 0.75 or edge >= 15:
        return "Medium"
    else:
        return "Low"


def add_bets_sheet(wb, bets, clv_lookup):
    ws = wb.active
    ws.title = "All Bets"

    headers = [
        "Date", "Pitcher", "Bet",
        "Projection", "Line", "Gap", "Confidence",
        "Edge %", "Open Odds", "Close Odds", "CLV %",
        "Actual Ks", "Result",
    ]

    thin   = Side(style="thin", color="BBCFE0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    # header
    ws.row_dimensions[1].height = 24
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = center
        c.border    = border

    win_fill  = PatternFill("solid", fgColor=C_WIN)
    loss_fill = PatternFill("solid", fgColor=C_LOSS)
    alt_fill  = PatternFill("solid", fgColor=C_ALT_ROW)
    white     = PatternFill("solid", fgColor="FFFFFF")
    clv_pos   = PatternFill("solid", fgColor=C_CLV_POS)
    clv_neg   = PatternFill("solid", fgColor=C_CLV_NEG)
    push_fill = PatternFill("solid", fgColor=C_PUSH)

    high_conf_font  = Font(bold=True, size=10, color="1F3864")
    win_font        = Font(bold=True, size=10, color=C_WIN_FONT)
    loss_font       = Font(bold=True, size=10, color=C_LOSS_FONT)
    base_font       = Font(size=10)

    for ri, (_, bet) in enumerate(bets.iterrows(), start=2):
        row_fill = alt_fill if ri % 2 == 0 else white
        side     = bet["best_side"]
        line     = float(bet["line"])
        gap      = float(bet["gap"])
        edge     = float(bet["edge_pct"])
        proj     = float(bet["strikeouts_projection"])
        won      = bool(bet["won"])
        conf     = confidence_tier(gap, edge)

        # CLV lookup
        key = (bet["game_date"], bet["pitcher_name"].strip().lower(),
               line, str(bet.get("market", "strikeouts")))
        lk = clv_lookup.get(key, {})
        open_o  = lk.get("open_over",  np.nan)
        open_u  = lk.get("open_under", np.nan)
        close_o = lk.get("close_over", np.nan)
        close_u = lk.get("close_under",np.nan)
        entry   = open_o  if side == "over" else open_u
        closing = close_o if side == "over" else close_u
        e_dec   = american_to_decimal(entry)
        c_dec   = american_to_decimal(closing)
        if not pd.isna(e_dec) and not pd.isna(c_dec) and c_dec > 1:
            clv = ((e_dec / c_dec) - 1) * 100
        else:
            clv = np.nan

        close_odds_val = closing if not pd.isna(closing) else np.nan
        close_odds_str = fmt_american(close_odds_val) if not pd.isna(close_odds_val) else "—"

        bet_label    = f"{'Over' if side == 'over' else 'Under'} {line:.1f}"
        result_label = "WIN" if won else "LOSS"

        open_odds_str = fmt_american(entry) if not pd.isna(entry) else "—"

        row_data = [
            bet["game_date"],
            bet["pitcher_name"],
            bet_label,
            round(proj, 2),
            line,
            round(gap, 2),
            conf,
            round(edge, 1),
            open_odds_str,
            close_odds_str,
            f"{clv:+.2f}%" if not pd.isna(clv) else "—",
            int(bet["strikeouts"]) if not pd.isna(bet.get("strikeouts")) else "—",
            result_label,
        ]

        res_fill = win_fill if won else loss_fill
        res_font = win_font if won else loss_font

        for ci, val in enumerate(row_data, 1):
            col_name = headers[ci - 1]
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = border
            c.font   = base_font

            if col_name == "Result":
                c.fill = res_fill
                c.font = res_font
                c.alignment = center
            elif col_name == "CLV %":
                c.fill = (clv_pos if (not pd.isna(clv) and clv > 0)
                           else clv_neg if (not pd.isna(clv) and clv <= 0)
                           else row_fill)
                c.alignment = center
            elif col_name == "Confidence":
                if conf == "High":
                    c.fill = PatternFill("solid", fgColor="D9E8FF")
                    c.font = high_conf_font
                else:
                    c.fill = row_fill
                c.alignment = center
            elif col_name == "Gap":
                c.fill = row_fill
                c.alignment = center
                # bold if gap >= 1.0
                if abs(gap) >= 1.0:
                    c.font = Font(bold=True, size=10)
            elif col_name in ("Open Odds", "Close Odds"):
                c.fill = row_fill
                c.alignment = center
                if val == "—":
                    c.font = Font(size=10, color="AAAAAA")
            elif col_name in ("Projection", "Line", "Edge %", "Actual Ks"):
                c.fill = row_fill
                c.alignment = center
            elif col_name in ("Pitcher", "Bet"):
                c.fill = row_fill
                c.alignment = left
            else:
                c.fill = row_fill
                c.alignment = center

        ws.row_dimensions[ri].height = 18

    col_widths = {
        "Date": 11, "Pitcher": 22, "Bet": 12,
        "Projection": 11, "Line": 7, "Gap": 7,
        "Confidence": 12, "Edge %": 8,
        "Open Odds": 11, "Close Odds": 11, "CLV %": 9,
        "Actual Ks": 10, "Result": 8,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def add_summary_sheet(wb, bets, clv_lookup):
    ws = wb.create_sheet("Summary")

    thin   = Side(style="thin", color="BBCFE0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    def write(row, label, val, bold=False, header=False):
        for ci, v in enumerate([label, val], 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = border
            c.alignment = center
            if header:
                c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
                c.font = Font(bold=True, color="FFFFFF", size=11)
            else:
                c.font = Font(bold=bold, size=10)
        ws.row_dimensions[row].height = 20

    wins = bets["won"].sum()
    n    = len(bets)
    clvs = []
    for _, bet in bets.iterrows():
        side = bet["best_side"]
        key  = (bet["game_date"], bet["pitcher_name"].strip().lower(),
                float(bet["line"]), str(bet.get("market", "strikeouts")))
        lk = clv_lookup.get(key, {})
        entry   = lk.get("open_over",  np.nan) if side == "over" else lk.get("open_under",  np.nan)
        closing = lk.get("close_over", np.nan) if side == "over" else lk.get("close_under", np.nan)
        e_dec = american_to_decimal(entry)
        c_dec = american_to_decimal(closing)
        if not pd.isna(e_dec) and not pd.isna(c_dec) and c_dec > 1:
            clvs.append(((e_dec / c_dec) - 1) * 100)

    gap_05 = bets[bets["gap"].abs() >= 0.5]
    gap_10 = bets[bets["gap"].abs() >= 1.0]

    rows = [
        ("JUNE 2026 SUMMARY",         "",                                True,  True),
        ("Period",                     "Jun 1 – 16, 2026",               False, False),
        ("Min Edge Filter",            "10%",                            False, False),
        ("",                           "",                               False, False),
        ("── Overall ──",              "",                               True,  False),
        ("Total Bets",                 n,                                False, False),
        ("Win Rate",                   f"{wins/n:.1%}",                  True,  False),
        ("Bets with CLV data",         len(clvs),                        False, False),
        ("Mean CLV",                   f"{np.mean(clvs):+.2f}%" if clvs else "N/A", True, False),
        ("CLV > 0",                    f"{sum(1 for c in clvs if c>0)}/{len(clvs)}", False, False),
        ("",                           "",                               False, False),
        ("── By Gap Tier ──",          "",                               True,  False),
        ("Gap >= 0.5  (n)",            len(gap_05),                      False, False),
        ("Gap >= 0.5  Win%",           f"{gap_05['won'].mean():.1%}" if len(gap_05) else "N/A", True, False),
        ("Gap >= 1.0  (n)",            len(gap_10),                      False, False),
        ("Gap >= 1.0  Win%",           f"{gap_10['won'].mean():.1%}" if len(gap_10) else "N/A", True, False),
    ]

    for i, (label, val, bold, header) in enumerate(rows, 1):
        write(i, label, val, bold=bold, header=header)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--edges",     default=EDGES_FILE)
    p.add_argument("--odds",      default=ODDS_FILE)
    p.add_argument("--output",    default=OUTPUT_FILE)
    p.add_argument("--min-edge",  type=float, default=10.0)
    args = p.parse_args()

    print("Loading bets...")
    bets = load_bets(args.edges, args.min_edge)
    print(f"  {len(bets)} bets at edge >= {args.min_edge}%")

    print("Building CLV lookup from open/close odds...")
    clv_lookup = build_clv_lookup(args.odds)

    print("Building workbook...")
    wb = Workbook()
    add_bets_sheet(wb, bets, clv_lookup)
    add_summary_sheet(wb, bets, clv_lookup)

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"\nSaved -> {args.output}")


if __name__ == "__main__":
    main()

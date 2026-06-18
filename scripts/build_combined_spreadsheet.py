"""
Combined 2025 + 2026 backtest spreadsheet.
All bets, all metrics. 2025 = open odds only. 2026 = open + close + CLV.
"""
import pandas as pd
import numpy as np
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

EDGES_2025  = "data/processed_2024/thresh_sel_2025_dk_edges.csv"
EDGES_2026  = "data/processed/frozen_thresh_2026_edges.csv"
ODDS_2025   = "data/odds/hist_2025_dk_open.csv"        # open only
ODDS_2026   = "data/odds/full_2026_odds_dk.csv"        # open + close
OUTPUT      = "data/exports/combined_2025_2026_backtest.xlsx"
MIN_EDGE    = 5.0   # show everything meaningful

# ── colours ─────────────────────────────────────────────────────────────────
C_HDR       = "1F3864"
C_WIN       = "C6EFCE"; C_WIN_F   = "276221"
C_LOSS      = "FFCCCC"; C_LOSS_F  = "9C0006"
C_ALT       = "EEF4FB"
C_CLV_POS   = "E2EFDA"; C_CLV_NEG = "FCE4D6"
C_YR25      = "FFF8E7"   # warm tint for 2025 rows
C_YR26      = "F0F4FB"   # cool tint for 2026 rows
C_MONTH_25  = "7D6608"
C_MONTH_26  = "2E4057"
C_HIGH_CONF = "D9E8FF"


def american_to_decimal(odds):
    try:
        o = float(odds)
    except Exception:
        return np.nan
    if pd.isna(o): return np.nan
    return (100 / abs(o) + 1) if o < 0 else (o / 100 + 1)


def fmt_american(odds):
    try:
        o = int(round(float(odds)))
        return f"+{o}" if o >= 0 else str(o)
    except Exception:
        return "—"


def build_odds_lookup(path, has_close):
    """Returns dict keyed by (date_str, pitcher_lower, line, market)."""
    df = pd.read_csv(path)
    df["game_date"] = pd.to_datetime(df["game_date"]).dt.strftime("%Y-%m-%d")
    df["_key"] = df["pitcher_name"].str.strip().str.lower()
    df["line"]  = df["line"].astype(float)
    df["market"] = df.get("market", "strikeouts")

    result = {}
    snaps = ["open", "close"] if has_close else ["open"]

    for snap in snaps:
        if has_close:
            sub = df[df["snapshot_type"] == snap]
        else:
            sub = df  # single snapshot, treat as open

        for _, row in sub.iterrows():
            key = (row["game_date"], row["_key"], row["line"],
                   str(row.get("market", "strikeouts")))
            prev = result.get(key, {})
            o = row["over_odds"];  po = prev.get(f"{snap}_over",  np.nan)
            u = row["under_odds"]; pu = prev.get(f"{snap}_under", np.nan)
            prev[f"{snap}_over"]  = o if pd.isna(po) else max(po, o)
            prev[f"{snap}_under"] = u if pd.isna(pu) else max(pu, u)
            result[key] = prev
    return result


def load_bets(path, year_label, min_edge):
    df = pd.read_csv(path)
    df = df[df["market"] == "strikeouts"].copy()
    df = df[df["edge_pct"] >= min_edge].copy()
    df["game_date"] = pd.to_datetime(df["game_date"]).dt.strftime("%Y-%m-%d")
    df["gap"] = df["strikeouts_projection"] - df["line"]
    df["won"] = df.apply(
        lambda r: (r["strikeouts"] > r["line"]) if r["best_side"] == "over"
                  else (r["strikeouts"] < r["line"]), axis=1)
    df["year"] = year_label
    df = (df.sort_values("edge_pct", ascending=False)
            .drop_duplicates(subset=["game_date", "pitcher_name", "line", "best_side"])
            .sort_values(["game_date", "pitcher_name"])
            .reset_index(drop=True))
    return df


def confidence_tier(gap, edge):
    if abs(gap) >= 1.5 or edge >= 25: return "High"
    if abs(gap) >= 0.75 or edge >= 15: return "Medium"
    return "Low"


def border(color="BBCFE0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def add_bets_sheet(wb, bets, lk25, lk26):
    ws = wb.active
    ws.title = "All Bets"

    headers = ["Year", "Date", "Pitcher", "Bet", "Projection", "Line", "Gap",
               "Confidence", "Edge %", "Open Odds", "Close Odds", "CLV %",
               "Actual Ks", "Result"]

    bd  = border()
    ctr = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left",   vertical="center")

    ws.row_dimensions[1].height = 26
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor=C_HDR)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = ctr
        c.border = bd

    win_fill  = PatternFill("solid", fgColor=C_WIN)
    loss_fill = PatternFill("solid", fgColor=C_LOSS)
    clv_pos   = PatternFill("solid", fgColor=C_CLV_POS)
    clv_neg   = PatternFill("solid", fgColor=C_CLV_NEG)
    hi_fill   = PatternFill("solid", fgColor=C_HIGH_CONF)
    win_font  = Font(bold=True, size=10, color=C_WIN_F)
    loss_font = Font(bold=True, size=10, color=C_LOSS_F)
    base_font = Font(size=10)
    hi_font   = Font(bold=True, size=10, color="1F3864")

    prev_month = None
    ri = 2

    for _, bet in bets.iterrows():
        month = bet["game_date"][:7]
        yr    = bet["year"]

        if month != prev_month:
            month_color = C_MONTH_25 if yr == 2025 else C_MONTH_26
            label = pd.to_datetime(bet["game_date"]).strftime("%B %Y")
            ws.row_dimensions[ri].height = 15
            c = ws.cell(row=ri, column=1, value=f"  {label}")
            c.fill = PatternFill("solid", fgColor=month_color)
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.alignment = lft
            ws.merge_cells(start_row=ri, start_column=1,
                           end_row=ri, end_column=len(headers))
            prev_month = month
            ri += 1

        row_base = C_YR25 if yr == 2025 else C_YR26
        row_fill = PatternFill("solid", fgColor=row_base)

        side = bet["best_side"]
        line = float(bet["line"])
        gap  = float(bet["gap"])
        edge = float(bet["edge_pct"])
        proj = float(bet["strikeouts_projection"])
        won  = bool(bet["won"])
        conf = confidence_tier(gap, edge)

        lk = lk26 if yr == 2026 else lk25
        key = (bet["game_date"], bet["pitcher_name"].strip().lower(),
               line, str(bet.get("market", "strikeouts")))
        d = lk.get(key, {})
        entry   = d.get("open_over",  np.nan) if side == "over" else d.get("open_under",  np.nan)
        closing = d.get("close_over", np.nan) if side == "over" else d.get("close_under", np.nan)
        e_dec = american_to_decimal(entry)
        c_dec = american_to_decimal(closing)
        clv   = ((e_dec / c_dec) - 1) * 100 \
                if not pd.isna(e_dec) and not pd.isna(c_dec) and c_dec > 1 else np.nan

        close_str = fmt_american(closing) if not pd.isna(closing) else ("N/A (2025)" if yr == 2025 else "—")
        clv_str   = f"{clv:+.2f}%" if not pd.isna(clv) else ("N/A (2025)" if yr == 2025 else "—")

        row_data = [
            str(yr),
            bet["game_date"],
            bet["pitcher_name"],
            f"{'Over' if side=='over' else 'Under'} {line:.1f}",
            round(proj, 2),
            line,
            round(gap, 2),
            conf,
            round(edge, 1),
            fmt_american(entry) if not pd.isna(entry) else "—",
            close_str,
            clv_str,
            int(bet["strikeouts"]) if not pd.isna(bet.get("strikeouts")) else "—",
            "WIN" if won else "LOSS",
        ]

        ws.row_dimensions[ri].height = 18
        for ci, val in enumerate(row_data, 1):
            col = headers[ci - 1]
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = bd
            c.font   = base_font

            if col == "Result":
                c.fill = win_fill if won else loss_fill
                c.font = win_font if won else loss_font
                c.alignment = ctr
            elif col == "CLV %":
                if not pd.isna(clv):
                    c.fill = clv_pos if clv > 0 else clv_neg
                else:
                    c.fill = row_fill
                    c.font = Font(size=9, color="999999")
                c.alignment = ctr
            elif col == "Confidence":
                c.fill = hi_fill if conf == "High" else row_fill
                c.font = hi_font if conf == "High" else base_font
                c.alignment = ctr
            elif col == "Gap":
                c.fill = row_fill; c.alignment = ctr
                if abs(gap) >= 1.0: c.font = Font(bold=True, size=10)
            elif col in ("Year", "Projection", "Line", "Edge %", "Actual Ks",
                         "Open Odds", "Close Odds", "Date"):
                c.fill = row_fill; c.alignment = ctr
            elif col in ("Pitcher", "Bet"):
                c.fill = row_fill; c.alignment = lft
            else:
                c.fill = row_fill; c.alignment = ctr

        ri += 1

    widths = {"Year": 6, "Date": 11, "Pitcher": 22, "Bet": 12,
              "Projection": 11, "Line": 7, "Gap": 7, "Confidence": 12,
              "Edge %": 8, "Open Odds": 11, "Close Odds": 12, "CLV %": 12,
              "Actual Ks": 10, "Result": 8}
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 12)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def add_summary_sheet(wb, bets, lk26):
    ws = wb.create_sheet("Summary")
    bd  = border()
    ctr = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def hdr(r, text, cols=3):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        c = ws.cell(row=r, column=1, value=text)
        c.fill = PatternFill("solid", fgColor=C_HDR)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = lft; c.border = bd
        ws.row_dimensions[r].height = 24

    def row(r, label, val, bold=False, note=""):
        cells = [(1, label), (2, val), (3, note)]
        for ci, v in cells:
            c = ws.cell(row=r, column=ci, value=v)
            c.border = bd; c.alignment = ctr
            c.font = Font(bold=bold, size=10)
            if ci == 1: c.alignment = lft
            if ci == 3: c.font = Font(size=9, color="777777"); c.alignment = lft
        ws.row_dimensions[r].height = 20

    def blank(r):
        for ci in range(1, 4):
            c = ws.cell(row=r, column=ci, value="")
            c.fill = PatternFill("solid", fgColor="F8FAFC"); c.border = bd
        ws.row_dimensions[r].height = 7

    bets = bets.copy()
    bets["pay"] = bets.apply(
        lambda r: (r["over_odds"]/100 if r["over_odds"]>0 else 100/abs(r["over_odds"]))
                  if r["best_side"]=="over"
                  else (r["under_odds"]/100 if r["under_odds"]>0 else 100/abs(r["under_odds"])), axis=1)
    bets["profit"] = bets.apply(lambda r: r["pay"] if r["won"] else -1.0, axis=1)
    bets["game_date"] = pd.to_datetime(bets["game_date"])

    # CLV for 2026 only
    clvs_26 = []
    for _, bet in bets[bets["year"]==2026].iterrows():
        side = bet["best_side"]
        key  = (bet["game_date"].strftime("%Y-%m-%d"),
                bet["pitcher_name"].strip().lower(),
                float(bet["line"]), str(bet.get("market","strikeouts")))
        d  = lk26.get(key, {})
        e  = d.get("open_over",  np.nan) if side=="over" else d.get("open_under",  np.nan)
        cl = d.get("close_over", np.nan) if side=="over" else d.get("close_under", np.nan)
        ed = american_to_decimal(e); cd = american_to_decimal(cl)
        if not pd.isna(ed) and not pd.isna(cd) and cd > 1:
            clvs_26.append(((ed/cd)-1)*100)

    r = 1
    hdr(r, "COMBINED 2025 + 2026 BACKTEST — WALK-FORWARD VALIDATED"); r+=1
    row(r, "Methodology", "Threshold frozen on 2025 OOS, applied blind to 2026"); r+=1
    row(r, "Threshold", "Edge >= 15% (DraftKings lines only)"); r+=1
    row(r, "Models", "2024-only model for 2025 | Old model (2023-25) for 2026"); r+=1
    blank(r); r+=1

    for yr_label, yr_df in [("2025", bets[bets["year"]==2025]),
                             ("2026", bets[bets["year"]==2026]),
                             ("COMBINED", bets)]:
        is_combined = (yr_label == "COMBINED")
        hdr(r, f"{'  ' if not is_combined else ''}{'★ ' if is_combined else ''}{yr_label}"); r+=1
        n   = len(yr_df)
        wins = yr_df["won"].sum()
        roi  = yr_df["profit"].mean()
        sharpe = roi / yr_df["profit"].std() * n**0.5 if yr_df["profit"].std() > 0 else 0
        row(r, "Total bets",  n,                  bold=is_combined); r+=1
        row(r, "Win rate",    f"{wins/n:.1%}",     bold=is_combined); r+=1
        row(r, "ROI",         f"{roi:+.1%}",       bold=is_combined); r+=1
        row(r, "Sharpe",      f"{sharpe:.2f}",     bold=is_combined); r+=1
        if yr_label == "2026":
            row(r, "Mean CLV (2026)", f"{np.mean(clvs_26):+.2f}%" if clvs_26 else "N/A",
                bold=True, note=f"{len(clvs_26)} bets"); r+=1
        blank(r); r+=1

    # Monthly table
    hdr(r, "MONTHLY BREAKDOWN"); r+=1
    for ci, h in enumerate(["Month", "Bets", "Win%", "ROI"], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor="2E4057")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = ctr; c.border = bd
    ws.row_dimensions[r].height = 20; r+=1

    bets["month"] = bets["game_date"].dt.to_period("M")
    alt = ["FFFFFF", "F0F4FA"]
    for mi, (m, g) in enumerate(bets.groupby("month")):
        yr_val = g["year"].iloc[0]
        roi_m  = g["profit"].mean()
        rfill  = PatternFill("solid", fgColor="C6EFCE" if roi_m>0 else "FFCCCC")
        bfill  = PatternFill("solid", fgColor=alt[mi%2])
        for ci, v in enumerate([str(m), len(g), f"{g['won'].mean():.1%}", f"{roi_m:+.1%}"], 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = Font(size=10); c.alignment = ctr; c.border = bd
            c.fill = rfill if ci==4 else bfill
        ws.row_dimensions[r].height = 20; r+=1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12


def add_stats_sheet(wb, bets, lk26):
    ws = wb.create_sheet("Deep Stats")
    bd  = border()
    ctr = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left",   vertical="center")

    bets = bets.copy()
    bets["pay"] = bets.apply(
        lambda r: (r["over_odds"]/100 if r["over_odds"]>0 else 100/abs(r["over_odds"]))
                  if r["best_side"]=="over"
                  else (r["under_odds"]/100 if r["under_odds"]>0 else 100/abs(r["under_odds"])), axis=1)
    bets["profit"] = bets.apply(lambda r: r["pay"] if r["won"] else -1.0, axis=1)

    def section_hdr(r, text, ncols=5):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1, value=text)
        c.fill = PatternFill("solid", fgColor=C_HDR)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = lft; c.border = bd
        ws.row_dimensions[r].height = 22

    def table_hdr(r, cols):
        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.fill = PatternFill("solid", fgColor="2E4057")
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.alignment = ctr; c.border = bd
        ws.row_dimensions[r].height = 20

    def data_row(r, vals, highlight_col=None, green_if_pos_col=None):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = Font(size=10); c.alignment = ctr; c.border = bd
            bg = "F8FAFC" if r % 2 == 0 else "FFFFFF"
            c.fill = PatternFill("solid", fgColor=bg)
            if green_if_pos_col and ci == green_if_pos_col:
                try:
                    num = float(str(v).replace("%","").replace("+",""))
                    c.fill = PatternFill("solid", fgColor="C6EFCE" if num>=0 else "FFCCCC")
                except: pass
        ws.row_dimensions[r].height = 18

    r = 1

    # 1. Edge threshold breakdown
    section_hdr(r, "EDGE THRESHOLD BREAKDOWN (both years combined)"); r+=1
    table_hdr(r, ["Min Edge", "Bets", "Win%", "ROI", "Sharpe"]); r+=1
    for t in [5, 7, 10, 12, 15, 18, 20, 25]:
        sub = bets[bets["edge_pct"] >= t]
        if len(sub) < 20: continue
        sh = sub["profit"].mean()/sub["profit"].std()*len(sub)**0.5
        data_row(r, [f">= {t}%", len(sub), f"{sub['won'].mean():.1%}",
                     f"{sub['profit'].mean():+.1%}", f"{sh:.2f}"],
                 green_if_pos_col=4)
        r+=1

    r+=1
    # 2. Gap tier breakdown
    section_hdr(r, "PROJECTION GAP BREAKDOWN  (gap = model projection - line)"); r+=1
    table_hdr(r, ["Gap Band", "Bets", "Win%", "ROI", "Notes"]); r+=1
    gap_bands = [
        ("gap < 0  (under bet)",       bets[bets["gap"] < 0],             "Model below line"),
        ("gap 0.0–0.5  (weak over)",   bets[(bets["gap"]>=0)&(bets["gap"]<0.5)],  "Dead-money zone"),
        ("gap 0.5–1.0",                bets[(bets["gap"]>=0.5)&(bets["gap"]<1.0)], ""),
        ("gap 1.0–1.5",                bets[(bets["gap"]>=1.0)&(bets["gap"]<1.5)], ""),
        ("gap 1.5+",                   bets[bets["gap"] >= 1.5],                   "High conviction"),
    ]
    for label, sub, note in gap_bands:
        if len(sub) == 0: continue
        data_row(r, [label, len(sub), f"{sub['won'].mean():.1%}",
                     f"{sub['profit'].mean():+.1%}", note], green_if_pos_col=4)
        r+=1

    r+=1
    # 3. Over vs Under
    section_hdr(r, "OVER vs UNDER SPLIT"); r+=1
    table_hdr(r, ["Side", "Bets", "Win%", "ROI", "Avg Odds"]); r+=1
    for side, label in [("over","OVER"), ("under","UNDER")]:
        sub = bets[bets["best_side"]==side]
        avg_odds = sub.apply(lambda x: x["over_odds"] if x["best_side"]=="over"
                             else x["under_odds"], axis=1).mean()
        data_row(r, [label, len(sub), f"{sub['won'].mean():.1%}",
                     f"{sub['profit'].mean():+.1%}", f"{avg_odds:+.0f}"],
                 green_if_pos_col=4)
        r+=1

    r+=1
    # 4. Significance tests
    from scipy import stats
    section_hdr(r, "STATISTICAL SIGNIFICANCE TESTS"); r+=1
    table_hdr(r, ["Period", "p-value (t-test)", "p-value (binomial)", "95% CI lo", "95% CI hi"]); r+=1
    np.random.seed(42)

    for label, sub in [("2025", bets[bets["year"]==2025]),
                        ("2026", bets[bets["year"]==2026]),
                        ("Combined", bets)]:
        p = sub["profit"].values
        w = sub["won"].values
        implied = []
        for _, bet in sub.iterrows():
            o = bet["over_odds"] if bet["best_side"]=="over" else bet["under_odds"]
            implied.append(abs(o)/(abs(o)+100) if o<0 else 100/(o+100))
        bep = np.mean(implied)
        t_stat, p2 = stats.ttest_1samp(p, 0)
        p_t = p2/2
        p_b = stats.binom_test(int(w.sum()), len(w), bep, alternative="greater")
        boot = [np.random.choice(p, len(p), replace=True).mean() for _ in range(10000)]
        lo, hi = np.percentile(boot, [2.5, 97.5])
        data_row(r, [label, f"{p_t:.4f}", f"{p_b:.4f}",
                     f"{lo:+.2%}", f"{hi:+.2%}"], green_if_pos_col=4)
        r+=1

    for col, w in zip(["A","B","C","D","E"], [30,18,20,12,12]):
        ws.column_dimensions[col].width = w


def main():
    print("Loading bets...")
    s25 = load_bets(EDGES_2025, 2025, MIN_EDGE)
    s26 = load_bets(EDGES_2026, 2026, MIN_EDGE)
    all_bets = pd.concat([s25, s26], ignore_index=True)
    all_bets = all_bets.sort_values("game_date").reset_index(drop=True)
    print(f"  2025: {len(s25)} bets  |  2026: {len(s26)} bets  |  Total: {len(all_bets)}")

    print("Building odds lookups...")
    lk25 = build_odds_lookup(ODDS_2025, has_close=False)
    lk26 = build_odds_lookup(ODDS_2026, has_close=True)

    print("Building workbook...")
    wb = Workbook()
    add_bets_sheet(wb, all_bets, lk25, lk26)
    add_summary_sheet(wb, all_bets, lk26)
    add_stats_sheet(wb, all_bets, lk26)

    Path(OUTPUT).parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\nSaved -> {OUTPUT}")


if __name__ == "__main__":
    main()

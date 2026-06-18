"""
Bet sizing simulation spreadsheet.
Compares 4 strategies on the walk-forward 2026 validated bets (edge >= 15%).
Starting bankroll: $1,000

Strategies:
  Flat 1%       – $10 every bet, fixed
  Flat 2%       – $20 every bet, fixed
  Half Kelly    – kelly_fraction * 0.5 * current_bankroll  (kelly capped at 10%, so eff cap 5%)
  Full Kelly    – kelly_fraction * 1.0 * current_bankroll  (kelly capped at 10%)
"""
import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as numfmt)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── helpers ──────────────────────────────────────────────────────────────────
def hex_fill(h): return PatternFill("solid", fgColor=h)
def bold(ws, cell): ws[cell].font = Font(bold=True)
def hdr_style(ws, row, col, val, bg="1F3864", fg="FFFFFF"):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=True, color=fg, size=10)
    c.fill      = hex_fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = Border(
        bottom=Side(style="thin", color="FFFFFF"),
        right =Side(style="thin", color="FFFFFF"))
    return c

def auto_width(ws, padding=2):
    for col in ws.columns:
        mx = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + padding, 40)

def money(v):    return f"${v:,.2f}"
def pct(v):      return f"{v:+.1%}"
def pct2(v):     return f"{v:.1%}"

STARTING = 1_000.0
KELLY_CAP = 0.10   # raw Kelly capped at 10% before fractioning


# ── load walk-forward data ────────────────────────────────────────────────────
def load(pfx, d):
    df = pd.read_csv(f"{d}/{pfx}_edges.csv")
    df = df[df["market"] == "strikeouts"].copy()
    df["won"] = df.apply(
        lambda r: (r["strikeouts"] > r["line"]) if r["best_side"] == "over"
                  else (r["strikeouts"] < r["line"]), axis=1)
    df["payout_odds"] = df.apply(
        lambda r: r["over_odds"]  if r["best_side"] == "over" else r["under_odds"], axis=1)
    df["decimal"] = df["payout_odds"].apply(
        lambda o: o/100 + 1 if o > 0 else 100/abs(o) + 1)
    df["net_b"]   = df["decimal"] - 1   # net payout per $1 staked
    df["win_prob"] = df.apply(
        lambda r: r["over_probability"] if r["best_side"] == "over" else r["under_probability"], axis=1)
    # raw uncapped Kelly
    df["kelly_raw"] = df.apply(
        lambda r: max(0.0, (r["win_prob"] * r["net_b"] - (1 - r["win_prob"])) / r["net_b"]),
        axis=1)
    df["kelly_capped"] = df["kelly_raw"].clip(upper=KELLY_CAP)
    df["gap"]   = df["strikeouts_projection"] - df["line"]
    df = (df.sort_values("edge_pct", ascending=False)
            .drop_duplicates(subset=["game_date","pitcher_name","line","best_side"])
            .reset_index(drop=True))
    return df[df["edge_pct"] >= 15].copy()

p1 = load("wf2026_p1_mar_apr", "data/processed");        p1["period"] = "Mar-Apr"
p2 = load("wf2026_p2_may",     "data/processed_apr2026"); p2["period"] = "May"
p3 = load("wf2026_p3_jun",     "data/processed");        p3["period"] = "Jun"
bets = (pd.concat([p1, p2, p3], ignore_index=True)
          .sort_values(["game_date","pitcher_name"])
          .reset_index(drop=True))
bets["game_date"] = pd.to_datetime(bets["game_date"])

N = len(bets)

# ── simulate 4 strategies ─────────────────────────────────────────────────────
STRATEGIES = {
    "Flat 1%":      {"type": "flat",  "frac": 0.01},
    "Flat 2%":      {"type": "flat",  "frac": 0.02},
    "Qtr Kelly":    {"type": "kelly", "mult": 0.25},
    "Half Kelly":   {"type": "kelly", "mult": 0.50},
    "Full Kelly":   {"type": "kelly", "mult": 1.00},
}

def simulate(bets, strat):
    bk = STARTING
    rows = []
    for _, bet in bets.iterrows():
        if strat["type"] == "flat":
            stake = STARTING * strat["frac"]
        else:
            k = bet["kelly_capped"] * strat["mult"]
            stake = bk * k
        stake = max(stake, 0.0)
        if bet["won"]:
            profit = stake * bet["net_b"]
        else:
            profit = -stake
        bk += profit
        rows.append({
            "stake":  stake,
            "profit": profit,
            "bk":     bk,
        })
    return pd.DataFrame(rows)

sims = {name: simulate(bets, cfg) for name, cfg in STRATEGIES.items()}

def summary(sim, bets):
    n       = len(sim)
    profits = sim["profit"].values
    roi     = profits.sum() / (STARTING * 1)   # relative to starting bankroll
    final   = sim["bk"].iloc[-1]
    gross   = final - STARTING
    staked  = sim["stake"].sum()
    roi_on_staked = profits.sum() / staked
    wins    = bets["won"].sum()

    # drawdown
    pk = sim["bk"].cummax()
    dd = (sim["bk"] - pk) / pk
    max_dd_pct = dd.min()
    max_dd_amt = (sim["bk"] - pk).min()

    # per-bet sharpe (profit/stake, annualised? No — just per-bet units)
    unit_p  = profits / sim["stake"].replace(0, np.nan)
    sharpe  = unit_p.mean() / unit_p.std() * n**0.5

    # longest losing streak
    streak = cur = 0
    for w in bets["won"]:
        cur = cur + 1 if not w else 0
        streak = max(streak, cur)

    return {
        "Bets":             n,
        "Wins":             int(wins),
        "Win Rate":         wins / n,
        "Total Staked ($)": staked,
        "Total Profit ($)": profits.sum(),
        "Final Bankroll($)": final,
        "Bankroll Growth":  (final - STARTING) / STARTING,
        "ROI on Staked":    roi_on_staked,
        "Max Drawdown ($)": max_dd_amt,
        "Max Drawdown (%)": max_dd_pct,
        "Sharpe":           sharpe,
        "Longest Lose Str": streak,
    }

summaries = {name: summary(sims[name], bets) for name in STRATEGIES}


# ── build spreadsheet ─────────────────────────────────────────────────────────
wb = Workbook()

# ── TAB 1: GUIDE ─────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Guide"

guide_rows = [
    ["MLB Pitcher K-Props  —  Bet Sizing Simulator", None],
    [None, None],
    ["What is Sharpe Ratio?", None],
    ["Sharpe measures return per unit of risk. Formula: (avg profit per bet) / (std dev of profit) × sqrt(bets).", None],
    ["In finance, Sharpe > 1 = good, > 2 = excellent. Our walk-forward 2026 has Sharpe 1.86, meaning", None],
    ["we earn +1.86 standard deviations of return per bet. It penalises volatile streaks — a model that", None],
    ["wins 60% at -110 odds has high Sharpe; one that wins big infrequently has low Sharpe even at same ROI.", None],
    [None, None],
    ["What is Kelly Criterion?", None],
    ["Kelly tells you the mathematically optimal % of bankroll to bet to maximise long-run growth.", None],
    ["Formula:  f* = (p × b – q) / b    where p = model win probability, q = 1–p, b = net payout per $1 staked.", None],
    ["Kelly is aggressive — it maximises growth but produces large drawdowns. Most pros use Half Kelly.", None],
    [None, None],
    ["The 4 Strategies Compared", None],
    ["Strategy",  "Description"],
    ["Flat 1%",   "Bet $10 every single bet (1% of $1,000 starting bankroll). Never changes."],
    ["Flat 2%",   "Bet $20 every single bet (2% of $1,000 starting bankroll). Never changes."],
    ["Half Kelly","Bet (kelly × 0.5) of CURRENT bankroll each bet. Kelly capped at 10% before halving."],
    ["Full Kelly","Bet kelly × 1.0 of CURRENT bankroll each bet. Kelly capped at 10%."],
    [None, None],
    ["Key concepts:", None],
    ["- Flat sizing: simple, low variance, bankroll doesn't compound.", None],
    ["- Kelly sizing: stakes grow with your bankroll (compounding), shrink after losses (self-hedging).", None],
    ["- Half Kelly is the practical standard — roughly 50% of the variance of Full Kelly.", None],
    ["- Kelly cap at 10%: even when the model sees a 30%+ edge, we never bet more than 10% of bankroll.", None],
    [None, None],
    ["Important caveat:", None],
    ["Kelly assumes your edge estimate is exact. Ours comes from a walk-forward model, so there is", None],
    ["estimation error. Half Kelly is safer because it accounts for that uncertainty implicitly.", None],
    [None, None],
    ["Starting bankroll: $1,000.  See 'Strategy Summary' and 'Bet Log' tabs.", None],
]

ws.column_dimensions["A"].width = 70
ws.column_dimensions["B"].width = 70
for i, (a, b) in enumerate(guide_rows, 1):
    wa = ws.cell(row=i, column=1, value=a)
    wb_ = ws.cell(row=i, column=2, value=b)
    if i == 1:
        wa.font = Font(bold=True, size=14, color="1F3864")
    elif a in ("What is Sharpe Ratio?", "What is Kelly Criterion?",
               "The 4 Strategies Compared", "Key concepts:", "Important caveat:"):
        wa.font = Font(bold=True, size=11, color="2E4057")
    elif a == "Strategy" and b == "Description":
        for c in (wa, wb_): c.font = Font(bold=True); c.fill = hex_fill("D9E2F3")
    elif a in ("Flat 1%", "Flat 2%", "Half Kelly", "Full Kelly"):
        wa.font = Font(bold=True)
ws.freeze_panes = None


# ── TAB 2: STRATEGY SUMMARY ───────────────────────────────────────────────────
ws2 = wb.create_sheet("Strategy Summary")
ws2.column_dimensions["A"].width = 26

metric_names = list(list(summaries.values())[0].keys())
strat_names  = list(STRATEGIES.keys())

hdr_style(ws2, 1, 1, "Metric", "1F3864")
for ci, s in enumerate(strat_names, 2):
    hdr_style(ws2, 1, ci, s, "2E4057")

COLORS = {"Flat 1%": "EBF5FB", "Flat 2%": "D6EAF8",
          "Qtr Kelly": "FEF9E7",
          "Half Kelly": "E9F7EF", "Full Kelly": "D5F5E3"}

for ri, metric in enumerate(metric_names, 2):
    ws2.cell(row=ri, column=1, value=metric).font = Font(bold=True)
    for ci, s in enumerate(strat_names, 2):
        val = summaries[s][metric]
        c = ws2.cell(row=ri, column=ci, value=val)
        c.fill = hex_fill(COLORS[s])
        if metric in ("Win Rate", "Bankroll Growth", "ROI on Staked", "Max Drawdown (%)"):
            c.number_format = "0.0%"
        elif metric in ("Total Staked ($)", "Total Profit ($)", "Final Bankroll($)", "Max Drawdown ($)"):
            c.number_format = '"$"#,##0.00'
        elif metric == "Sharpe":
            c.number_format = "0.00"
        c.alignment = Alignment(horizontal="center")

# highlight best per row
highlight_best = {
    "Total Profit ($)": max, "Final Bankroll($)": max, "Bankroll Growth": max,
    "ROI on Staked": max, "Sharpe": max,
    "Max Drawdown ($)": min, "Max Drawdown (%)": min, "Longest Lose Str": min,
}
for ri, metric in enumerate(metric_names, 2):
    if metric not in highlight_best: continue
    fn = highlight_best[metric]
    vals = {s: summaries[s][metric] for s in strat_names}
    best = fn(vals.values())
    for ci, s in enumerate(strat_names, 2):
        if vals[s] == best:
            ws2.cell(row=ri, column=ci).font = Font(bold=True, color="155724")

auto_width(ws2)
ws2.freeze_panes = "B2"


# ── TAB 3: BET LOG ────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Bet Log")
PERIOD_COLORS = {"Mar-Apr": "EBF5FB", "May": "FEF9E7", "Jun": "F0FFF4"}

cols = [
    ("#",           "#",              None),
    ("Date",        "game_date",      "date"),
    ("Pitcher",     "pitcher_name",   None),
    ("Side",        "best_side",      None),
    ("Line",        "line",           "0.5"),
    ("Projection",  "strikeouts_projection", "0.00"),
    ("Gap",         "gap",            "+0.00"),
    ("Actual K",    "strikeouts",     "0"),
    ("Odds",        "payout_odds",    "+0"),
    ("Edge %",      "edge_pct",       "0.0%"),
    ("Win Prob",    "win_prob",       "0.0%"),
    ("Raw Kelly",   "kelly_raw",      "0.0%"),
    ("Result",      "won",            None),
    ("──── Flat 1% ────", None, None),
    ("Stake",       "_f1_stake",      '"$"0.00'),
    ("Profit",      "_f1_profit",     '"$"+0.00;[Red]"$"-0.00'),
    ("Bankroll",    "_f1_bk",         '"$"#,##0.00'),
    ("──── Flat 2% ────", None, None),
    ("Stake",       "_f2_stake",      '"$"0.00'),
    ("Profit",      "_f2_profit",     '"$"+0.00;[Red]"$"-0.00'),
    ("Bankroll",    "_f2_bk",         '"$"#,##0.00'),
    ("── Qtr Kelly ──",   None, None),
    ("Kelly %",     "_qk_pct",        "0.0%"),
    ("Stake",       "_qk_stake",      '"$"0.00'),
    ("Profit",      "_qk_profit",     '"$"+0.00;[Red]"$"-0.00'),
    ("Bankroll",    "_qk_bk",         '"$"#,##0.00'),
    ("── Half Kelly ──",  None, None),
    ("Kelly %",     "_hk_pct",        "0.0%"),
    ("Stake",       "_hk_stake",      '"$"0.00'),
    ("Profit",      "_hk_profit",     '"$"+0.00;[Red]"$"-0.00'),
    ("Bankroll",    "_hk_bk",         '"$"#,##0.00'),
    ("── Full Kelly ──",  None, None),
    ("Kelly %",     "_fk_pct",        "0.0%"),
    ("Stake",       "_fk_stake",      '"$"0.00'),
    ("Profit",      "_fk_profit",     '"$"+0.00;[Red]"$"-0.00'),
    ("Bankroll",    "_fk_bk",         '"$"#,##0.00'),
]

# header
for ci, (hdr, _, _) in enumerate(cols, 1):
    if hdr.startswith("────") or hdr.startswith("──"):
        ws3.cell(row=1, column=ci, value=hdr).fill = hex_fill("2C3E50")
        ws3.cell(row=1, column=ci).font = Font(bold=True, color="FFFFFF", size=8)
    else:
        hdr_style(ws3, 1, ci, hdr, "1F3864")

# build annotated rows
f1 = sims["Flat 1%"];  f2 = sims["Flat 2%"]
qk = sims["Qtr Kelly"]; hk = sims["Half Kelly"]; fk = sims["Full Kelly"]

for ri, (_, bet) in enumerate(bets.iterrows(), 2):
    i = ri - 2
    bg = PERIOD_COLORS.get(bet["period"], "FFFFFF")
    fill = hex_fill(bg)
    won_fill = hex_fill("D5F5E3") if bet["won"] else hex_fill("FADBD8")

    row_data = {
        "#":                 i + 1,
        "game_date":         bet["game_date"].date(),
        "pitcher_name":      bet["pitcher_name"],
        "best_side":         bet["best_side"].upper(),
        "line":              bet["line"],
        "strikeouts_projection": round(bet["strikeouts_projection"], 2),
        "gap":               round(bet["gap"], 2),
        "strikeouts":        int(bet["strikeouts"]),
        "payout_odds":       int(bet["payout_odds"]),
        "edge_pct":          bet["edge_pct"] / 100,
        "win_prob":          bet["win_prob"],
        "kelly_raw":         bet["kelly_raw"],
        "won":               "WIN" if bet["won"] else "LOSS",
        None:                None,
        "_f1_stake":         f1.iloc[i]["stake"],
        "_f1_profit":        f1.iloc[i]["profit"],
        "_f1_bk":            f1.iloc[i]["bk"],
        None: None,
        "_f2_stake":         f2.iloc[i]["stake"],
        "_f2_profit":        f2.iloc[i]["profit"],
        "_f2_bk":            f2.iloc[i]["bk"],
        None: None,
        "_qk_pct":           bets.iloc[i]["kelly_capped"] * 0.25,
        "_qk_stake":         qk.iloc[i]["stake"],
        "_qk_profit":        qk.iloc[i]["profit"],
        "_qk_bk":            qk.iloc[i]["bk"],
        None: None,
        "_hk_pct":           bets.iloc[i]["kelly_capped"] * 0.50,
        "_hk_stake":         hk.iloc[i]["stake"],
        "_hk_profit":        hk.iloc[i]["profit"],
        "_hk_bk":            hk.iloc[i]["bk"],
        None: None,
        "_fk_pct":           bets.iloc[i]["kelly_capped"] * 1.00,
        "_fk_stake":         fk.iloc[i]["stake"],
        "_fk_profit":        fk.iloc[i]["profit"],
        "_fk_bk":            fk.iloc[i]["bk"],
    }

    for ci, (_, field, fmt) in enumerate(cols, 1):
        if field is None:
            c = ws3.cell(row=ri, column=ci, value="")
            c.fill = hex_fill("2C3E50")
            continue
        val = row_data.get(field)
        c = ws3.cell(row=ri, column=ci, value=val)
        c.fill = won_fill if field == "won" else fill
        if field == "won":
            c.font = Font(bold=True, color="155724" if bet["won"] else "922B21")
            c.alignment = Alignment(horizontal="center")
        if fmt:
            c.number_format = fmt
        c.border = Border(bottom=Side(style="hair", color="CCCCCC"))

ws3.freeze_panes = "C2"
ws3.row_dimensions[1].height = 28
# rough widths
widths = [4,12,22,6,5,10,5,7,6,7,8,8,6,  # base cols
          1,7,8,12, 1,7,8,12, 1,7,7,8,12, 1,7,7,8,12, 1,7,7,8,12]
for ci, w in enumerate(widths, 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w


# ── TAB 4: BANKROLL CURVES (chart data) ──────────────────────────────────────
ws4 = wb.create_sheet("Bankroll Curves")

headers = ["Bet #", "Date", "Flat 1%", "Flat 2%", "Qtr Kelly", "Half Kelly", "Full Kelly"]
for ci, h in enumerate(headers, 1):
    hdr_style(ws4, 1, ci, h, "1F3864")

# row 2 = starting point (bet 0)
ws4.cell(row=2, column=1, value=0)
ws4.cell(row=2, column=2, value="Start")
for ci in range(3, 8): ws4.cell(row=2, column=ci, value=STARTING)

for i, (_, bet) in enumerate(bets.iterrows(), 1):
    ri = i + 2
    ws4.cell(row=ri, column=1, value=i)
    ws4.cell(row=ri, column=2, value=bet["game_date"].strftime("%m/%d"))
    ws4.cell(row=ri, column=3, value=round(f1.iloc[i-1]["bk"], 2))
    ws4.cell(row=ri, column=4, value=round(f2.iloc[i-1]["bk"], 2))
    ws4.cell(row=ri, column=5, value=round(qk.iloc[i-1]["bk"], 2))
    ws4.cell(row=ri, column=6, value=round(hk.iloc[i-1]["bk"], 2))
    ws4.cell(row=ri, column=7, value=round(fk.iloc[i-1]["bk"], 2))

# build chart
chart = LineChart()
chart.title    = "Bankroll Growth by Strategy  (Walk-Forward 2026, $1,000 start)"
chart.style    = 10
chart.y_axis.title = "Bankroll ($)"
chart.x_axis.title = "Bet #"
chart.width    = 24
chart.height   = 14
chart.y_axis.numFmt = '"$"#,##0'

total_rows = N + 3
for col_i, name, color in [
        (3, "Flat 1%",    "4472C4"),
        (4, "Flat 2%",    "ED7D31"),
        (5, "Qtr Kelly",  "FFC000"),
        (6, "Half Kelly", "70AD47"),
        (7, "Full Kelly", "FF0000"),
]:
    data = Reference(ws4, min_col=col_i, min_row=1, max_row=total_rows)
    chart.add_data(data, titles_from_data=True)
    chart.series[-1].graphicalProperties.line.solidFill = color
    chart.series[-1].graphicalProperties.line.width = 20000 if "Kelly" in name else 15000

# x-axis labels
cats = Reference(ws4, min_col=1, min_row=2, max_row=total_rows)
chart.set_categories(cats)
ws4.add_chart(chart, "H2")

auto_width(ws4)


# ── SAVE ─────────────────────────────────────────────────────────────────────
out = "data/exports/sizing_simulation_2026.xlsx"
wb.save(out)
print(f"Saved -> {out}")
print()
print("STRATEGY SUMMARY")
print(f"{'':24} {'Flat 1%':>12} {'Flat 2%':>12} {'Half Kelly':>12} {'Full Kelly':>12}")
print("-" * 78)
for metric in metric_names:
    row = f"{metric:<24}"
    for s in strat_names:
        v = summaries[s][metric]
        if metric in ("Win Rate","Bankroll Growth","ROI on Staked","Max Drawdown (%)"):
            row += f"{v:>12.1%}"
        elif metric in ("Total Staked ($)","Total Profit ($)","Final Bankroll($)","Max Drawdown ($)"):
            row += f"{v:>12,.0f}"
        elif metric == "Sharpe":
            row += f"{v:>12.2f}"
        else:
            row += f"{v:>12}"
    print(row)

"""Export picks to a polished Excel workbook with dashboard, bankroll chart, and color-coded bet log."""
from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, GradientFill, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
UNIT            = 100
MIN_EDGE        = 7.0
MAX_STAKE       = 300
START_BANKROLL  = 10_000

# ── Palette ───────────────────────────────────────────────────────────────────
NAVY        = "0D1F3C"
GOLD        = "C9A84C"
LIGHT_GOLD  = "F5E6C8"
GREEN_DARK  = "1E6F3A"
GREEN_MID   = "27AE60"
GREEN_LIGHT = "D5F5E3"
RED_DARK    = "922B21"
RED_LIGHT   = "FADBD8"
WHITE       = "FFFFFF"
OFF_WHITE   = "F8F9FA"
GREY_LIGHT  = "E8ECEF"
GREY_MED    = "BDC3C7"
YELLOW_SOFT = "FEF9E7"
TEAL_SOFT   = "EBF5FB"

thin  = Side(style="thin",   color="D5D8DC")
thick = Side(style="medium", color="BDC3C7")

def _fill(c):  return PatternFill("solid", fgColor=c)
def _font(bold=False, color="1A1A2E", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def _border(left=thin, right=thin, top=thin, bottom=thin):
    return Border(left=left, right=right, top=top, bottom=bottom)
def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def _left():
    return Alignment(horizontal="left", vertical="center")


def _prep_data(raw_csv: str) -> pd.DataFrame:
    df = pd.read_csv(raw_csv)
    df["edge_pct"] = pd.to_numeric(df["edge_pct"], errors="coerce")
    df = df[df["market"] == "strikeouts"].copy() if "market" in df.columns else df

    if "projection" not in df.columns and "strikeouts_projection" in df.columns:
        df["projection"] = df["strikeouts_projection"]
    if "actual" not in df.columns and "strikeouts" in df.columns:
        df["actual"] = df["strikeouts"]
    if "bet_odds" not in df.columns:
        df["bet_odds"] = np.where(df["best_side"] == "over", df["over_odds"], df["under_odds"])
    if "won" not in df.columns:
        df["won"] = np.where(
            df["best_side"] == "over",
            df["actual"] > df["line"],
            df["actual"] < df["line"],
        )

    df = df[df["edge_pct"] >= MIN_EDGE].copy()
    df = (
        df.sort_values("edge_pct", ascending=False)
        .drop_duplicates(subset=["game_date", "pitcher_name", "line", "best_side"])
        .sort_values("game_date")
        .reset_index(drop=True)
    )

    # Staking: $10 per 1% edge, capped at MAX_STAKE
    df["units"] = (df["edge_pct"] / 10).clip(upper=MAX_STAKE / UNIT).round(2)
    df["stake"] = (df["units"] * UNIT).round(2)

    odds = df["bet_odds"].astype(float)
    df["decimal_odds"] = np.where(odds > 0, 1 + odds / 100, 1 + 100 / odds.abs())

    bank = float(START_BANKROLL)
    profits, banks = [], []
    for _, row in df.iterrows():
        stake = float(row["stake"])
        won   = bool(row["won"])
        pnl   = round(stake * (float(row["decimal_odds"]) - 1), 2) if won else -stake
        bank  = round(bank + pnl, 2)
        profits.append(pnl)
        banks.append(bank)

    df["profit"]           = profits
    df["rolling_bankroll"] = banks
    df["return_pct"]       = ((df["rolling_bankroll"] - START_BANKROLL) / START_BANKROLL * 100).round(1)
    df["game_date"]        = pd.to_datetime(df["game_date"]).dt.date
    return df


# ── Sheet 1: Dashboard ────────────────────────────────────────────────────────
def _write_dashboard(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    # Set column widths
    col_widths = [2, 18, 18, 18, 18, 18, 18, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(1, 50):
        ws.row_dimensions[r].height = 18

    # ── Title bar ────────────────────────────────────────────────────────────
    ws.merge_cells("B2:G3")
    tc = ws["B2"]
    tc.value = "MLB PITCHER STRIKEOUTS — 2026 MODEL PERFORMANCE"
    tc.font  = Font(bold=True, color=GOLD, size=18, name="Calibri")
    tc.fill  = _fill(NAVY)
    tc.alignment = _center()
    for col in range(2, 8):
        for row in range(2, 4):
            ws.cell(row=row, column=col).fill = _fill(NAVY)

    # subtitle
    ws.merge_cells("B4:G4")
    sc = ws["B4"]
    sc.value = f"Bets with 7%+ model edge  |  $10 per 1% edge staking  |  Starting bankroll ${START_BANKROLL:,}"
    sc.font  = Font(italic=True, color=GREY_MED, size=10, name="Calibri")
    sc.fill  = _fill(NAVY)
    sc.alignment = _center()

    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 4
    ws.row_dimensions[4].height = 20

    # ── KPI cards (row 6-9) ───────────────────────────────────────────────────
    n_bets       = len(df)
    n_wins       = int(df["won"].sum())
    win_rate     = n_wins / n_bets if n_bets else 0
    total_staked = df["stake"].sum()
    total_profit = df["profit"].sum()
    roi          = total_profit / total_staked * 100 if total_staked else 0
    final_bank   = df["rolling_bankroll"].iloc[-1] if n_bets else START_BANKROLL
    max_dd       = 0.0
    peak = START_BANKROLL
    for b in df["rolling_bankroll"]:
        if b > peak: peak = b
        dd = peak - b
        if dd > max_dd: max_dd = dd
    streak_max = 0
    streak_cur = 0
    for w in df["won"]:
        if w: streak_cur += 1; streak_max = max(streak_max, streak_cur)
        else: streak_cur = 0

    kpis = [
        ("TOTAL BETS",    str(n_bets),              NAVY,       WHITE),
        ("WIN RATE",      f"{win_rate:.1%}",         GREEN_DARK, WHITE),
        ("ROI",           f"{roi:+.1f}%",            GREEN_DARK if roi >= 0 else RED_DARK, WHITE),
        ("TOTAL PROFIT",  f"${total_profit:+,.0f}",  GREEN_DARK if total_profit >= 0 else RED_DARK, WHITE),
        ("FINAL BANKROLL",f"${final_bank:,.0f}",     NAVY,       GOLD),
        ("MAX DRAWDOWN",  f"-${max_dd:,.0f}",        RED_DARK,   WHITE),
    ]

    ws.row_dimensions[6].height = 14
    ws.row_dimensions[7].height = 30
    ws.row_dimensions[8].height = 34
    ws.row_dimensions[9].height = 14

    kpi_cols = [2, 3, 4, 5, 6, 7]
    for col_i, (label, value, bg, fg) in zip(kpi_cols, kpis):
        col_letter = get_column_letter(col_i)
        # label cell
        lc = ws.cell(row=7, column=col_i, value=label)
        lc.font      = Font(bold=True, color=GREY_MED, size=8, name="Calibri")
        lc.fill      = _fill(bg)
        lc.alignment = _center()
        # value cell
        vc = ws.cell(row=8, column=col_i, value=value)
        vc.font      = Font(bold=True, color=fg, size=16, name="Calibri")
        vc.fill      = _fill(bg)
        vc.alignment = _center()

    # ── Date range strip ─────────────────────────────────────────────────────
    ws.row_dimensions[10].height = 22
    ws.merge_cells("B10:G10")
    dr = ws["B10"]
    start_d = df["game_date"].min()
    end_d   = df["game_date"].max()
    dr.value = f"{start_d}  →  {end_d}"
    dr.font  = Font(italic=True, color=NAVY, size=10, name="Calibri")
    dr.fill  = _fill(LIGHT_GOLD)
    dr.alignment = _center()

    # ── Edge bucket breakdown (rows 12-18) ───────────────────────────────────
    ws.row_dimensions[12].height = 20
    ws.merge_cells("B12:G12")
    eh = ws["B12"]
    eh.value     = "PERFORMANCE BY EDGE RANGE"
    eh.font      = Font(bold=True, color=WHITE, size=11, name="Calibri")
    eh.fill      = _fill(NAVY)
    eh.alignment = _center()

    bucket_headers = ["Edge Range", "Bets", "Wins", "Win Rate", "Staked", "Profit", "ROI"]
    bucket_cols    = [2, 3, 4, 5, 6, 7]
    ws.row_dimensions[13].height = 18
    for ci, (col, hdr) in enumerate(zip(bucket_cols, bucket_headers)):
        c = ws.cell(row=13, column=col, value=hdr)
        c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
        c.fill = _fill("2C3E50")
        c.alignment = _center()
        c.border = _border()

    buckets = [
        ("7 – 9%",   7,  9,  "FFF9C4", "000000"),
        ("9 – 12%",  9,  12, "C8E6C9", "000000"),
        ("12%+",     12, 999,"1B5E20", "FFFFFF"),
        ("7%+ TOTAL",7,  999,"0D47A1", "FFFFFF"),
    ]
    for ri, (label, lo, hi, bg, fg) in enumerate(buckets, start=14):
        sub = df[(df["edge_pct"] >= lo) & (df["edge_pct"] < hi)]
        ws.row_dimensions[ri].height = 18
        if sub.empty:
            continue
        b = len(sub); w = int(sub["won"].sum())
        stk = sub["stake"].sum(); pnl = sub["profit"].sum()
        wr  = w / b * 100; roi_v = pnl / stk * 100 if stk else 0
        vals = [label, b, w, f"{wr:.1f}%", f"${stk:,.0f}", f"${pnl:+,.0f}", f"{roi_v:+.1f}%"]
        for col, v in zip(bucket_cols, vals):
            c = ws.cell(row=ri, column=col, value=v)
            c.font = Font(bold=(label.endswith("TOTAL")), color=fg, size=9, name="Calibri")
            c.fill = _fill(bg)
            c.alignment = _center()
            c.border = _border()

    # ── Monthly summary (rows 20-) ────────────────────────────────────────────
    df2 = df.copy()
    df2["month"] = pd.to_datetime(df2["game_date"]).dt.to_period("M").astype(str)
    monthly = (
        df2.groupby("month")
        .agg(bets=("won","count"), wins=("won","sum"),
             staked=("stake","sum"), profit=("profit","sum"))
        .reset_index()
    )
    monthly["wr"]  = (monthly["wins"]  / monthly["bets"]  * 100).round(1)
    monthly["roi"] = (monthly["profit"] / monthly["staked"] * 100).round(1)

    start_row = 20
    ws.row_dimensions[start_row].height = 20
    ws.merge_cells(f"B{start_row}:G{start_row}")
    mh = ws[f"B{start_row}"]
    mh.value = "MONTHLY BREAKDOWN"
    mh.font  = Font(bold=True, color=WHITE, size=11, name="Calibri")
    mh.fill  = _fill(NAVY)
    mh.alignment = _center()

    m_headers = ["Month", "Bets", "Wins", "Win Rate", "Staked", "Profit", "ROI"]
    ws.row_dimensions[start_row + 1].height = 18
    for ci, (col, hdr) in enumerate(zip(bucket_cols, m_headers)):
        c = ws.cell(row=start_row+1, column=col, value=hdr)
        c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
        c.fill = _fill("2C3E50")
        c.alignment = _center()
        c.border = _border()

    for ri, row in enumerate(monthly.itertuples(), start=start_row+2):
        ws.row_dimensions[ri].height = 17
        is_pos = row.roi >= 0
        row_bg = "E8F8E8" if is_pos else "FDECEA"
        vals = [
            row.month, row.bets, int(row.wins),
            f"{row.wr:.1f}%", f"${row.staked:,.0f}",
            f"${row.profit:+,.0f}", f"{row.roi:+.1f}%"
        ]
        for col, v in zip(bucket_cols, vals):
            c = ws.cell(row=ri, column=col, value=v)
            c.fill = _fill(row_bg)
            c.border = _border()
            c.alignment = _center()
            c.font = Font(
                bold=(col == 7),
                color=GREEN_DARK if (col == 7 and is_pos) else (RED_DARK if col == 7 else "1A1A2E"),
                size=9, name="Calibri"
            )


# ── Sheet 2: Bet Log ──────────────────────────────────────────────────────────
def _write_bet_log(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Bet Log")
    ws.sheet_view.showGridLines = False

    col_widths = [13, 22, 7, 7, 9, 7, 7, 9, 9, 7, 10, 9, 8, 11, 13, 11]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    headers = [
        "Date", "Pitcher", "Team", "Opp",
        "Proj Ks", "Line", "Side", "Odds",
        "Edge %", "Units", "Stake $",
        "Actual Ks", "Result", "P&L $", "Bankroll $", "Return %",
    ]
    col_keys = [
        "game_date", "pitcher_name", "team", "opponent",
        "projection", "line", "best_side", "bet_odds",
        "edge_pct", "units", "stake",
        "actual", "won", "profit", "rolling_bankroll", "return_pct",
    ]

    # header row
    ws.row_dimensions[1].height = 26
    for ci, hdr in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
        c.fill      = _fill(NAVY)
        c.alignment = _center()
        c.border    = _border(left=thick, right=thick, top=thick, bottom=thick)
    ws.freeze_panes = "A2"

    for row_i, (_, row) in enumerate(df.iterrows(), start=2):
        won  = bool(row["won"])
        edge = float(row["edge_pct"])
        pnl  = float(row["profit"])
        rb   = float(row["rolling_bankroll"])

        if won:
            row_bg = GREEN_LIGHT
        elif row_i % 2 == 0:
            row_bg = GREY_LIGHT
        else:
            row_bg = OFF_WHITE

        ws.row_dimensions[row_i].height = 17

        for ci, key in enumerate(col_keys, 1):
            val = row[key]
            if isinstance(val, (np.bool_,)):
                val = bool(val)

            # Format values
            if key == "won":
                disp = "✓ WIN" if val else "✗ LOSS"
            elif key == "best_side":
                disp = str(val).upper()
            elif key in ("projection", "line", "actual"):
                disp = round(float(val), 1) if pd.notna(val) else ""
            elif key == "edge_pct":
                disp = round(float(val), 1)
            elif key == "units":
                disp = round(float(val), 2)
            elif key in ("stake", "profit", "rolling_bankroll"):
                disp = round(float(val), 2)
            elif key == "return_pct":
                disp = round(float(val), 1)
            elif key == "bet_odds":
                v = int(round(float(val)))
                disp = f"+{v}" if v > 0 else str(v)
            else:
                disp = val

            c = ws.cell(row=row_i, column=ci, value=disp)
            c.fill      = _fill(row_bg)
            c.border    = _border()
            c.alignment = _center()
            c.font      = _font(size=9)

            # Special cell styling
            if key == "edge_pct":
                if edge >= 12:
                    c.fill = _fill(GREEN_DARK); c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
                elif edge >= 9:
                    c.fill = _fill("A9DFBF");   c.font = Font(bold=True, color=GREEN_DARK, size=9, name="Calibri")
                else:
                    c.fill = _fill("FDFDE7");   c.font = Font(bold=True, color="7D6608", size=9, name="Calibri")

            elif key == "won":
                if won:
                    c.fill = _fill(GREEN_MID); c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
                else:
                    c.fill = _fill("E74C3C");  c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")

            elif key == "profit":
                c.font = Font(bold=True, color=GREEN_DARK if pnl >= 0 else RED_DARK, size=9, name="Calibri")

            elif key == "rolling_bankroll":
                c.font = Font(bold=True, color=GREEN_DARK if rb >= START_BANKROLL else RED_DARK, size=9, name="Calibri")

            elif key == "return_pct":
                rv = float(row["return_pct"])
                c.font = Font(bold=True, color=GREEN_DARK if rv >= 0 else RED_DARK, size=9, name="Calibri")


# ── Sheet 3: Bankroll Chart ───────────────────────────────────────────────────
def _write_chart_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Bankroll Chart")
    ws.sheet_view.showGridLines = False

    # Write data for chart (hidden cols)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws["A1"] = "Bet #"
    ws["B1"] = "Bankroll"
    for i, val in enumerate(df["rolling_bankroll"], start=2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=round(float(val), 2))

    n = len(df) + 1

    chart = LineChart()
    chart.title       = "Bankroll Progression"
    chart.style       = 10
    chart.y_axis.title = "Bankroll ($)"
    chart.x_axis.title = "Bet Number"
    chart.height      = 14
    chart.width       = 28

    data = Reference(ws, min_col=2, min_row=1, max_row=n)
    chart.add_data(data, titles_from_data=True)
    chart.series[0].graphicalProperties.line.solidFill = "1F6FEB"
    chart.series[0].graphicalProperties.line.width = 18000

    # Baseline
    base_data = [[START_BANKROLL]] * (n - 1)
    ws["D1"] = "Baseline"
    for i, v in enumerate(base_data, start=2):
        ws.cell(row=i, column=4, value=START_BANKROLL)
    base_ref = Reference(ws, min_col=4, min_row=1, max_row=n)
    chart.add_data(base_ref, titles_from_data=True)
    chart.series[1].graphicalProperties.line.solidFill = "FF4444"
    chart.series[1].graphicalProperties.line.dashDot = "dash"
    chart.series[1].graphicalProperties.line.width = 12000

    ws.add_chart(chart, "F2")

    # Title label
    ws.merge_cells("F1:P1")
    t = ws["F1"]
    t.value     = f"Bankroll Growth  |  Starting ${START_BANKROLL:,}  →  Final ${df['rolling_bankroll'].iloc[-1]:,.0f}"
    t.font      = Font(bold=True, color=NAVY, size=12, name="Calibri")
    t.alignment = _center()


# ── Main ──────────────────────────────────────────────────────────────────────
def build_excel(raw_csv: str, out_path: str) -> None:
    df = _prep_data(raw_csv)

    wb = Workbook()
    _write_dashboard(wb, df)
    _write_bet_log(wb, df)
    _write_chart_sheet(wb, df)

    # Tab colors
    wb["Dashboard"].sheet_properties.tabColor       = NAVY[2:] if NAVY.startswith("FF") else NAVY
    wb["Bet Log"].sheet_properties.tabColor         = "27AE60"
    wb["Bankroll Chart"].sheet_properties.tabColor  = "1F6FEB"

    # Make Dashboard active
    wb.active = wb["Dashboard"]

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    final = df["rolling_bankroll"].iloc[-1]
    roi   = (df["profit"].sum() / df["stake"].sum() * 100) if df["stake"].sum() > 0 else 0
    print(f"Saved: {out_path}")
    print(f"  {len(df)} bets  |  {df['won'].mean():.1%} win rate  |  {roi:+.1f}% ROI  |  ${START_BANKROLL:,} to ${final:,.0f}")


if __name__ == "__main__":
    raw_csv = sys.argv[1] if len(sys.argv) > 1 else "data/exports/2026_all_picks_raw_v2.csv"
    out_xls = sys.argv[2] if len(sys.argv) > 2 else "data/exports/2026_picks_board_v3.xlsx"
    build_excel(raw_csv, out_xls)

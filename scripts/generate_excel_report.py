#!/usr/bin/env python3
"""
Generate a professional Excel performance workbook for the V4 pitcher strikeouts model.
Output: reports/performance/V4_Performance_Workbook.xlsx
"""

import sys
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

# ── Constants ──────────────────────────────────────────────────────────────────

STARTING_BANKROLL  = 10_000.0
FLAT_STAKE         = 100.0
BANKROLL_PCT_STAKE = 0.01          # 1 % of current bankroll
V4_MIN_EGP         = 6.0
V2_CORE_THRESH     = 12.0
SKIP_MONTHS        = {6}
KELLY_CAP          = 0.20          # cap Kelly fraction at 20 %

OUT_DIR   = Path("reports/performance")
OUT_FILE  = OUT_DIR / "V4_Performance_Workbook.xlsx"

# ── Color palette ──────────────────────────────────────────────────────────────

HDR_DARK    = "1F3864"
HDR_MID     = "2E75B6"
HDR_LIGHT   = "BDD7EE"
WIN_BG      = "C6EFCE"
WIN_FG      = "276221"
LOSS_BG     = "FFC7CE"
LOSS_FG     = "9C0006"
PUSH_BG     = "EDEDED"
ALT_BG      = "F2F7FF"
V2_BG       = "DEEAF1"
V4EX_BG     = "E2EFDA"
GOLD        = "FFD700"
ORANGE_BG   = "FCE4D6"
RED_BG      = "FF0000"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
MED_GRAY    = "D9D9D9"

# ── Style helpers ──────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=None, size=11, name="Calibri"):
    return Font(bold=bold, color=color or "000000", size=size, name=name)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def left():
    return Alignment(horizontal="left", vertical="center")

def right():
    return Alignment(horizontal="right", vertical="center")

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="666666")
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(ws, row, col, text, width=None, bg=HDR_DARK, fg=WHITE):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(bold=True, color=fg, size=10)
    c.alignment = center()
    c.border = thin_border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def kpi_card(ws, row, col, label, value, bg=HDR_DARK, label_size=9, val_size=14):
    """Write a 2-row KPI card (label on row, value on row+1)."""
    lc = ws.cell(row=row, column=col, value=label)
    lc.fill = fill(bg)
    lc.font = font(bold=False, color="BDD7EE", size=label_size)
    lc.alignment = center()

    vc = ws.cell(row=row + 1, column=col, value=value)
    vc.fill = fill(bg)
    vc.font = font(bold=True, color=WHITE, size=val_size)
    vc.alignment = center()
    return vc

# ── Financial helpers ──────────────────────────────────────────────────────────

def amer_to_dec(odds):
    """American to decimal odds."""
    if pd.isna(odds) or odds == 0:
        return np.nan
    if odds > 0:
        return odds / 100 + 1
    return 100 / abs(odds) + 1

def implied_prob(odds):
    if pd.isna(odds) or odds == 0:
        return np.nan
    if odds > 0:
        return 100 / (odds + 100)
    return abs(odds) / (abs(odds) + 100)

def calc_pnl(row, stake):
    """Profit/loss given a stake."""
    if pd.isna(row["won"]) or pd.isna(row["odds_used"]):
        return np.nan
    dec = amer_to_dec(row["odds_used"])
    if pd.isna(dec):
        return np.nan
    return stake * (dec - 1) if row["won"] == 1.0 else -stake

def kelly_fraction(edge_pct, odds):
    """Return the full Kelly fraction of bankroll (uncapped)."""
    if pd.isna(edge_pct) or pd.isna(odds) or odds == 0:
        return 0.0
    if odds > 0:
        b = odds / 100
    else:
        b = 100 / abs(odds)
    kf = (edge_pct / 100.0) / b
    return max(0.0, kf)

def running_stats(df, stake_col="stake_a"):
    """Add running bankroll, P&L, units, win%, ROI to df (sorted by date)."""
    df = df.copy().reset_index(drop=True)
    bankroll_a = STARTING_BANKROLL
    bankroll_b = STARTING_BANKROLL
    bankroll_c = STARTING_BANKROLL
    bankroll_d = STARTING_BANKROLL

    rows_run = []
    for i, row in df.iterrows():
        # Strategy A: flat $100
        st_a = FLAT_STAKE
        pnl_a = calc_pnl(row, st_a)

        # Strategy B: 1% bankroll
        st_b = bankroll_b * BANKROLL_PCT_STAKE
        pnl_b = calc_pnl(row, st_b)

        # Strategy C: quarter Kelly
        kf = min(kelly_fraction(row.get("edge_pct", 0), row["odds_used"]), KELLY_CAP)
        st_c = bankroll_c * kf * 0.25
        pnl_c = calc_pnl(row, st_c)

        # Strategy D: half Kelly
        st_d = bankroll_d * kf * 0.5
        pnl_d = calc_pnl(row, st_d)

        bankroll_a += (pnl_a or 0)
        bankroll_b += (pnl_b or 0)
        bankroll_c += (pnl_c or 0)
        bankroll_d += (pnl_d or 0)

        rows_run.append({
            "stake_a": st_a, "pnl_a": pnl_a, "bankroll_a": bankroll_a,
            "stake_b": st_b, "pnl_b": pnl_b, "bankroll_b": bankroll_b,
            "stake_c": st_c, "pnl_c": pnl_c, "bankroll_c": bankroll_c,
            "stake_d": st_d, "pnl_d": pnl_d, "bankroll_d": bankroll_d,
        })

    run_df = pd.DataFrame(rows_run)
    df = pd.concat([df.reset_index(drop=True), run_df], axis=1)

    # Running accumulators (strategy A as primary for display)
    df["cum_pnl"]     = df["pnl_a"].cumsum()
    df["cum_bets"]    = range(1, len(df) + 1)
    df["cum_wins"]    = (df["won"] == 1.0).cumsum()
    df["run_win_pct"] = df["cum_wins"] / df["cum_bets"]
    df["run_roi"]     = df["cum_pnl"] / (df["cum_bets"] * FLAT_STAKE)
    df["run_units"]   = df["cum_pnl"] / FLAT_STAKE

    return df

def drawdown_series(bankroll_series, start=STARTING_BANKROLL):
    """Return drawdown % series (0 to -X)."""
    peak = pd.Series(bankroll_series).cummax()
    peak = peak.where(peak > 0, start)
    dd = (bankroll_series - peak) / peak
    return dd

def max_streak(wins):
    """Return (max_win_streak, max_loss_streak) from boolean win series."""
    ws = ls = cur_w = cur_l = 0
    for w in wins:
        if w:
            cur_w += 1; cur_l = 0
        else:
            cur_l += 1; cur_w = 0
        ws = max(ws, cur_w); ls = max(ls, cur_l)
    return ws, ls

# ── Data loading ───────────────────────────────────────────────────────────────

def load_and_normalize():
    """Load all source files and return a unified, enriched DataFrame."""

    def _load(path, tag):
        df = pd.read_csv(path)
        df["source"] = tag
        return df

    df25   = _load("data/exports/2025_backtest.csv", "2025_bt")
    df26   = _load("data/exports/2026_backtest_extended.csv", "2026_bt")
    df_log = _load("data/exports/picks_log.csv", "picks_log")

    # Rename to common schema
    for df in [df25, df26, df_log]:
        df.rename(columns={
            "best_side": "side",
            "strikeouts_projection": "projection",
            "game_date": "date",
        }, inplace=True)
        df["date"] = pd.to_datetime(df["date"])
        df["month"] = df["date"].dt.month
        df["year"]  = df["date"].dt.year

    # picks_log gap is signed (proj - line); backtest gaps are already absolute
    df_log["abs_gap"] = df_log["gap"].abs()
    df25["abs_gap"]   = df25["gap"].abs()
    df26["abs_gap"]   = df26["gap"].abs()

    # edge_gap_product
    for df in [df25, df26, df_log]:
        df["egp"] = df["abs_gap"] * df["edge_pct"]

    # Strategy bucket
    def bucket(egp):
        if egp >= V2_CORE_THRESH:   return "V2_core"
        if egp >= V4_MIN_EGP:       return "V4_extra"
        return "Filtered"

    for df in [df25, df26, df_log]:
        df["bucket"] = df["egp"].apply(bucket)

    # Decimal / implied prob
    for df in [df25, df26, df_log]:
        df["dec_odds"]    = df["odds_used"].apply(amer_to_dec)
        df["mkt_prob"]    = df["odds_used"].apply(implied_prob)

    # pnl at $100 flat
    for df in [df25, df26, df_log]:
        if "pnl" not in df.columns:
            df["pnl"] = df.apply(lambda r: calc_pnl(r, FLAT_STAKE), axis=1)

    # result label
    def result_label(row):
        if pd.isna(row.get("won")):   return "Pending"
        return "Win" if row["won"] == 1.0 else "Loss"

    for df in [df25, df26, df_log]:
        df["result_label"] = df.apply(result_label, axis=1)

    # CLV – normalise column name
    for df in [df26, df_log]:
        if "clv_pct" in df.columns:
            df["clv"] = df["clv_pct"]
    if "clv" not in df25.columns:
        df25["clv"] = np.nan

    # opening / closing odds normalise
    for df in [df25]:
        if "opening_odds" not in df.columns:
            df["opening_odds"] = np.nan
        if "closing_odds" not in df.columns:
            df["closing_odds"] = np.nan

    # ── Assign to sheets ───────────────────────────────────────────────────────
    # 2025: all non-June V4 bets from 2025 backtest
    sh_2025 = df25[(df25["egp"] >= V4_MIN_EGP) & (~df25["month"].isin(SKIP_MONTHS))].copy()

    # 2026 Pre-June: March–May 2026 from 2026 backtest
    sh_2026_pre = df26[(df26["egp"] >= V4_MIN_EGP) & (~df26["month"].isin(SKIP_MONTHS))].copy()

    # June Holdout: June from picks_log (live bets)
    sh_june = df_log[df_log["month"].isin(SKIP_MONTHS)].copy()

    # Combined: deployed V4 bets only (no June)
    sh_combined = pd.concat([sh_2025, sh_2026_pre], ignore_index=True).sort_values("date")

    return sh_2025, sh_2026_pre, sh_june, sh_combined

# ── Bet sheet writer ───────────────────────────────────────────────────────────

BET_COLUMNS = [
    ("Bet ID",            "bet_id",        6,   "0"),
    ("Date",              "date",          11,  "YYYY-MM-DD"),
    ("Pitcher",           "pitcher_name",  20,  "@"),
    ("Side",              "side",           6,  "@"),
    ("Line",              "line",           6,  "0.0"),
    ("Projection",        "projection",     11, "0.00"),
    ("Abs Gap",           "abs_gap",        8,  "0.00"),
    ("Edge %",            "edge_pct",       8,  "0.00%"),
    ("EGP",               "egp",            8,  "0.00"),
    ("Bucket",            "bucket",         9,  "@"),
    ("Opening Odds",      "opening_odds",  12,  "+0;-0;0"),
    ("Executed Odds",     "odds_used",     13,  "+0;-0;0"),
    ("Closing Odds",      "closing_odds",  12,  "+0;-0;0"),
    ("CLV",               "clv",            7,  "0.00"),
    ("Dec Odds",          "dec_odds",      9,   "0.000"),
    ("Mkt Prob",          "mkt_prob",      9,   "0.0%"),
    ("Result",            "result_label",   8,  "@"),
    ("Stake ($)",         "stake_a",       10,  "$#,##0.00"),
    ("P&L ($)",           "pnl_a",         10,  "$#,##0.00;[Red]-$#,##0.00"),
    ("ROI",               "bet_roi",       8,   "0.0%"),
    ("Run Win%",          "run_win_pct",   10,  "0.0%"),
    ("Run ROI",           "run_roi",       9,   "0.0%"),
    ("Run Units",         "run_units",     10,  "+0.00;-0.00;0.00"),
    ("Run Bankroll",      "bankroll_a",    13,  "$#,##0"),
]


def write_bet_sheet(ws, df_raw, sheet_label):
    """Write a standardized bet sheet with formatting and conditional formatting."""
    df = df_raw.copy().sort_values("date").reset_index(drop=True)
    df["bet_id"] = range(1, len(df) + 1)

    df["bet_roi"] = df.apply(
        lambda r: (r["pnl_a"] / FLAT_STAKE) if not pd.isna(r.get("pnl_a")) else np.nan,
        axis=1
    )

    n_rows = len(df)
    n_cols = len(BET_COLUMNS)

    # ── Header row ─────────────────────────────────────────────────────────────
    for c_idx, (label, _, width, _) in enumerate(BET_COLUMNS, start=1):
        header_style(ws, 1, c_idx, label, width=width)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    # ── Data rows ──────────────────────────────────────────────────────────────
    for r_idx, row in df.iterrows():
        excel_row = r_idx + 2
        alt = (r_idx % 2 == 0)

        for c_idx, (_, col, _, fmt) in enumerate(BET_COLUMNS, start=1):
            val = row.get(col)

            # Convert numpy types
            if isinstance(val, (np.integer,)):  val = int(val)
            elif isinstance(val, (np.floating,)): val = None if np.isnan(val) else float(val)
            elif isinstance(val, pd.Timestamp): val = val.date()

            cell = ws.cell(row=excel_row, column=c_idx, value=val)

            if fmt != "@" and val is not None:
                cell.number_format = fmt

            cell.border = thin_border()

            # Base row color
            bucket_val = str(row.get("bucket", ""))
            result_val = str(row.get("result_label", ""))

            if result_val == "Win":
                base_bg = WIN_BG
            elif result_val == "Loss":
                base_bg = LOSS_BG
            elif alt:
                base_bg = ALT_BG
            else:
                base_bg = WHITE

            cell.fill = fill(base_bg)
            cell.alignment = center() if fmt not in ("@",) else left()

    # ── Conditional formatting ─────────────────────────────────────────────────
    data_range = f"A2:{get_column_letter(n_cols)}{n_rows + 1}"
    pnl_col_letter = get_column_letter([col for _, col, _, _ in BET_COLUMNS].index("pnl_a") + 1)

    ws.conditional_formatting.add(
        f"{pnl_col_letter}2:{pnl_col_letter}{n_rows + 1}",
        ColorScaleRule(
            start_type="min", start_color=LOSS_BG,
            mid_type="num",   mid_value=0,  mid_color="FFFFFF",
            end_type="max",   end_color=WIN_BG,
        )
    )

    roi_col_letter = get_column_letter([col for _, col, _, _ in BET_COLUMNS].index("run_roi") + 1)
    ws.conditional_formatting.add(
        f"{roi_col_letter}2:{roi_col_letter}{n_rows + 1}",
        ColorScaleRule(
            start_type="min", start_color=LOSS_BG,
            mid_type="num",   mid_value=0,  mid_color="FFFFFF",
            end_type="max",   end_color=WIN_BG,
        )
    )

    # ── Sheet info footer ──────────────────────────────────────────────────────
    footer_row = n_rows + 3
    ws.cell(row=footer_row, column=1, value=f"{sheet_label} — {n_rows} bets")
    ws.cell(row=footer_row, column=1).font = font(bold=True, size=9, color="888888")

    # ── Summary mini-table ─────────────────────────────────────────────────────
    s_row = footer_row + 1
    settled = df[df["result_label"].isin(["Win", "Loss"])]
    n_s    = len(settled)
    n_w    = (settled["won"] == 1.0).sum()
    roi    = settled["pnl_a"].sum() / (n_s * FLAT_STAKE) if n_s > 0 else 0
    final  = df["bankroll_a"].iloc[-1] if len(df) > 0 else STARTING_BANKROLL

    for text, val, fmt_str in [
        ("Settled Bets",  n_s,  "0"),
        ("Wins",          int(n_w), "0"),
        ("Win Rate",      n_w / n_s if n_s else 0, "0.0%"),
        ("ROI",           roi, "0.0%"),
        ("Total P&L",     settled["pnl_a"].sum(), "$#,##0"),
        ("Final BR",      final, "$#,##0"),
    ]:
        c = ws.cell(row=s_row, column=1, value=text)
        c.font = font(bold=True, size=9, color=HDR_DARK)
        v = ws.cell(row=s_row, column=2, value=val)
        v.number_format = fmt_str
        v.font = font(size=9)
        s_row += 1

    return df  # return enriched df for use in other sheets

# ── Monthly stats sheet ────────────────────────────────────────────────────────

def write_monthly_sheet(ws, df_combined):
    title_labels = [
        "Year", "Month", "Bets", "Wins", "Losses", "Win Rate", "ROI",
        "Units", "P&L ($)", "Avg Odds", "Avg Edge%", "Avg EGP",
        "Avg CLV", "End Bankroll",
    ]
    for c_idx, label in enumerate(title_labels, start=1):
        header_style(ws, 1, c_idx, label)
    ws.freeze_panes = "A2"

    df = df_combined[df_combined["result_label"].isin(["Win","Loss"])].copy()
    df["ym"] = df["date"].dt.to_period("M")
    groups = df.groupby("ym")

    bankroll = STARTING_BANKROLL
    r = 2
    for period, g in sorted(groups, key=lambda x: str(x[0])):
        g = g.sort_values("date")
        n    = len(g)
        wins = (g["won"] == 1.0).sum()
        pnl  = g["pnl_a"].sum()
        bankroll += pnl
        units     = pnl / FLAT_STAKE
        roi       = pnl / (n * FLAT_STAKE)

        vals = [
            period.year, period.strftime("%B"), n, int(wins), n - int(wins),
            wins / n, roi, units, pnl,
            g["odds_used"].mean(), g["edge_pct"].mean(), g["egp"].mean(),
            g["clv"].mean() if g["clv"].notna().any() else None,
            bankroll,
        ]
        fmts = [
            "0", "@", "0", "0", "0", "0.0%", "0.0%",
            "+0.00;-0.00", "$#,##0", "+0;-0", "0.00", "0.00",
            "0.00", "$#,##0",
        ]

        alt = (r % 2 == 0)
        row_bg = ALT_BG if alt else WHITE
        if roi > 0:   row_bg = WIN_BG
        elif roi < 0: row_bg = LOSS_BG

        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.number_format = fmt
            cell.alignment = center()
            cell.border = thin_border()
            cell.fill = fill(row_bg)
        r += 1

    ws.column_dimensions["B"].width = 12
    for col in "ACDEFGHIJKLN":
        ws.column_dimensions[col].width = 11

# ── Drawdown sheet ────────────────────────────────────────────────────────────

def write_drawdown_sheet(ws, df_combined):
    headers = [
        "Period #", "Peak Date", "Trough Date", "Recovery Date",
        "DD %", "DD ($)", "Duration (days)", "Bets in DD",
    ]
    for c_idx, h in enumerate(headers, start=1):
        header_style(ws, 1, c_idx, h, width=14)
    ws.freeze_panes = "A2"

    df = df_combined.sort_values("date").copy()
    bk = df["bankroll_a"].values
    dates = df["date"].values

    peak = STARTING_BANKROLL
    peak_date = dates[0] if len(dates) > 0 else None
    in_dd = False
    dd_start = None
    dd_peak = None
    dd_peak_date = None
    dd_bets = 0
    dd_min = STARTING_BANKROLL
    dd_min_date = None

    periods = []
    for i, (b, d) in enumerate(zip(bk, dates)):
        if b > peak:
            if in_dd:
                # recovery
                periods.append({
                    "peak_date":     dd_peak_date,
                    "trough_date":   dd_min_date,
                    "recovery_date": d,
                    "dd_pct":        (dd_min - dd_peak) / dd_peak,
                    "dd_usd":        dd_min - dd_peak,
                    "duration":      (pd.Timestamp(d) - pd.Timestamp(dd_peak_date)).days,
                    "bets":          dd_bets,
                })
                in_dd = False
            peak = b
            peak_date = d
        elif b < peak:
            if not in_dd:
                in_dd = True
                dd_peak = peak
                dd_peak_date = peak_date
                dd_min = b
                dd_min_date = d
                dd_bets = 1
            else:
                dd_bets += 1
                if b < dd_min:
                    dd_min = b
                    dd_min_date = d

    periods.sort(key=lambda x: x["dd_pct"])

    for p_idx, p in enumerate(periods, start=1):
        r = p_idx + 1
        dd_pct = p["dd_pct"]
        color = "FF0000" if dd_pct < -0.15 else ("FFC000" if dd_pct < -0.05 else WHITE)
        vals = [
            p_idx,
            pd.Timestamp(p["peak_date"]).date() if p["peak_date"] is not None else "",
            pd.Timestamp(p["trough_date"]).date() if p["trough_date"] is not None else "",
            pd.Timestamp(p["recovery_date"]).date() if p["recovery_date"] is not None else "Open",
            dd_pct,
            p["dd_usd"],
            p["duration"],
            p["bets"],
        ]
        fmts = ["0","YYYY-MM-DD","YYYY-MM-DD","YYYY-MM-DD","0.0%","$#,##0","0","0"]
        for c_idx, (v, f) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row=r, column=c_idx, value=v)
            cell.number_format = f
            cell.alignment = center()
            cell.border = thin_border()
            cell.fill = fill(color if color != WHITE else (ALT_BG if r % 2 == 0 else WHITE))

# ── Bankroll simulation sheet ──────────────────────────────────────────────────

def write_bankroll_sheet(ws, df_combined):
    df = df_combined.sort_values("date").reset_index(drop=True)

    # ── Note on Kelly ─────────────────────────────────────────────────────────
    ws.cell(row=1, column=1, value="Bankroll Simulation — Four Staking Strategies")
    ws.cell(row=1, column=1).font = font(bold=True, size=13, color=HDR_DARK)

    note = (
        "Kelly fractions are approximated from edge_pct ÷ net_odds. "
        "Full model probabilities are unavailable in archived backtests, so "
        "exact Kelly cannot be derived from first principles. "
        "These values are illustrative; treat with caution."
    )
    ws.cell(row=2, column=1, value=note)
    ws.cell(row=2, column=1).font = font(size=9, color="888888")
    ws.merge_cells(f"A2:{get_column_letter(18)}2")

    headers = [
        "Bet #", "Date", "Pitcher", "Side", "Odds",
        "Won",
        "Stk A",  "P&L A",  "BR A",
        "Stk B",  "P&L B",  "BR B",
        "Stk C",  "P&L C",  "BR C",
        "Stk D",  "P&L D",  "BR D",
    ]
    col_widths = [6,11,20,6,7,5, 9,10,11, 9,10,11, 9,10,11, 9,10,11]
    for c_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        header_style(ws, 3, c_idx, h, width=w, bg=HDR_DARK)
    ws.freeze_panes = "A4"

    for r_idx, row in df.iterrows():
        excel_r = r_idx + 4
        alt = r_idx % 2 == 0
        bg = ALT_BG if alt else WHITE

        vals = [
            r_idx + 1,
            row["date"].date() if hasattr(row["date"], "date") else row["date"],
            row.get("pitcher_name", ""),
            row.get("side", ""),
            row.get("odds_used"),
            "W" if row.get("won") == 1.0 else ("L" if row.get("won") == 0.0 else ""),
            row.get("stake_a"), row.get("pnl_a"), row.get("bankroll_a"),
            row.get("stake_b"), row.get("pnl_b"), row.get("bankroll_b"),
            row.get("stake_c"), row.get("pnl_c"), row.get("bankroll_c"),
            row.get("stake_d"), row.get("pnl_d"), row.get("bankroll_d"),
        ]
        fmts = [
            "0","YYYY-MM-DD","@","@","+0;-0",
            "@",
            "$#,##0","$#,##0.00;[Red]-$#,##0.00","$#,##0",
            "$#,##0","$#,##0.00;[Red]-$#,##0.00","$#,##0",
            "$#,##0","$#,##0.00;[Red]-$#,##0.00","$#,##0",
            "$#,##0","$#,##0.00;[Red]-$#,##0.00","$#,##0",
        ]
        for c_idx, (v, f) in enumerate(zip(vals, fmts), start=1):
            if isinstance(v, float) and np.isnan(v): v = None
            cell = ws.cell(row=excel_r, column=c_idx, value=v)
            cell.number_format = f
            cell.alignment = center()
            cell.border = thin_border()
            if v == "W": cell.fill = fill(WIN_BG)
            elif v == "L": cell.fill = fill(LOSS_BG)
            else: cell.fill = fill(bg)

    # ── Strategy summary table ────────────────────────────────────────────────
    footer_r = len(df) + 6
    ws.cell(row=footer_r, column=1, value="Strategy Summary").font = font(bold=True, size=11, color=HDR_DARK)
    footer_r += 1

    strats = [
        ("A — Flat $100",         "bankroll_a", "pnl_a", "stake_a"),
        ("B — 1% Bankroll",       "bankroll_b", "pnl_b", "stake_b"),
        ("C — Quarter Kelly",     "bankroll_c", "pnl_c", "stake_c"),
        ("D — Half Kelly",        "bankroll_d", "pnl_d", "stake_d"),
    ]
    sum_hdrs = ["Strategy","Final BR","Total P&L","Total Return","Avg Stake","Max Drawdown","Peak BR"]
    for c_idx, h in enumerate(sum_hdrs, start=1):
        header_style(ws, footer_r, c_idx, h, bg=HDR_MID)
    footer_r += 1

    settled = df[df["result_label"].isin(["Win","Loss"])]
    for name, br_col, pnl_col, stk_col in strats:
        if br_col not in df.columns: continue
        final_br  = df[br_col].iloc[-1] if len(df) > 0 else STARTING_BANKROLL
        total_pnl = df[pnl_col].sum() if pnl_col in df.columns else 0
        total_ret = (final_br - STARTING_BANKROLL) / STARTING_BANKROLL
        avg_stake = df[stk_col].mean() if stk_col in df.columns else 0
        peak_br   = df[br_col].max() if br_col in df.columns else STARTING_BANKROLL
        max_dd    = drawdown_series(df[br_col]).min()

        vals = [name, final_br, total_pnl, total_ret, avg_stake, max_dd, peak_br]
        fmts = ["@","$#,##0","$#,##0","0.0%","$#,##0","0.0%","$#,##0"]
        for c_idx, (v, f) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row=footer_r, column=c_idx, value=v)
            cell.number_format = f
            cell.alignment = center()
            cell.border = thin_border()
            cell.fill = fill(WIN_BG if (total_pnl or 0) > 0 else LOSS_BG)
        footer_r += 1

# ── Dashboard sheet ────────────────────────────────────────────────────────────

def write_dashboard(ws, df_combined, sh_2025, sh_2026_pre, sh_june):
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:P2")
    title = ws["A1"]
    title.value = "V4 Pitcher Strikeouts Model — Performance Dashboard"
    title.fill  = fill(HDR_DARK)
    title.font  = Font(bold=True, size=18, color=WHITE, name="Calibri")
    title.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    gen_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.merge_cells("A3:P3")
    sub = ws["A3"]
    sub.value = (f"Generated {gen_date}  |  "
                 f"V4: min_EGP≥6, skip June, overs+unders, $100 flat  |  "
                 f"Walk-forward cutoff 2024-12-31")
    sub.fill  = fill(HDR_MID)
    sub.font  = Font(italic=True, size=9, color=WHITE, name="Calibri")
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # ── KPI cards ─────────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[7].height = 8   # spacer

    settled_all = df_combined[df_combined["result_label"].isin(["Win","Loss"])]
    n_total     = len(settled_all)
    n_wins      = int((settled_all["won"] == 1.0).sum())
    win_pct     = n_wins / n_total if n_total else 0
    roi         = settled_all["pnl_a"].sum() / (n_total * FLAT_STAKE) if n_total else 0
    total_pnl   = settled_all["pnl_a"].sum()
    final_br    = df_combined["bankroll_a"].iloc[-1] if len(df_combined) > 0 else STARTING_BANKROLL
    avg_odds    = settled_all["odds_used"].mean()
    avg_clv     = settled_all["clv"].mean() if settled_all["clv"].notna().any() else 0
    units       = total_pnl / FLAT_STAKE
    max_dd      = drawdown_series(df_combined["bankroll_a"]).min() if len(df_combined) > 0 else 0
    ws_val, ls_val = max_streak(list(settled_all["won"] == 1.0))
    avg_edge    = settled_all["edge_pct"].mean()
    avg_egp     = settled_all["egp"].mean()
    avg_stake   = FLAT_STAKE

    kpis = [
        ("Total Bets",   n_total,  "0"),
        ("Win Rate",     win_pct,  "0.0%"),
        ("ROI",          roi,      "+0.0%;-0.0%"),
        ("Units P&L",    units,    "+0.00;-0.00"),
        ("Total Profit", total_pnl,"$#,##0"),
        ("Final BR",     final_br, "$#,##0"),
        ("Avg Odds",     avg_odds, "+0;-0"),
        ("Avg CLV",      avg_clv,  "0.00"),
        ("Max Drawdown", max_dd,   "0.0%"),
        ("Win Streak",   ws_val,   "0"),
        ("Loss Streak",  ls_val,   "0"),
        ("Avg Edge %",   avg_edge, "0.00"),
    ]

    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 26

    for col_offset, (label, val, fmt) in enumerate(kpis):
        col = col_offset + 1

        lc = ws.cell(row=5, column=col, value=label)
        lc.fill = fill(HDR_DARK)
        lc.font = Font(bold=False, color="BDD7EE", size=8, name="Calibri")
        lc.alignment = Alignment(horizontal="center", vertical="center")

        vc = ws.cell(row=6, column=col, value=val)
        vc.fill = fill(HDR_DARK)
        vc.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.number_format = fmt

        ws.column_dimensions[get_column_letter(col)].width = 13

    # ── Period breakdown mini-table ───────────────────────────────────────────
    ws.row_dimensions[8].height = 8
    ws.row_dimensions[9].height = 18
    r = 9
    period_hdrs = ["Period","Bets","Wins","Win%","ROI","P&L","Avg EGP","Avg CLV"]
    for c_idx, h in enumerate(period_hdrs, start=1):
        header_style(ws, r, c_idx, h, bg=HDR_MID, fg=WHITE)
    r += 1

    period_data = []
    for label, dset in [
        ("2025 (non-Jun)", sh_2025),
        ("2026 Pre-Jun",   sh_2026_pre),
        ("June Holdout",   sh_june),
        ("Combined (dep)", df_combined),
    ]:
        s = dset[dset["result_label"].isin(["Win","Loss"])]
        if len(s) == 0: continue
        period_data.append((
            label, len(s),
            int((s["won"] == 1.0).sum()),
            (s["won"] == 1.0).mean(),
            s["pnl_a"].sum() / (len(s) * FLAT_STAKE),
            s["pnl_a"].sum(),
            s["egp"].mean(),
            s["clv"].mean() if s["clv"].notna().any() else None,
        ))

    for i, row_data in enumerate(period_data):
        label_p, n, w, wp, roi_p, pnl_p, egp_m, clv_m = row_data
        bg = WIN_BG if roi_p > 0 else LOSS_BG
        row_vals = [label_p, n, w, wp, roi_p, pnl_p, egp_m, clv_m]
        row_fmts = ["@","0","0","0.0%","0.0%","$#,##0","0.00","0.00"]
        for c_idx, (v, f) in enumerate(zip(row_vals, row_fmts), start=1):
            cell = ws.cell(row=r, column=c_idx, value=v)
            cell.number_format = f
            cell.alignment = center()
            cell.border = thin_border()
            cell.fill = fill(bg)
        r += 1

    # ── Bucket breakdown ──────────────────────────────────────────────────────
    r += 1
    bucket_hdrs = ["Bucket","Bets","Win%","ROI","P&L","Avg Edge%","Avg EGP"]
    for c_idx, h in enumerate(bucket_hdrs, start=1):
        header_style(ws, r, c_idx, h, bg=HDR_MID)
    r += 1

    settled_all_with_june = pd.concat([df_combined, sh_june], ignore_index=True)
    settled_all_with_june = settled_all_with_june[settled_all_with_june["result_label"].isin(["Win","Loss"])]
    for bkt in ["V2_core", "V4_extra", "Filtered"]:
        g = settled_all_with_june[settled_all_with_june["bucket"] == bkt]
        if len(g) == 0: continue
        bg = V2_BG if bkt == "V2_core" else (V4EX_BG if bkt == "V4_extra" else PUSH_BG)
        vals = [
            bkt, len(g), (g["won"]==1.0).mean(),
            g["pnl_a"].sum()/(len(g)*FLAT_STAKE), g["pnl_a"].sum(),
            g["edge_pct"].mean(), g["egp"].mean(),
        ]
        fmts = ["@","0","0.0%","0.0%","$#,##0","0.00","0.00"]
        for c_idx, (v, f) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row=r, column=c_idx, value=v)
            cell.number_format = f
            cell.alignment = center()
            cell.border = thin_border()
            cell.fill = fill(bg)
        r += 1

    # ── Running P&L chart ─────────────────────────────────────────────────────
    chart_start_row = 8

    df_chart = df_combined.sort_values("date").reset_index(drop=True)
    # Write chart data to a hidden area
    chart_data_col = 14
    ws.cell(row=chart_start_row, column=chart_data_col, value="Bet#")
    ws.cell(row=chart_start_row, column=chart_data_col + 1, value="Bankroll")
    ws.cell(row=chart_start_row, column=chart_data_col + 2, value="P&L")

    for i, (_, row) in enumerate(df_chart.iterrows()):
        rr = chart_start_row + 1 + i
        ws.cell(row=rr, column=chart_data_col,     value=i + 1)
        ws.cell(row=rr, column=chart_data_col + 1, value=row.get("bankroll_a"))
        ws.cell(row=rr, column=chart_data_col + 2, value=row.get("cum_pnl"))

    n_chart = len(df_chart)

    # Bankroll chart
    br_chart = LineChart()
    br_chart.title   = "Running Bankroll (Strategy A — Flat $100)"
    br_chart.style   = 10
    br_chart.y_axis.title = "Bankroll ($)"
    br_chart.x_axis.title = "Bet Number"
    br_chart.height  = 12
    br_chart.width   = 22

    br_data = Reference(ws, min_col=chart_data_col + 1, min_row=chart_start_row,
                        max_row=chart_start_row + n_chart)
    br_chart.add_data(br_data, titles_from_data=True)
    br_chart.series[0].graphicalProperties.line.solidFill = HDR_MID
    br_chart.series[0].graphicalProperties.line.width = 15000

    ws.add_chart(br_chart, "A16")

    # Cumulative P&L chart
    pnl_chart = LineChart()
    pnl_chart.title   = "Cumulative Profit / Loss ($)"
    pnl_chart.style   = 10
    pnl_chart.y_axis.title = "P&L ($)"
    pnl_chart.height  = 12
    pnl_chart.width   = 22

    pnl_data = Reference(ws, min_col=chart_data_col + 2, min_row=chart_start_row,
                         max_row=chart_start_row + n_chart)
    pnl_chart.add_data(pnl_data, titles_from_data=True)
    pnl_chart.series[0].graphicalProperties.line.solidFill = WIN_FG

    ws.add_chart(pnl_chart, "I16")

# ── README sheet ───────────────────────────────────────────────────────────────

def write_readme(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 110

    lines = [
        ("V4 Pitcher Strikeouts Model — Workbook Guide", True, 16, HDR_DARK),
        ("", False, 11, "000000"),
        ("WORKBOOK STRUCTURE", True, 12, HDR_MID),
        ("Sheet: 2025             — V4-qualified bets from 2025 walk-forward backtest (non-June, EGP≥6)", False, 10, "000000"),
        ("Sheet: 2026 Pre-June    — V4-qualified bets from 2026 season through May 31 (EGP≥6)", False, 10, "000000"),
        ("Sheet: June Holdout     — June 2026 live bets from picks_log (model-excluded month, for reference)", False, 10, "000000"),
        ("Sheet: Combined         — All deployed V4 bets (2025 + 2026 Pre-Jun, no June)", False, 10, "000000"),
        ("Sheet: Dashboard        — KPI summary, period breakdown, bucket analysis, charts", False, 10, "000000"),
        ("Sheet: Monthly Stats    — One row per calendar month", False, 10, "000000"),
        ("Sheet: Drawdown         — Every drawdown period, sorted by severity", False, 10, "000000"),
        ("Sheet: Bankroll Sim     — Four staking strategies side-by-side", False, 10, "000000"),
        ("", False, 11, "000000"),
        ("MODEL PARAMETERS", True, 12, HDR_MID),
        ("Walk-forward cutoff:    2024-12-31  (train=2022–2024, OOS=2025–present)", False, 10, "000000"),
        ("Min edge-gap product:   6.0         (V4 threshold; V2_core=≥12, V4_extra=6–12)", False, 10, "000000"),
        ("Skip months:            June (month 6)  — confirmed underperformance 2025 & 2026", False, 10, "000000"),
        ("Direction:              Both sides (overs + unders)", False, 10, "000000"),
        ("", False, 11, "000000"),
        ("STAKING STRATEGIES", True, 12, HDR_MID),
        ("Strategy A — Flat $100 per bet.  Simple benchmark.  Starting bankroll $10,000.", False, 10, "000000"),
        ("Strategy B — 1% of current bankroll per bet.  Geometric growth, shrinks in drawdowns.", False, 10, "000000"),
        ("Strategy C — Quarter Kelly.  Kelly fraction ≈ edge_pct / net_odds, then × 0.25.", False, 10, "000000"),
        ("Strategy D — Half Kelly.  Same as C but × 0.50.  Higher variance, higher ceiling.", False, 10, "000000"),
        ("", False, 11, "000000"),
        ("KELLY CAVEAT", True, 12, "9C0006"),
        ("Exact Kelly requires the model's raw probability output, which is not stored in the backtest CSVs.", False, 10, "000000"),
        ("The approximation edge_pct ÷ net_odds is used instead.  Treat Kelly columns as illustrative.", False, 10, "000000"),
        ("If full model probabilities are needed, re-run project_daily.py with --export-probs.", False, 10, "000000"),
        ("", False, 11, "000000"),
        ("FORMULAS", True, 12, HDR_MID),
        ("P&L (win):    stake × (dec_odds − 1)", False, 10, "000000"),
        ("P&L (loss):   −stake", False, 10, "000000"),
        ("Decimal odds: (100/|amer|)+1 if amer<0, else (amer/100)+1", False, 10, "000000"),
        ("ROI:          Σ P&L / (N × $100)", False, 10, "000000"),
        ("EGP:          |projection − line| × edge_pct", False, 10, "000000"),
        ("CLV:          closing_odds_american − executed_odds_american  (positive = good)", False, 10, "000000"),
        ("Drawdown:     (current_bankroll − peak_bankroll) / peak_bankroll", False, 10, "000000"),
        ("", False, 11, "000000"),
        ("COLOR CODING", True, 12, HDR_MID),
        ("Green row:    Win", False, 10, "000000"),
        ("Red row:      Loss", False, 10, "000000"),
        ("Blue shading: V2_core bucket (EGP ≥ 12)", False, 10, "000000"),
        ("Green shade:  V4_extra bucket (6 ≤ EGP < 12)", False, 10, "000000"),
        ("Orange:       Moderate drawdown (> 5%)", False, 10, "000000"),
        ("Dark red:     Severe drawdown (> 15%)", False, 10, "000000"),
        ("", False, 11, "000000"),
        ("DATA SOURCES", True, 12, HDR_MID),
        ("data/exports/2025_backtest.csv          — 2025 walk-forward backtest output", False, 10, "000000"),
        ("data/exports/2026_backtest_extended.csv — 2026 season-to-date backtest", False, 10, "000000"),
        ("data/exports/picks_log.csv              — Live model output June 2026 (V3 pipeline)", False, 10, "000000"),
        ("", False, 11, "000000"),
        ("EXCLUDED BETS", True, 12, HDR_MID),
        ("June 2026 bets are in the 'June Holdout' sheet.  They are NOT in Combined or Dashboard totals.", False, 10, "000000"),
        ("Bets with EGP < 6.0 are excluded entirely (below V4 threshold).", False, 10, "000000"),
        ("Pending bets (no settled result) are excluded from ROI / P&L calculations.", False, 10, "000000"),
    ]

    for row_i, (text, bold, size, color) in enumerate(lines, start=1):
        cell = ws.cell(row=row_i, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color=color, name="Calibri")
        ws.row_dimensions[row_i].height = max(14, size + 4)

# ── Pivot summaries sheet ──────────────────────────────────────────────────────

def write_pivot_sheet(ws, df_combined, sh_june):
    all_bets = pd.concat([df_combined, sh_june], ignore_index=True)
    all_bets = all_bets[all_bets["result_label"].isin(["Win","Loss"])]

    def mini_pivot(df, group_col, start_col, start_row, label):
        ws.cell(row=start_row, column=start_col, value=label)
        ws.cell(row=start_row, column=start_col).font = font(bold=True, size=11, color=HDR_DARK)

        hdrs = ["Group","Bets","Wins","Win%","ROI","P&L","Avg Odds","Avg EGP"]
        for c_idx, h in enumerate(hdrs, start=start_col):
            header_style(ws, start_row + 1, c_idx, h, bg=HDR_MID)

        r = start_row + 2
        for grp_name, g in df.groupby(group_col):
            n  = len(g)
            w  = int((g["won"] == 1.0).sum())
            pnl = g["pnl_a"].sum()
            roi = pnl / (n * FLAT_STAKE)
            bg  = WIN_BG if roi > 0 else LOSS_BG
            vals = [str(grp_name), n, w, w/n, roi, pnl, g["odds_used"].mean(), g["egp"].mean()]
            fmts = ["@","0","0","0.0%","0.0%","$#,##0","+0;-0","0.00"]
            for c_idx, (v, f) in enumerate(zip(vals, fmts), start=start_col):
                cell = ws.cell(row=r, column=c_idx, value=v)
                cell.number_format = f; cell.alignment = center()
                cell.border = thin_border(); cell.fill = fill(bg)
            r += 1
        return r + 1

    # Odds bucket pivot helper
    def odds_bucket(odds):
        if pd.isna(odds): return "Unknown"
        if odds <= -160:  return "≤−160 (heavy fav)"
        if odds <= -120:  return "−120 to −160"
        if odds < 0:      return "−100 to −120"
        if odds <= 110:   return "+100 to +110"
        if odds <= 130:   return "+111 to +130"
        return "≥+130 (dog)"

    all_bets["odds_bucket"] = all_bets["odds_used"].apply(odds_bucket)
    all_bets["line_bucket"] = all_bets["line"].apply(
        lambda x: f"{x:.1f}" if not pd.isna(x) else "?"
    )

    pivots = [
        ("side",        "By Side (Over/Under)",       1,  1),
        ("bucket",      "By Strategy Bucket",          1, 13),
        ("odds_bucket", "By Odds Bucket",              1, 25),
    ]
    for col, lbl, sc, sr in pivots:
        mini_pivot(all_bets, col, sc, sr, lbl)

    # Year pivot
    all_bets["year_str"] = all_bets["year"].astype(str)
    pivots2 = [
        ("year_str",   "By Year",                     1, 45),
        ("line_bucket","By Strikeout Line",            1, 60),
    ]
    for col, lbl, sc, sr in pivots2:
        mini_pivot(all_bets, col, sc, sr, lbl)

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Loading source data...")
    sh_2025, sh_2026_pre, sh_june, sh_combined = load_and_normalize()

    print(f"  2025 deployed bets:    {len(sh_2025)}")
    print(f"  2026 pre-June bets:    {len(sh_2026_pre)}")
    print(f"  June holdout bets:     {len(sh_june)}")
    print(f"  Combined deployed:     {len(sh_combined)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Build combined with running stats first (needed for Dashboard/Bankroll)
    print("Computing running stats...")
    combined_run = running_stats(sh_combined.sort_values("date").reset_index(drop=True))
    # Enrich sub-sheets too
    sh_2025_run   = running_stats(sh_2025.sort_values("date").reset_index(drop=True))
    sh_2026_run   = running_stats(sh_2026_pre.sort_values("date").reset_index(drop=True))
    sh_june_run   = running_stats(sh_june.sort_values("date").reset_index(drop=True))

    # Propagate bet_roi
    for df in [combined_run, sh_2025_run, sh_2026_run, sh_june_run]:
        df["bet_roi"] = df.apply(
            lambda r: (r["pnl_a"] / FLAT_STAKE) if not pd.isna(r.get("pnl_a")) else np.nan,
            axis=1
        )

    # ── Individual bet sheets ─────────────────────────────────────────────────
    print("Writing bet sheets...")

    for sheet_name, df_sheet, label in [
        ("2025",          sh_2025_run,   "2025 WF Backtest"),
        ("2026 Pre-June", sh_2026_run,   "2026 Pre-June"),
        ("June Holdout",  sh_june_run,   "June 2026 Holdout"),
        ("Combined",      combined_run,  "Combined Deployed"),
    ]:
        ws = wb.create_sheet(sheet_name)
        write_bet_sheet(ws, df_sheet, label)
        print(f"  OK {sheet_name}")

    # ── Summary sheets ────────────────────────────────────────────────────────
    print("Writing summary sheets...")

    ws_dash = wb.create_sheet("Dashboard")
    write_dashboard(ws_dash, combined_run, sh_2025_run, sh_2026_run, sh_june_run)
    print("  OK Dashboard")

    ws_monthly = wb.create_sheet("Monthly Stats")
    write_monthly_sheet(ws_monthly, combined_run)
    print("  OK Monthly Stats")

    ws_dd = wb.create_sheet("Drawdown")
    write_drawdown_sheet(ws_dd, combined_run)
    print("  OK Drawdown")

    ws_bk = wb.create_sheet("Bankroll Sim")
    write_bankroll_sheet(ws_bk, combined_run)
    print("  OK Bankroll Sim")

    ws_piv = wb.create_sheet("Pivot Summaries")
    write_pivot_sheet(ws_piv, combined_run, sh_june_run)
    print("  OK Pivot Summaries")

    ws_rm = wb.create_sheet("README")
    write_readme(ws_rm)
    print("  OK README")

    # ── Save ─────────────────────────────────────────────────────────────────
    wb.save(OUT_FILE)
    print(f"\nWorkbook saved: {OUT_FILE}")

    # ── Also export combined CSV ──────────────────────────────────────────────
    csv_path = OUT_DIR / "V4_all_bets.csv"
    export_cols = [
        "date","pitcher_name","side","line","projection","abs_gap","edge_pct",
        "egp","bucket","opening_odds","odds_used","closing_odds","clv",
        "dec_odds","mkt_prob","result_label","won","pnl_a","bet_roi",
        "run_win_pct","run_roi","run_units","bankroll_a","source","year","month",
    ]
    all_export = pd.concat([combined_run, sh_june_run], ignore_index=True).sort_values("date")
    avail_cols = [c for c in export_cols if c in all_export.columns]
    all_export[avail_cols].to_csv(csv_path, index=False)
    print(f"CSV export saved: {csv_path}")

    # ── Quick stats summary ──────────────────────────────────────────────────
    s = combined_run[combined_run["result_label"].isin(["Win","Loss"])]
    print(f"\n{'─'*50}")
    print(f"  COMBINED DEPLOYED (no June)")
    print(f"{'─'*50}")
    print(f"  Bets:      {len(s)}")
    print(f"  Win rate:  {(s['won']==1.0).mean():.1%}")
    print(f"  ROI:       {s['pnl_a'].sum()/(len(s)*FLAT_STAKE):.1%}")
    print(f"  Total P&L: ${s['pnl_a'].sum():,.0f}")
    print(f"  Final BR:  ${combined_run['bankroll_a'].iloc[-1]:,.0f}")
    print(f"  Max DD:    {drawdown_series(combined_run['bankroll_a']).min():.1%}")
    print(f"{'─'*50}")

    s2 = sh_june_run[sh_june_run["result_label"].isin(["Win","Loss"])]
    if len(s2):
        print(f"\n  JUNE HOLDOUT")
        print(f"  Bets:      {len(s2)}")
        print(f"  Win rate:  {(s2['won']==1.0).mean():.1%}")
        print(f"  ROI:       {s2['pnl_a'].sum()/(len(s2)*FLAT_STAKE):.1%}")
        print(f"  Total P&L: ${s2['pnl_a'].sum():,.0f}")


if __name__ == "__main__":
    main()

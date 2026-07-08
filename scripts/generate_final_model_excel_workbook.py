from __future__ import annotations

import math
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from scipy.stats import poisson
from sklearn.impute import SimpleImputer
from sklearn.linear_model import PoissonRegressor
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler

from no_leak_2025_to_2026_model_search import (
    EDGE_FILE,
    add_betting_frame,
    filter_strategy,
    fit_model,
    is_feature_col,
    predict_model,
    top_corr_features,
)


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "reports" / "final_model_workbook"
WORKBOOK_PATH = OUT_DIR / "MLB_Pitcher_K_Final_Model_Bet_Ledger.xlsx"
CSV_PATH = OUT_DIR / "final_model_all_bets.csv"
REPORT_PATH = OUT_DIR / "final_model_workbook_summary.md"

START_BANKROLL = 10_000.0
FLAT_STAKE = 100.0
FINAL_MODEL = "poisson_top120_a2"
EDGE_MIN = 7.0
EG_MIN = 0.0
SIDE_FILTER = "under"

BET_COLUMNS = [
    "Bet ID",
    "Date",
    "Game Date",
    "Pitcher",
    "Team",
    "Opponent",
    "Home/Away",
    "Sportsbook",
    "Opening Odds",
    "Closing Odds",
    "Executed Odds",
    "Exchange-adjusted Odds",
    "Strikeout Line",
    "Model Projection",
    "Market Projection",
    "Margin",
    "Devig Edge",
    "Model Probability",
    "Market Implied Probability",
    "Bet Side",
    "Odds",
    "Decimal Odds",
    "American Odds",
    "Units Risked",
    "Units Won/Lost",
    "Stake ($)",
    "Profit/Loss ($)",
    "ROI for Bet",
    "CLV",
    "Result",
    "Running Win %",
    "Running ROI",
    "Running Units",
    "Running Bankroll",
    "Year",
    "Month",
    "Odds Bucket",
    "Model Sample",
    "Kelly Fraction",
]


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
CARD_FILL = PatternFill("solid", fgColor="F4F7FA")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")
DARK_RED_FILL = PatternFill("solid", fgColor="C00000")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")
BOX = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def american_to_decimal_value(odds: float) -> float:
    if pd.isna(odds):
        return np.nan
    odds = float(odds)
    return 1.0 + odds / 100.0 if odds > 0 else 1.0 + 100.0 / abs(odds)


def implied_prob_from_american(odds: pd.Series) -> pd.Series:
    odds = pd.to_numeric(odds, errors="coerce")
    return pd.Series(np.where(odds > 0, 100.0 / (odds + 100.0), odds.abs() / (odds.abs() + 100.0)), index=odds.index)


def odds_bucket(odds: float) -> str:
    if pd.isna(odds):
        return "Missing"
    odds = float(odds)
    if odds <= -150:
        return "<= -150"
    if odds <= -120:
        return "-149 to -120"
    if odds < 100:
        return "-119 to -101"
    if odds <= 120:
        return "+100 to +120"
    if odds <= 150:
        return "+121 to +150"
    return ">= +151"


def build_candidate_bets() -> pd.DataFrame:
    raw = pd.read_csv(EDGE_FILE)
    raw["game_date"] = pd.to_datetime(raw["game_date"])
    raw = raw[raw["market"].eq("strikeouts")].copy()

    numeric_cols = set(raw.select_dtypes(include=[np.number]).columns)
    all_features = [col for col in raw.columns if is_feature_col(col, numeric_cols)]

    train_games = (
        raw[raw["game_date"].dt.year == 2025]
        .sort_values("game_date")
        .drop_duplicates(["game_date", "pitcher_id"], keep="first")
        .reset_index(drop=True)
    )
    train_inner = train_games[train_games["game_date"] < "2025-07-01"].copy()
    val_inner = raw[(raw["game_date"].dt.year == 2025) & (raw["game_date"] >= "2025-07-01")].copy()
    test_2026 = raw[raw["game_date"].dt.year == 2026].copy()

    top120 = top_corr_features(train_inner, all_features, 120)
    estimator = Pipeline(
        [
            ("imputer", SimpleImputer(strategy="median")),
            ("scaler", StandardScaler()),
            ("model", PoissonRegressor(alpha=2.0, max_iter=2000)),
        ]
    )

    fitted_inner = fit_model(FINAL_MODEL, train_inner, top120, estimator)
    val_pred = predict_model(fitted_inner, val_inner)
    val_bets = add_betting_frame(val_inner, val_pred, fitted_inner.residual_std, FINAL_MODEL)
    bets_2025 = filter_strategy(val_bets, EDGE_MIN, EG_MIN, SIDE_FILTER)
    bets_2025["Model Sample"] = "2025 Validation"

    final_estimator = Pipeline(
        [
            ("imputer", SimpleImputer(strategy="median")),
            ("scaler", StandardScaler()),
            ("model", PoissonRegressor(alpha=2.0, max_iter=2000)),
        ]
    )
    fitted_final = fit_model(FINAL_MODEL, train_games, top120, final_estimator)
    test_pred = predict_model(fitted_final, test_2026)
    test_bets = add_betting_frame(test_2026, test_pred, fitted_final.residual_std, FINAL_MODEL)
    bets_2026 = filter_strategy(test_bets, EDGE_MIN, EG_MIN, SIDE_FILTER)
    bets_2026["Model Sample"] = "2026 Frozen Test"

    out = pd.concat([bets_2025, bets_2026], ignore_index=True)
    out = out.sort_values(["game_date", "pitcher_name", "line", "side", "bet_odds"]).reset_index(drop=True)
    return out


def calculate_kelly_fraction(prob: pd.Series, decimal_odds: pd.Series) -> pd.Series:
    b = pd.to_numeric(decimal_odds, errors="coerce") - 1.0
    p = pd.to_numeric(prob, errors="coerce")
    q = 1.0 - p
    kelly = ((b * p) - q) / b
    return kelly.clip(lower=0.0).replace([np.inf, -np.inf], np.nan).fillna(0.0)


def transform_bets(raw_bets: pd.DataFrame) -> pd.DataFrame:
    df = raw_bets.copy()
    df["fetched_at"] = pd.to_datetime(df["fetched_at"], errors="coerce")
    df["game_date"] = pd.to_datetime(df["game_date"], errors="coerce")

    over_imp = implied_prob_from_american(df["over_odds"])
    under_imp = implied_prob_from_american(df["under_odds"])
    denom = over_imp + under_imp
    market_under = under_imp / denom
    market_over = over_imp / denom

    model_prob = np.where(df["side"].eq("over"), df["over_probability"], df["under_probability"])
    market_prob = np.where(df["side"].eq("over"), market_over, market_under)
    side_margin = np.where(df["side"].eq("over"), df["projection"] - df["line"], df["line"] - df["projection"])
    sportsbook = np.where(df["side"].eq("over"), df["over_bookmaker"], df["under_bookmaker"])

    result = np.select([df["push"], df["won"]], ["Push", "Win"], default="Loss")
    american = pd.to_numeric(df["bet_odds"], errors="coerce")
    decimal = pd.to_numeric(df["decimal_odds"], errors="coerce")
    profit_unit = pd.to_numeric(df["profit_unit"], errors="coerce")

    transformed = pd.DataFrame(
        {
            "Date": df["fetched_at"].dt.tz_localize(None),
            "Game Date": df["game_date"].dt.date,
            "Pitcher": df["pitcher_name"],
            "Team": df["team"],
            "Opponent": df["opponent"],
            "Home/Away": np.where(df["is_home"].astype(bool), "Home", "Away"),
            "Sportsbook": sportsbook,
            "Opening Odds": np.nan,
            "Closing Odds": np.nan,
            "Executed Odds": american,
            "Exchange-adjusted Odds": american,
            "Strikeout Line": pd.to_numeric(df["line"], errors="coerce"),
            "Model Projection": pd.to_numeric(df["projection"], errors="coerce"),
            "Market Projection": pd.to_numeric(df["line"], errors="coerce"),
            "Margin": side_margin,
            "Devig Edge": pd.to_numeric(df["edge_pct"], errors="coerce") / 100.0,
            "Model Probability": model_prob,
            "Market Implied Probability": market_prob,
            "Bet Side": df["side"].str.title(),
            "Odds": american,
            "Decimal Odds": decimal,
            "American Odds": american,
            "Units Risked": 1.0,
            "Units Won/Lost": profit_unit,
            "Stake ($)": FLAT_STAKE,
            "Profit/Loss ($)": profit_unit * FLAT_STAKE,
            "ROI for Bet": profit_unit,
            "CLV": np.nan,
            "Result": result,
            "Year": df["game_date"].dt.year,
            "Month": df["game_date"].dt.to_period("M").astype(str),
            "Odds Bucket": [odds_bucket(x) for x in american],
            "Model Sample": df["Model Sample"],
        }
    )
    transformed["Kelly Fraction"] = calculate_kelly_fraction(transformed["Model Probability"], transformed["Decimal Odds"])
    transformed.insert(0, "Bet ID", [f"KPROP-{i:05d}" for i in range(1, len(transformed) + 1)])

    transformed["Running Win %"] = running_win_rate(transformed["Result"])
    transformed["Running ROI"] = transformed["Units Won/Lost"].cumsum() / transformed["Units Risked"].cumsum()
    transformed["Running Units"] = transformed["Units Won/Lost"].cumsum()
    transformed["Running Bankroll"] = START_BANKROLL + transformed["Profit/Loss ($)"].cumsum()
    return transformed[BET_COLUMNS]


def running_win_rate(results: pd.Series) -> list[float]:
    wins = 0
    decisions = 0
    rates = []
    for value in results:
        if value == "Win":
            wins += 1
            decisions += 1
        elif value == "Loss":
            decisions += 1
        rates.append(wins / decisions if decisions else np.nan)
    return rates


def summary_stats(df: pd.DataFrame) -> dict[str, float]:
    wins = int(df["Result"].eq("Win").sum())
    losses = int(df["Result"].eq("Loss").sum())
    pushes = int(df["Result"].eq("Push").sum())
    decisions = wins + losses
    profit = float(df["Profit/Loss ($)"].sum())
    stake = float(df["Stake ($)"].sum())
    return {
        "Bets": int(len(df)),
        "Wins": wins,
        "Losses": losses,
        "Pushes": pushes,
        "Win Rate": wins / decisions if decisions else np.nan,
        "ROI": profit / stake if stake else np.nan,
        "Units": float(df["Units Won/Lost"].sum()),
        "Profit": profit,
        "CLV": float(df["CLV"].mean()) if df["CLV"].notna().any() else np.nan,
        "Average Odds": float(df["American Odds"].mean()),
        "Average Margin": float(df["Margin"].mean()),
        "Average Devig Edge": float(df["Devig Edge"].mean()),
        "Average Stake": float(df["Stake ($)"].mean()),
    }


def make_group_summary(df: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    rows = []
    for key, group in df.groupby(group_cols, dropna=False):
        if not isinstance(key, tuple):
            key = (key,)
        row = {col: value for col, value in zip(group_cols, key)}
        row.update(summary_stats(group))
        rows.append(row)
    out = pd.DataFrame(rows)
    sort_cols = group_cols if group_cols else ["Bets"]
    return out.sort_values(sort_cols).reset_index(drop=True)


def simulate_bankroll(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    states = {
        "A": {"bankroll": START_BANKROLL, "peak": START_BANKROLL},
        "B": {"bankroll": START_BANKROLL, "peak": START_BANKROLL},
        "C": {"bankroll": START_BANKROLL, "peak": START_BANKROLL},
        "D": {"bankroll": START_BANKROLL, "peak": START_BANKROLL},
    }
    for i, bet in df.reset_index(drop=True).iterrows():
        profit_unit = float(bet["Units Won/Lost"])
        kelly = float(bet["Kelly Fraction"])
        stakes = {
            "A": FLAT_STAKE,
            "B": states["B"]["bankroll"] * 0.01,
            "C": states["C"]["bankroll"] * kelly * 0.25,
            "D": states["D"]["bankroll"] * kelly * 0.50,
        }
        row = {
            "Bet #": i + 1,
            "Bet ID": bet["Bet ID"],
            "Game Date": bet["Game Date"],
            "Pitcher": bet["Pitcher"],
            "Result": bet["Result"],
            "Profit Unit": profit_unit,
            "Kelly Fraction": kelly,
        }
        for label in ["A", "B", "C", "D"]:
            stake = stakes[label]
            pnl = stake * profit_unit
            states[label]["bankroll"] += pnl
            states[label]["peak"] = max(states[label]["peak"], states[label]["bankroll"])
            drawdown = states[label]["bankroll"] - states[label]["peak"]
            dd_pct = drawdown / states[label]["peak"] if states[label]["peak"] else 0.0
            row[f"Strategy {label} Stake"] = stake
            row[f"Strategy {label} P/L"] = pnl
            row[f"Strategy {label} Bankroll"] = states[label]["bankroll"]
            row[f"Strategy {label} Peak"] = states[label]["peak"]
            row[f"Strategy {label} Drawdown $"] = drawdown
            row[f"Strategy {label} Drawdown %"] = dd_pct
        rows.append(row)
    return pd.DataFrame(rows)


def longest_streak(results: Iterable[str], target: str) -> int:
    best = 0
    current = 0
    for result in results:
        if result == target:
            current += 1
            best = max(best, current)
        elif result != "Push":
            current = 0
    return best


def drawdown_periods(df: pd.DataFrame) -> pd.DataFrame:
    bankroll = df["Running Bankroll"].to_numpy(dtype=float)
    dates = pd.to_datetime(df["Game Date"]).dt.date.to_list()
    periods = []
    peak_value = START_BANKROLL
    peak_date = dates[0] if dates else None
    in_drawdown = False
    trough_value = START_BANKROLL
    trough_date = None
    start_bet = 0
    trough_bet = 0

    for idx, value in enumerate(bankroll, start=1):
        date = dates[idx - 1]
        if value >= peak_value:
            if in_drawdown:
                periods.append(
                    {
                        "Peak Date": peak_date,
                        "Trough Date": trough_date,
                        "Recovery Date": date,
                        "Drawdown %": (trough_value - peak_value) / peak_value,
                        "Drawdown $": trough_value - peak_value,
                        "Duration": (pd.Timestamp(date) - pd.Timestamp(peak_date)).days,
                        "Number of Bets": idx - start_bet + 1,
                    }
                )
                in_drawdown = False
            peak_value = value
            peak_date = date
            start_bet = idx
            trough_value = value
            trough_date = date
        else:
            if not in_drawdown:
                in_drawdown = True
                trough_value = value
                trough_date = date
                trough_bet = idx
            elif value < trough_value:
                trough_value = value
                trough_date = date
                trough_bet = idx

    if in_drawdown:
        periods.append(
            {
                "Peak Date": peak_date,
                "Trough Date": trough_date,
                "Recovery Date": "",
                "Drawdown %": (trough_value - peak_value) / peak_value,
                "Drawdown $": trough_value - peak_value,
                "Duration": (pd.Timestamp(dates[-1]) - pd.Timestamp(peak_date)).days if dates else 0,
                "Number of Bets": max(0, len(df) - start_bet + 1),
            }
        )
    periods_df = pd.DataFrame(periods)
    if not periods_df.empty:
        periods_df = periods_df.sort_values("Drawdown %").reset_index(drop=True)
    return periods_df


def max_drawdown_for_series(values: pd.Series) -> float:
    peak = values.cummax()
    dd = values - peak
    return float(dd.min()) if len(dd) else 0.0


def monthly_stats(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for month, group in df.groupby("Month", sort=True):
        row = {"Month": month}
        row.update(summary_stats(group))
        row["Maximum Drawdown"] = max_drawdown_for_series(group["Running Bankroll"])
        row["Ending Bankroll"] = float(group["Running Bankroll"].iloc[-1])
        rows.append(row)
    return pd.DataFrame(rows)


def set_sheet_basics(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False


def write_dataframe(ws, df: pd.DataFrame, table_name: str | None = None, start_row: int = 1, start_col: int = 1) -> None:
    for c_idx, col in enumerate(df.columns, start_col):
        cell = ws.cell(start_row, c_idx, col)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX
    for r_offset, row in enumerate(df.itertuples(index=False, name=None), 1):
        for c_offset, value in enumerate(row, 0):
            cell = ws.cell(start_row + r_offset, start_col + c_offset, clean_excel_value(value))
            cell.border = BOX
            cell.alignment = Alignment(vertical="center")
    end_row = start_row + len(df)
    end_col = start_col + len(df.columns) - 1
    if table_name and len(df) > 0:
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)
    autosize(ws)


def clean_excel_value(value):
    if isinstance(value, (pd.Timestamp,)):
        return value.to_pydatetime()
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def autosize(ws, max_width: int = 32) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells[:200]:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def apply_number_formats(ws, headers: list[str], start_row: int = 2) -> None:
    pct_cols = {"Devig Edge", "Model Probability", "Market Implied Probability", "ROI for Bet", "CLV", "Running Win %", "Running ROI", "Kelly Fraction", "Win Rate", "ROI", "Average Devig Edge", "Drawdown %", "Total Return", "CAGR"}
    money_cols = {"Stake ($)", "Profit/Loss ($)", "Running Bankroll", "Profit", "Average Stake", "Ending Bankroll", "Maximum Drawdown", "Final Bankroll", "Stake", "Peak Bankroll", "Drawdown $"}
    odds_cols = {"Opening Odds", "Closing Odds", "Executed Odds", "Exchange-adjusted Odds", "Odds", "American Odds", "Average Odds"}
    for idx, header in enumerate(headers, 1):
        letter = get_column_letter(idx)
        for cell in ws[f"{letter}{start_row}": f"{letter}{ws.max_row}"]:
            c = cell[0]
            if header in pct_cols:
                c.number_format = "0.0%"
            elif header in money_cols or "$" in header or "Bankroll" in header:
                c.number_format = '$#,##0.00;[Red]($#,##0.00)'
            elif header in odds_cols:
                c.number_format = '0'
            elif "Date" in header:
                c.number_format = "yyyy-mm-dd"
            elif header in {"Strikeout Line", "Model Projection", "Market Projection", "Margin", "Units Won/Lost", "Running Units"}:
                c.number_format = "0.00"


def add_bet_sheet(wb: Workbook, name: str, df: pd.DataFrame, table_name: str) -> None:
    ws = wb.create_sheet(name)
    set_sheet_basics(ws)
    write_dataframe(ws, df, table_name=table_name)

    headers = list(df.columns)
    col = {name: idx + 1 for idx, name in enumerate(headers)}
    for row in range(2, ws.max_row + 1):
        excel_row = row
        units_col = get_column_letter(col["Units Won/Lost"])
        stake_col = get_column_letter(col["Stake ($)"])
        profit_col = get_column_letter(col["Profit/Loss ($)"])
        roi_col = get_column_letter(col["ROI for Bet"])
        result_col = get_column_letter(col["Result"])
        unit_risk_col = get_column_letter(col["Units Risked"])
        ws.cell(row, col["Profit/Loss ($)"], f"={units_col}{excel_row}*{stake_col}{excel_row}")
        ws.cell(row, col["ROI for Bet"], f"=IF({stake_col}{excel_row}=0,0,{profit_col}{excel_row}/{stake_col}{excel_row})")
        ws.cell(row, col["Running Win %"], f'=IF((COUNTIF(${result_col}$2:{result_col}{excel_row},"Win")+COUNTIF(${result_col}$2:{result_col}{excel_row},"Loss"))=0,0,COUNTIF(${result_col}$2:{result_col}{excel_row},"Win")/(COUNTIF(${result_col}$2:{result_col}{excel_row},"Win")+COUNTIF(${result_col}$2:{result_col}{excel_row},"Loss")))')
        ws.cell(row, col["Running ROI"], f"=SUM(${units_col}$2:{units_col}{excel_row})/SUM(${unit_risk_col}$2:{unit_risk_col}{excel_row})")
        ws.cell(row, col["Running Units"], f"=SUM(${units_col}$2:{units_col}{excel_row})")
        ws.cell(row, col["Running Bankroll"], f"={START_BANKROLL}+SUM(${profit_col}$2:{profit_col}{excel_row})")

    apply_number_formats(ws, headers)
    apply_bet_conditional_formatting(ws, col)


def apply_bet_conditional_formatting(ws, col: dict[str, int]) -> None:
    max_row = ws.max_row
    result_letter = get_column_letter(col["Result"])
    roi_letter = get_column_letter(col["ROI for Bet"])
    clv_letter = get_column_letter(col["CLV"])
    pnl_letter = get_column_letter(col["Profit/Loss ($)"])
    running_bankroll_letter = get_column_letter(col["Running Bankroll"])

    ws.conditional_formatting.add(f"{result_letter}2:{result_letter}{max_row}", CellIsRule(operator="equal", formula=['"Win"'], fill=GREEN_FILL))
    ws.conditional_formatting.add(f"{result_letter}2:{result_letter}{max_row}", CellIsRule(operator="equal", formula=['"Loss"'], fill=RED_FILL))
    ws.conditional_formatting.add(f"{result_letter}2:{result_letter}{max_row}", CellIsRule(operator="equal", formula=['"Push"'], fill=GRAY_FILL))
    ws.conditional_formatting.add(f"{roi_letter}2:{roi_letter}{max_row}", ColorScaleRule(start_type="min", start_color="FFC7CE", mid_type="num", mid_value=0, mid_color="FFFFFF", end_type="max", end_color="C6EFCE"))
    ws.conditional_formatting.add(f"{clv_letter}2:{clv_letter}{max_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL))
    ws.conditional_formatting.add(f"{clv_letter}2:{clv_letter}{max_row}", CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL))
    ws.conditional_formatting.add(f"{pnl_letter}2:{pnl_letter}{max_row}", FormulaRule(formula=[f"{pnl_letter}2>=LARGE(${pnl_letter}$2:${pnl_letter}${max_row},10)"], fill=GREEN_FILL))
    ws.conditional_formatting.add(f"{pnl_letter}2:{pnl_letter}{max_row}", FormulaRule(formula=[f"{pnl_letter}2<=SMALL(${pnl_letter}$2:${pnl_letter}${max_row},10)"], fill=RED_FILL))
    ws.conditional_formatting.add(f"{running_bankroll_letter}2:{running_bankroll_letter}{max_row}", ColorScaleRule(start_type="min", start_color="F4B183", mid_type="percentile", mid_value=50, mid_color="FFFFFF", end_type="max", end_color="C6EFCE"))


def add_dashboard(wb: Workbook, combined: pd.DataFrame, monthly: pd.DataFrame, bankroll: pd.DataFrame, dd: pd.DataFrame) -> None:
    ws = wb.create_sheet("Dashboard")
    set_sheet_basics(ws, "A1")
    ws["A1"] = "MLB Pitcher Strikeout Model Ledger Dashboard"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    ws["A2"] = f"Model: {FINAL_MODEL} | Rule: {SIDE_FILTER} only, devig edge >= {EDGE_MIN:.0f}% | Flat ledger stake: ${FLAT_STAKE:,.0f}"
    ws["A2"].font = Font(italic=True, color="666666")

    stats = summary_stats(combined)
    max_dd = max_drawdown_for_series(combined["Running Bankroll"])
    final_bankroll = float(combined["Running Bankroll"].iloc[-1])
    kpis = [
        ("Total Bets", stats["Bets"]),
        ("Win Rate", stats["Win Rate"]),
        ("ROI", stats["ROI"]),
        ("Units", stats["Units"]),
        ("Profit", stats["Profit"]),
        ("Average Odds", stats["Average Odds"]),
        ("Average CLV", np.nan),
        ("Average Stake", stats["Average Stake"]),
        ("Max Drawdown", max_dd),
        ("Final Bankroll", final_bankroll),
        ("Biggest Winning Streak", longest_streak(combined["Result"], "Win")),
        ("Biggest Losing Streak", longest_streak(combined["Result"], "Loss")),
        ("Average Margin", stats["Average Margin"]),
        ("Average Devig Edge", stats["Average Devig Edge"]),
    ]

    for idx, (label, value) in enumerate(kpis):
        row = 4 + (idx // 4) * 3
        col = 1 + (idx % 4) * 3
        ws.cell(row, col, label)
        ws.cell(row + 1, col, clean_excel_value(value))
        ws.cell(row, col).fill = SUBHEADER_FILL
        ws.cell(row, col).font = BOLD_FONT
        ws.cell(row + 1, col).fill = CARD_FILL
        ws.cell(row + 1, col).font = Font(size=14, bold=True, color="1F4E78")
        ws.cell(row, col).border = BOX
        ws.cell(row + 1, col).border = BOX
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        if "Rate" in label or label in {"ROI", "Average CLV", "Average Devig Edge"}:
            ws.cell(row + 1, col).number_format = "0.0%"
        elif label in {"Profit", "Average Stake", "Max Drawdown", "Final Bankroll"}:
            ws.cell(row + 1, col).number_format = '$#,##0.00;[Red]($#,##0.00)'
        elif label in {"Units", "Average Margin"}:
            ws.cell(row + 1, col).number_format = "0.00"

    chart_start = 18
    chart_df = pd.DataFrame(
        {
            "Bet #": np.arange(1, len(combined) + 1),
            "Running Bankroll": combined["Running Bankroll"].to_numpy(),
            "Running Units": combined["Running Units"].to_numpy(),
            "Cumulative Profit": combined["Profit/Loss ($)"].cumsum().to_numpy(),
            "Drawdown": combined["Running Bankroll"].to_numpy() - np.maximum.accumulate(combined["Running Bankroll"].to_numpy()),
            "Rolling ROI (25)": combined["Units Won/Lost"].rolling(25).mean().to_numpy(),
            "Rolling Win Rate (25)": combined["Result"].eq("Win").rolling(25).mean().to_numpy(),
        }
    )
    write_dataframe(ws, chart_df, table_name=None, start_row=chart_start, start_col=1)
    monthly_chart_col = 10
    write_dataframe(ws, monthly[["Month", "ROI", "Win Rate", "Bets"]], table_name=None, start_row=chart_start, start_col=monthly_chart_col)

    add_line_chart(ws, "Running Bankroll", chart_start, 1, 2, len(chart_df), "J4")
    add_line_chart(ws, "Running Units", chart_start, 1, 3, len(chart_df), "J19")
    add_line_chart(ws, "Cumulative Profit", chart_start, 1, 4, len(chart_df), "J34")
    add_bar_chart(ws, "Monthly ROI", chart_start, monthly_chart_col, monthly_chart_col + 1, len(monthly), "R4")
    add_bar_chart(ws, "Monthly Win Rate", chart_start, monthly_chart_col, monthly_chart_col + 2, len(monthly), "R19")
    add_bar_chart(ws, "Monthly Bet Count", chart_start, monthly_chart_col, monthly_chart_col + 3, len(monthly), "R34")
    add_line_chart(ws, "Drawdown Curve", chart_start, 1, 5, len(chart_df), "Z4")
    add_line_chart(ws, "Rolling ROI (25 bets)", chart_start, 1, 6, len(chart_df), "Z19")
    add_line_chart(ws, "Rolling Win Rate (25 bets)", chart_start, 1, 7, len(chart_df), "Z34")

    for col_idx in range(1, 40):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13


def add_line_chart(ws, title: str, start_row: int, cat_col: int, data_col: int, rows: int, anchor: str) -> None:
    chart = LineChart()
    chart.title = title
    chart.height = 7
    chart.width = 14
    data = Reference(ws, min_col=data_col, min_row=start_row, max_row=start_row + rows)
    cats = Reference(ws, min_col=cat_col, min_row=start_row + 1, max_row=start_row + rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, anchor)


def add_bar_chart(ws, title: str, start_row: int, cat_col: int, data_col: int, rows: int, anchor: str) -> None:
    chart = BarChart()
    chart.title = title
    chart.height = 7
    chart.width = 14
    data = Reference(ws, min_col=data_col, min_row=start_row, max_row=start_row + rows)
    cats = Reference(ws, min_col=cat_col, min_row=start_row + 1, max_row=start_row + rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, anchor)


def add_bankroll_sheet(wb: Workbook, bankroll: pd.DataFrame, combined: pd.DataFrame) -> pd.DataFrame:
    ws = wb.create_sheet("Bankroll Simulation")
    set_sheet_basics(ws, "A9")

    summary_rows = []
    first_date = pd.to_datetime(combined["Game Date"]).min()
    last_date = pd.to_datetime(combined["Game Date"]).max()
    years = max((last_date - first_date).days / 365.25, 1 / 365.25)
    for label, description in [
        ("A", "Flat $100 per bet"),
        ("B", "Flat 1% of current bankroll"),
        ("C", "Quarter Kelly from model probability"),
        ("D", "Half Kelly from model probability"),
    ]:
        final = float(bankroll[f"Strategy {label} Bankroll"].iloc[-1])
        total_return = final / START_BANKROLL - 1.0
        cagr = (final / START_BANKROLL) ** (1 / years) - 1.0 if years > 0 and final > 0 else np.nan
        summary_rows.append(
            {
                "Strategy": f"Strategy {label}",
                "Description": description,
                "Final Bankroll": final,
                "Total Return": total_return,
                "CAGR": cagr,
                "Maximum Drawdown": float(bankroll[f"Strategy {label} Drawdown $"].min()),
                "Maximum Drawdown %": float(bankroll[f"Strategy {label} Drawdown %"].min()),
            }
        )
    summary = pd.DataFrame(summary_rows)
    write_dataframe(ws, summary, table_name="TblBankrollSummary", start_row=1, start_col=1)
    apply_number_formats(ws, list(summary.columns), start_row=2)

    write_dataframe(ws, bankroll, table_name="TblBankrollSimulation", start_row=9, start_col=1)
    for col_idx, header in enumerate(bankroll.columns, 1):
        if "Stake" in header or "P/L" in header or "Bankroll" in header or "Peak" in header or "Drawdown $" in header:
            for row in range(10, ws.max_row + 1):
                ws.cell(row, col_idx).number_format = '$#,##0.00;[Red]($#,##0.00)'
        elif "Drawdown %" in header or "Kelly" in header:
            for row in range(10, ws.max_row + 1):
                ws.cell(row, col_idx).number_format = "0.0%"
    for col_idx, header in enumerate(bankroll.columns, 1):
        if "Drawdown" in header:
            letter = get_column_letter(col_idx)
            ws.conditional_formatting.add(f"{letter}10:{letter}{ws.max_row}", ColorScaleRule(start_type="min", start_color="C00000", mid_type="percentile", mid_value=50, mid_color="FCE4D6", end_type="max", end_color="FFFFFF"))
    return summary


def add_drawdown_sheet(wb: Workbook, dd: pd.DataFrame) -> None:
    ws = wb.create_sheet("Drawdown Analysis")
    set_sheet_basics(ws)
    if dd.empty:
        dd = pd.DataFrame(columns=["Peak Date", "Trough Date", "Recovery Date", "Drawdown %", "Drawdown $", "Duration", "Number of Bets"])
    write_dataframe(ws, dd, table_name="TblDrawdowns")
    apply_number_formats(ws, list(dd.columns), start_row=2)
    if ws.max_row >= 2:
        dd_col = get_column_letter(list(dd.columns).index("Drawdown %") + 1)
        ws.conditional_formatting.add(f"{dd_col}2:{dd_col}{ws.max_row}", ColorScaleRule(start_type="min", start_color="C00000", mid_type="percentile", mid_value=50, mid_color="F4B183", end_type="max", end_color="FFFFFF"))


def add_summary_tables(wb: Workbook, combined: pd.DataFrame) -> None:
    ws = wb.create_sheet("Summary Tables")
    set_sheet_basics(ws, "A2")
    row = 1
    tables = [
        ("By Year", ["Year"]),
        ("By Month", ["Month"]),
        ("By Pitcher", ["Pitcher"]),
        ("By Opponent", ["Opponent"]),
        ("By Team", ["Team"]),
        ("By Odds Bucket", ["Odds Bucket"]),
        ("By Strikeout Line", ["Strikeout Line"]),
        ("By Sportsbook", ["Sportsbook"]),
        ("By Bet Side", ["Bet Side"]),
    ]
    for title, group_cols in tables:
        ws.cell(row, 1, title)
        ws.cell(row, 1).font = Font(size=13, bold=True, color="1F4E78")
        row += 1
        table = make_group_summary(combined, group_cols)
        safe_name = "Tbl" + "".join(ch for ch in title if ch.isalnum())
        write_dataframe(ws, table, table_name=safe_name, start_row=row, start_col=1)
        apply_number_formats(ws, list(table.columns), start_row=row + 1)
        row += len(table) + 4


def add_readme(wb: Workbook, validation: dict[str, object]) -> None:
    ws = wb.create_sheet("README")
    set_sheet_basics(ws, "A1")
    ws["A1"] = "README"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    lines = [
        "Workbook structure",
        "2025 contains validation bets generated by the no-leak inner model trained only on games before 2025-07-01.",
        "2026 Pre-June contains frozen 2026 test bets before 2026-06-01.",
        "June Holdout contains settled June 2026 bets available in the source data, ending at the latest settled game date present.",
        "Combined contains every bet in chronological order and is the source ledger for summary sheets.",
        "Dashboard, Monthly Statistics, Drawdown Analysis, Bankroll Simulation, and Summary Tables summarize performance.",
        "",
        "Model and bet rule",
        f"Model: {FINAL_MODEL}. Rule: {SIDE_FILTER} only with devig edge >= {EDGE_MIN:.0f}% and edge-gap >= {EG_MIN:.0f}.",
        "2026 bets use a model fit only on 2025 pitcher-game rows. The raw same-day league_k leakage feature is excluded by the model-search feature screen.",
        "",
        "Staking strategies",
        "Strategy A risks a flat $100 per bet.",
        "Strategy B risks 1% of current bankroll.",
        "Strategy C risks one-quarter Kelly using the model probability and executed decimal odds.",
        "Strategy D risks one-half Kelly using the model probability and executed decimal odds.",
        "Kelly simulations are included because model probabilities exist, but they are not staking recommendations and are not CLV-validated.",
        "",
        "Formulas",
        "Profit/Loss equals Units Won/Lost times Stake.",
        "ROI equals Profit/Loss divided by Stake.",
        "Running Bankroll equals $10,000 plus cumulative Profit/Loss in the flat-stake ledger.",
        "Kelly fraction equals max(0, (b*p - (1-p))/b), where b is decimal odds minus 1 and p is model probability.",
        "",
        "Color coding",
        "Wins are green, losses red, pushes gray. ROI uses a red-white-green scale. Drawdowns use orange/red scales.",
        "Top 10 wins and losses are highlighted in the individual bet sheets.",
        "",
        "Assumptions and excluded data",
        "Opening odds, closing odds, and CLV are unavailable for the clean no-leak candidate ledger, so those columns are intentionally blank.",
        "Exchange-adjusted odds equal executed odds because no exchange-specific adjustment was available in the source data.",
        "Market Projection uses the posted strikeout line as a market-center proxy because no separate market mean was available.",
        "No bets were excluded after applying the documented model rule and de-duplication key.",
        "",
        "Validation checks",
        f"Unique Bet IDs: {validation['unique_bet_ids']}",
        f"Duplicate bets detected: {validation['duplicate_bets']}",
        f"Flat bankroll reconciliation: {validation['bankroll_reconciles']}",
        f"Profit reconciliation: {validation['profit_reconciles']}",
        f"Units reconciliation: {validation['units_reconciles']}",
    ]
    for row, text in enumerate(lines, 3):
        ws.cell(row, 1, text)
        if text and row in {3, 10, 14, 21, 27, 31, 37}:
            ws.cell(row, 1).font = Font(size=13, bold=True, color="1F4E78")
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 120


def write_report(combined: pd.DataFrame, bankroll_summary: pd.DataFrame, validation: dict[str, object], monthly: pd.DataFrame) -> None:
    stats = summary_stats(combined)
    lines = [
        "# Final Model Workbook Summary",
        "",
        f"Workbook: `{WORKBOOK_PATH}`",
        f"All-bets CSV: `{CSV_PATH}`",
        "",
        "## Model Ledger",
        "",
        f"- Model: `{FINAL_MODEL}`",
        f"- Rule: `{SIDE_FILTER}` only, devig edge >= {EDGE_MIN:.0f}%, edge-gap >= {EG_MIN:.0f}",
        f"- Bet count: {stats['Bets']:,}",
        f"- Win rate: {stats['Win Rate']:.2%}",
        f"- Flat-stake ROI: {stats['ROI']:.2%}",
        f"- Units: {stats['Units']:.2f}",
        f"- Profit at $100 flat stake: ${stats['Profit']:,.2f}",
        f"- Average odds: {stats['Average Odds']:.0f}",
        f"- Average margin: {stats['Average Margin']:.2f} strikeouts",
        f"- Average devig edge: {stats['Average Devig Edge']:.2%}",
        "",
        "## Bankroll Simulations",
        "",
    ]
    for _, row in bankroll_summary.iterrows():
        lines.append(
            f"- {row['Strategy']} ({row['Description']}): final bankroll ${row['Final Bankroll']:,.2f}, "
            f"total return {row['Total Return']:.2%}, max drawdown ${row['Maximum Drawdown']:,.2f} ({row['Maximum Drawdown %']:.2%})"
        )
    monthly_rows = ["| Month | Bets | Win Rate | ROI | Units | Profit | Ending Bankroll |", "|---|---:|---:|---:|---:|---:|---:|"]
    for _, row in monthly.iterrows():
        monthly_rows.append(
            f"| {row['Month']} | {int(row['Bets']):,} | {row['Win Rate']:.2%} | {row['ROI']:.2%} | "
            f"{row['Units']:.2f} | ${row['Profit']:,.2f} | ${row['Ending Bankroll']:,.2f} |"
        )

    lines.extend(
        [
            "",
            "## Monthly Snapshot",
            "",
            "\n".join(monthly_rows),
            "",
            "## Validation Checks",
            "",
            f"- Unique Bet IDs: {validation['unique_bet_ids']}",
            f"- Duplicate bets detected: {validation['duplicate_bets']}",
            f"- Running bankroll reconciles: {validation['bankroll_reconciles']}",
            f"- Profit equals ending bankroll minus starting bankroll: {validation['profit_reconciles']}",
            f"- Units reconcile with total profit at $100/unit: {validation['units_reconciles']}",
            "",
            "## Data Notes",
            "",
            "- Opening odds, closing odds, and CLV were not available for the clean no-leak candidate ledger and were intentionally left blank.",
            "- 2025 bets are validation bets from the inner no-leak model. 2026 bets are frozen-model test bets.",
            "- The June holdout ends at the latest settled result available in the source data.",
        ]
    )
    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


def validation_checks(df: pd.DataFrame) -> dict[str, object]:
    dup_keys = ["Game Date", "Pitcher", "Team", "Opponent", "Strikeout Line", "Bet Side", "Sportsbook"]
    duplicate_bets = int(df.duplicated(dup_keys).sum())
    unique_bet_ids = bool(df["Bet ID"].is_unique)
    profit = float(df["Profit/Loss ($)"].sum())
    ending_bankroll = float(df["Running Bankroll"].iloc[-1])
    units = float(df["Units Won/Lost"].sum())
    return {
        "duplicate_bets": duplicate_bets,
        "unique_bet_ids": unique_bet_ids,
        "bankroll_reconciles": bool(abs(ending_bankroll - (START_BANKROLL + profit)) < 1e-8),
        "profit_reconciles": bool(abs(profit - (ending_bankroll - START_BANKROLL)) < 1e-8),
        "units_reconciles": bool(abs(units * FLAT_STAKE - profit) < 1e-8),
    }


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    raw_bets = build_candidate_bets()
    combined = transform_bets(raw_bets)
    combined.to_csv(CSV_PATH, index=False)

    bets_2025 = combined[combined["Year"].eq(2025)].copy()
    bets_2026_pre_june = combined[(combined["Year"].eq(2026)) & (pd.to_datetime(combined["Game Date"]) < pd.Timestamp("2026-06-01"))].copy()
    june_holdout = combined[(combined["Year"].eq(2026)) & (pd.to_datetime(combined["Game Date"]) >= pd.Timestamp("2026-06-01"))].copy()

    monthly = monthly_stats(combined)
    dd = drawdown_periods(combined)
    bankroll = simulate_bankroll(combined)
    validation = validation_checks(combined)

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    add_bet_sheet(wb, "2025", bets_2025, "TblBets2025")
    add_bet_sheet(wb, "2026 Pre-June", bets_2026_pre_june, "TblBets2026PreJune")
    add_bet_sheet(wb, "June Holdout", june_holdout, "TblBetsJuneHoldout")
    add_bet_sheet(wb, "Combined", combined, "TblCombinedBets")

    add_dashboard(wb, combined, monthly, bankroll, dd)

    ws_month = wb.create_sheet("Monthly Statistics")
    set_sheet_basics(ws_month)
    write_dataframe(ws_month, monthly, table_name="TblMonthlyStats")
    apply_number_formats(ws_month, list(monthly.columns), start_row=2)

    add_drawdown_sheet(wb, dd)
    bankroll_summary = add_bankroll_sheet(wb, bankroll, combined)
    add_summary_tables(wb, combined)
    add_readme(wb, validation)

    wb.save(WORKBOOK_PATH)
    write_report(combined, bankroll_summary, validation, monthly)

    print(f"Wrote workbook: {WORKBOOK_PATH}")
    print(f"Wrote all-bets CSV: {CSV_PATH}")
    print(f"Wrote summary report: {REPORT_PATH}")
    print(f"Bets: {len(combined):,}; 2025={len(bets_2025):,}; 2026 pre-June={len(bets_2026_pre_june):,}; June={len(june_holdout):,}")
    print(validation)


if __name__ == "__main__":
    main()

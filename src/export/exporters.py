from __future__ import annotations

from pathlib import Path

import pandas as pd


def export_csv(df: pd.DataFrame, path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)
    return path


def export_pretty_excel(df: pd.DataFrame, path: str | Path) -> Path:
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    display_cols = [
        "pitcher_name",
        "market",
        "best_side",
        "projection",
        "line",
        "recommended_odds",
        "fair_odds",
        "hit_probability_pct",
        "edge",
        "kelly",
        "bookmaker",
        "team",
        "opponent",
        "fetched_at",
    ]
    display = df[[col for col in display_cols if col in df.columns]].copy()
    if "hit_probability" in df.columns and "hit_probability_pct" in display.columns:
        display["hit_probability_pct"] = df["hit_probability"]
    if "edge_pct" in df.columns and "edge" in display.columns:
        display["edge"] = df["edge_pct"] / 100
    if "kelly_fraction" in df.columns and "kelly" in display.columns:
        display["kelly"] = df["kelly_fraction"]
    display = display.rename(
        columns={
            "pitcher_name": "Pitcher",
            "market": "Market",
            "best_side": "Side",
            "projection": "Projection",
            "line": "Line",
            "recommended_odds": "Odds",
            "fair_odds": "Fair Odds",
            "hit_probability_pct": "Hit %",
            "edge": "Edge",
            "kelly": "Kelly",
            "bookmaker": "Book",
            "team": "Team",
            "opponent": "Opp",
            "fetched_at": "Fetched At",
        }
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        display.to_excel(writer, index=False, sheet_name="Daily Board")

    workbook = load_workbook(path)
    sheet = workbook["Daily Board"]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    fills = {
        "over": PatternFill("solid", fgColor="DCFCE7"),
        "under": PatternFill("solid", fgColor="DBEAFE"),
        "high_edge": PatternFill("solid", fgColor="BBF7D0"),
        "medium_edge": PatternFill("solid", fgColor="FEF3C7"),
        "negative": PatternFill("solid", fgColor="FECACA"),
    }

    headers = {cell.value: cell.column for cell in sheet[1]}
    side_col = headers.get("Side")
    edge_col = headers.get("Edge")
    kelly_col = headers.get("Kelly")
    hit_col = headers.get("Hit %")
    projection_col = headers.get("Projection")
    line_col = headers.get("Line")

    for row in range(2, sheet.max_row + 1):
        if side_col:
            side = str(sheet.cell(row=row, column=side_col).value or "").lower()
            if side in fills:
                sheet.cell(row=row, column=side_col).fill = fills[side]
                sheet.cell(row=row, column=side_col).font = Font(bold=True)
        for col in [projection_col, line_col]:
            if col:
                sheet.cell(row=row, column=col).number_format = "0.00"
        for col in [hit_col, edge_col, kelly_col]:
            if col:
                sheet.cell(row=row, column=col).number_format = "0.00%"

    if edge_col:
        edge_letter = get_column_letter(edge_col)
        edge_range = f"{edge_letter}2:{edge_letter}{sheet.max_row}"
        sheet.conditional_formatting.add(
            edge_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["0.15"], fill=fills["high_edge"]),
        )
        sheet.conditional_formatting.add(
            edge_range,
            CellIsRule(operator="between", formula=["0.05", "0.14999"], fill=fills["medium_edge"]),
        )
        sheet.conditional_formatting.add(
            edge_range,
            CellIsRule(operator="lessThan", formula=["0"], fill=fills["negative"]),
        )

    if kelly_col:
        kelly_letter = get_column_letter(kelly_col)
        kelly_range = f"{kelly_letter}2:{kelly_letter}{sheet.max_row}"
        sheet.conditional_formatting.add(
            kelly_range,
            CellIsRule(operator="greaterThan", formula=["0"], fill=fills["medium_edge"]),
        )

    widths = {
        "A": 24,
        "B": 16,
        "C": 10,
        "D": 12,
        "E": 9,
        "F": 10,
        "G": 11,
        "H": 10,
        "I": 10,
        "J": 10,
        "K": 14,
        "L": 9,
        "M": 9,
        "N": 24,
    }
    for col_letter, width in widths.items():
        sheet.column_dimensions[col_letter].width = width

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    workbook.save(path)
    return path


def export_google_sheets(df: pd.DataFrame, sheet_name: str, worksheet_name: str = "Daily Picks") -> None:
    if not sheet_name:
        raise ValueError("Google Sheets export requested but no sheet name was configured.")
    try:
        import gspread
        from google.auth import default
    except ImportError as exc:
        raise ImportError("Install gspread and google-auth to export to Google Sheets.") from exc

    credentials, _ = default(scopes=["https://www.googleapis.com/auth/spreadsheets"])
    client = gspread.authorize(credentials)
    sheet = client.open(sheet_name)
    try:
        worksheet = sheet.worksheet(worksheet_name)
    except gspread.WorksheetNotFound:
        worksheet = sheet.add_worksheet(title=worksheet_name, rows=1000, cols=50)
    worksheet.clear()
    worksheet.update([df.columns.tolist()] + df.fillna("").astype(str).values.tolist())
